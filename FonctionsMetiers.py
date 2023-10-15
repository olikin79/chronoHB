# Auteur : Olivier Lacroix, olacroix@ac-montpellier.fr

#from ZODB import DB
#from ZODB.FileStorage import FileStorage
#from ZODB.PersistentMapping import PersistentMapping
#import persistent
#import transaction
import locale
locale.setlocale(locale.LC_TIME,'') # pour mettre en français le module time. Samedi au lieu de Saturday, etc...
import time, datetime
import os, sys, glob, subprocess
import shutil
import random
import csv, re

#import http.server
from server import *
import threading
from threading import Thread
import requests

import xlsxwriter # pour les exports excels des résultats

### A décommenter plus tard pour la mise en place des imports NG.
from openpyxl import load_workbook

from tkinter.messagebox import *

#### DEBUG
DEBUG = True

version = "1.8"

LOGDIR="logs"

def windows():
    if os.sep == "\\" :
        return True
    else :
        return False


if windows() :
    import win32ui,win32print,win32api,win32con # impression pdf sous windows.


#--------------------
# Sauvegarde, lecture, mise à jour des exercices attribués par évaluation.
#--------------------
import pickle
#import chronoHBClasses

# récupère les données de sauvegarde de à la carte en tant que variables globales pour être utilisées par les autres fonctions.
##def lire_sauvegarde(sauvegarde) :
##    if os.path.exists(sauvegarde+".db") :
##        #d =
##        retour = pickle.load(open(sauvegarde+".db","rb"))
##        #d = shelve.open(sauvegarde)
##        #retour = d['root']
##        d.close()
##    else :
##        retour = {}
##    return retour

def creerDir(path) :
    retour = True
    #print(path)
    dossier = os.path.abspath(path)
    try :
        #print("Création de", dossier, "si besoin")
        if dossier != "" : # sous windows, en cas de sauvegarde vers la racine d'un disque, basename est vide.
            os.makedirs(dossier, exist_ok=True)
    except :
        #print("Impossible de créer le dossier : la clé USB de sauvegarde n'est probablement pas branchée.")
        retour = False
    return retour

# enregistre les données de sauvegarde
def dump_sauvegarde() :
    d = open("Courses.db","wb")
    pickle.dump(root, d)
    d.close()

# récupère les données de sauvegarde
def ecrire_sauvegarde(sauvegarde, commentaire="", surCle=False, avecVideos=False) :
    #global noSauvegarde
    #print("sauvegarde", sauvegarde+".db", "noSauvegarde:", noSauvegarde)
    #d = shelve.open(sauvegarde)
    #creerDir(sauvegarde)
    if avecVideos :
        destination = sauvegarde
        sauvegarde = os.path.basename(sauvegarde) # dans ce cas sauvegarde est un dossier et non un fichier. La flemme de refaire plus propre.
    else :
        if surCle :
            # ajout d'une sauvegarde sur clé très régulière
            destination = Parametres["cheminSauvegardeUSB"]
            try :
                creerDir(destination)
            except :
                if os.sep == "/" :
                    print("Impossible de créer le dossier fixé en paramètre ", destination)
                else :
                    print("Le lecteur", destination[:3] ,"n'existe pas")
        else :
            destination = "db"
    date = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())
    nomFichierCopie = destination + os.sep + sauvegarde+"_"+ date + commentaire
    dump_sauvegarde()
    if os.path.exists("Courses.db") :
        if destination != "" and creerDir(destination) :
            print("Création de la sauvegarde", nomFichierCopie)
            if os.path.exists("Courses.db") :
                shutil.copy2("Courses.db",  nomFichierCopie + ".db")
            if os.path.exists("donneesModifLocale.txt"):
                shutil.copy2("donneesModifLocale.txt", nomFichierCopie + "_ML.txt")
            else :
                print("Pas de fichier de modifications locales, on place un fichier vide dans la sauvegarde pour assurer une cohérence.")
                open("donneesModifLocale.txt", 'a').close()
            if os.path.exists("donneesSmartphone.txt"):
                shutil.copy2("donneesSmartphone.txt", nomFichierCopie + "_DS.txt")
            else :
                print("Pas de fichier de données provenant des smartphones, on place un fichier vide dans la sauvegarde pour assurer une cohérence.")
                open("donneesSmartphone.txt", 'a').close()
            creerDir("videos")
            if avecVideos or surCle :
                creerDir(destination + os.sep + "chronoHBvideos")
                files = glob.glob("videos/*.avi")
                for file in files :
                    dest = destination + os.sep + "chronoHBvideos" + os.sep + os.path.basename(file)
                    if not os.path.exists(dest) and time.time() - os.path.getmtime(dest) > 15 :
                        # on copie les fichiers vidéos qui n'existent pas et qui ne sont pas en cours de création : ils ont plus de 15 secondes.
                        shutil.copy2(file, dest)
            #if avecVideos and os.path.exists("videos") : # par défaut, on ne sauvegardait pas les vidéos. Seulement à vocation d'archivage.
            # désormais, on sauvegarde snas overwrite pour limiter les les flux
                #shutil.copytree("videos", "chronoHBvideos")
        elif destination != "" :
            print("Pas de SAUVEGARDE CREE : chemin spécifié incorrect (" +destination+")")
            nomFichierCopie = "Pas de SAUVEGARDE CREEE : chemin spécifié incorrect : " +destination
        else :
            nomFichierCopie = "Pas de SAUVEGARDE CREEE : paramètre spécifié vide"
##        while os.path.exists(sauvegarde+"_"+ str(noSauvegarde)+".db"):
##            noSauvegarde += 1
##        print("Sauvegarde vers", sauvegarde+"_"+ str(noSauvegarde) +".db")
##        shutil.copy2(sauvegarde+".db", sauvegarde+"_"+ str(noSauvegarde) +".db")
        return nomFichierCopie


# enregistre les données de sauvegarde
# récupère les données de sauvegarde
def recupere_sauvegarde(sauvegardeChoisie) :
    global sauvegarde
    #nomFichier = os.path.basename(sauvegardeChoisie)[:-3]
    #rep = os.path.dirname(sauvegardeChoisie)
    #fichierDonnees = sauvegardeChoisie
    #print("Sauvegarde choisie",sauvegardeChoisie,"Fichier:",nomFichier,"Dossier",rep)
    fichierML = sauvegardeChoisie[:-3] + "_ML.txt"
    fichierDS = sauvegardeChoisie[:-3] + "_DS.txt"
    dossierVideos = os.path.dirname(sauvegardeChoisie) + os.sep + "chronoHBvideos"
    tousPresents = True
    ### tester si les trois fichiers existent.
    for fichier in [sauvegardeChoisie ,fichierML , fichierDS] :
        if not os.path.exists(fichier) :
            #print("Fichier",fichier,"absent")
            tousPresents = False
            break
    ### avertir sinon
    if not tousPresents :
        message = "Le fichier " + fichier + " est absent. La sauvegarde est incomplète. Import annulé."
        print(message)
        showinfo("ERREUR",message)
    else :
        ### sauvegarder les données actuelles de façon automatique avec ecrire_sauvegarde(...)
        ecrire_sauvegarde(sauvegarde, "-avant-import-autres-donnees",surCle=False)
        ### copier les trois fichiers : celui db à la place de l'ancien + 2 fichiers textes finissant par ML et DS
        shutil.copy2(sauvegardeChoisie,  sauvegarde+".db")
        shutil.copy2(fichierML, "donneesModifLocale.txt")
        shutil.copy2(fichierDS, "donneesSmartphone.txt")
        # restauration des vidéos sauvegardées
        if os.path.exists("videos") :
            shutil.rmtree("videos")
        if os.path.exists(dossierVideos) :
            shutil.copytree(dossierVideos,"videos")
        ### restaurer la base de données avec chargerDonnees() afin de charger les données en mémoire.
        retour = chargerDonnees()
        setParametres() # fichier à destination du smartphone à regéréner.
        return retour


##    d = open(sauvegarde+".db","wb")
##    pickle.dump(root, d)
##    #d['root'] = root
##    d.close()
##    if os.path.exists(sauvegarde+".db") :
##        if surCle :
##            # ajout d'une sauvegarde sur clé très régulière
##            destination = Parametres["cheminSauvegardeUSB"]
##        else :
##            destination = "db"
##        if destination != "" and creerDir(destination) :
##            date = time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())
##            nomFichierCopie = destination + os.sep + sauvegarde+"_"+ date + commentaire + ".db"
##            if not os.path.exists(nomFichierCopie) :
##                shutil.copy2(sauvegarde+".db",  nomFichierCopie)
##                print("La sauvegarde", nomFichierCopie, "est créée.")
##                os.path.copy("donneesModifLocale.txt", destination + os.sep + sauvegarde +"_"+ date + commentaire + "_ML.txt")
##                os.path.copy("donneesSmartphone.txt", destination + os.sep + sauvegarde +"_"+ date + commentaire + "_DS.txt")
##            else :
##                print("La sauvegarde", nomFichierCopie, "existe déjà. Elle n'est pas remplacée pour éviter tout risque.")
##        elif destination != "" :
##            print("Pas de SAUVEGARDE CREE : chemin spécifié incorrect (" +destination+")")
##            nomFichierCopie = "Pas de SAUVEGARDE CREE : chemin spécifié incorrect : " +destination
##        else :
##            nomFichierCopie = "Pas de SAUVEGARDE CREE : paramètre spécifié vide"
####        while os.path.exists(sauvegarde+"_"+ str(noSauvegarde)+".db"):
####            noSauvegarde += 1
####        print("Sauvegarde vers", sauvegarde+"_"+ str(noSauvegarde) +".db")
####        shutil.copy2(sauvegarde+".db", sauvegarde+"_"+ str(noSauvegarde) +".db")
##        return nomFichierCopie

def incrementeDecompteParCategoriesDAgeEtRetourneSonRang(catFFA , DecompteParCategoriesDAge, sexe, evolution = 1) :
    # en théorie inutile puisque toutes les catégories sont présentes. rangDansCategorie = 0
    for L in DecompteParCategoriesDAge : # on fait le test pour les petites catégories puis pour les vétérans.
        # on considère que les séniors doivent être meilleurs que toutes les autres catégories (au dessus et en dessous).
        ## tester d'abord si la catégorie existe dans la liste.
        present = False
        for el in L :
            if el[0] == catFFA:
                present = True
                break
        ## si oui, faire le parcours suivant.
        ## sinon, ne rien faire.
        if present :
            for couple in L :
                if sexe == "F" :
                    couple[2] += evolution
                else :
                    couple[1] += evolution
                if catFFA == couple[0] :
                    if sexe == "F" :
                        rangDansCategorie = couple[2]
                    else :
                        rangDansCategorie = couple[1]
                    break
            # pour éviter de compter deux fois la catégorie sénior (et pour les autres catégories, de faire un 2ème parcours inutile.
            break
    return rangDansCategorie

class DictionnaireDeCoureurs(dict) :
    def __init__(self, AncienneListeAImporter=[]):
        super().__init__()
        self.nombreDeCoureurs = len(AncienneListeAImporter)
        self.importerAncienneListe(AncienneListeAImporter)
        self["CoureursElimines"] = {"A" : []}
        self.initEffectifs() # initialisation pour les nouvelles bases. Méthode permettant de mettre à niveau les anciennes.
    def initEffectifs(self):
        """ permet de connaître le nombre total de coureurs de chaque sexe et de chaque catégorie : pour les catégories, on considère ceux qui sont inférieurs qui doivent être battus."""
        self.nombreDeCoureursParSexe = [0,0]
        L1 = [ ["SE",0,0 ], ["ES",0,0 ], ["JU",0,0 ], ["CA",0,0 ], ["MI",0,0 ], ["BE",0,0 ], ["PO",0,0 ], ["EA",0,0 ], ["BB",0,0 ]]
        L2 = [["SE",0,0 ] , ['M0', 0,0], ['M1', 0,0], ['M2', 0,0], ['M3', 0,0], ['M4', 0,0], ['M5', 0,0], ['M6', 0,0], ['M7', 0,0], ['M8', 0,0], ['M9', 0,0], ['M10', 0,0]]
        self.nombreDeCoureursParCategorie = [L1, L2]
        for coureur in self.liste() :
            if coureur.sexe == "F" :
                self.nombreDeCoureursParSexe[1] += 1
            else :
                self.nombreDeCoureursParSexe[0] += 1
            if Parametres["CategorieDAge"] > 0 :
                incrementeDecompteParCategoriesDAgeEtRetourneSonRang(coureur.categorieFFA() , self.nombreDeCoureursParCategorie, coureur.sexe)
    def evolutionDUnAuxEffectifsTotaux(self, coureur, evolution=1):
        self.nombreDeCoureurs += evolution
        if coureur.sexe == "F" :
            self.nombreDeCoureursParSexe[1] += evolution
        else :
            self.nombreDeCoureursParSexe[0] += evolution
        if Parametres["CategorieDAge"] > 0 :
            incrementeDecompteParCategoriesDAgeEtRetourneSonRang(coureur.categorieFFA() , self.nombreDeCoureursParCategorie, coureur.sexe, evolution=evolution)
    def getTotalParCategorie(self,catFFA, sexe):
        return getDecompteParCategoriesDAgeEtRetourneTotal(catFFA , self.nombreDeCoureursParCategorie, sexe)
    def importerAncienneListe(self,AncienneListeAImporter) : # pour convertir l'ancienne liste en ce dictionnaire.
        self["A"]=AncienneListeAImporter
        # actualiser les dossards de tous les coureurs déjà présents 
        for c in self["A"] :
            doss = str(c.dossard)
            if not doss[-1].isalpha():
                c.setDossard(doss + "A")
    def recuperer(self, dossard) :
        doss = formateDossardNG(dossard)
        try : 
            if doss[-1].isalpha():
                # nouvelle implémentation des dossards
                cle = doss[-1].upper()
                c = self[cle][int(doss[:-1])-1]
            else :
                # compatibilité ascendante avec l'ancienne application et les dossards entièrement numériques / utilise l'entrée A.
                c = self["A"][int(doss)-1]
        except :
            c = Coureur("","","") # on retourne un objet Coureur mais vide.
        return c
    def liste(self) : ##### On élimine du retour les indices libres présents dans self["CoureursElimines"]
        L = []
        for e in self.cles() :
            #print(self["CoureursElimines"][e])
            i = 0
            indicesLibres = self["CoureursElimines"][e]
            for c in self[e] :
                if not i in indicesLibres :
                    L.append(c)
                    #print("Coureur non effacé listé",c.nom, c.prenom, c.dossard)
                i += 1
        return L
    def effacerTout(self) :
        self.clear()
        for file in glob.glob("Coureurs?.txt") :
            print("Suppression du fichier", file)
            os.remove(file)
        self.__init__()
    def ajouter(self, coureur, course="", dossard = ""): # course serait une lettre "A", "B", "C", ...
        # ajout dans le premier coureur vide de la course.
        # retourne le dossard affecté
        if dossard : 
            course = dossard[-1]
            num = int(dossard[:-1])
        if self.existe(coureur) == "" :
            if not course in self.keys() :
                self[course] = []
                self["CoureursElimines"][course]=[]
            if dossard : # le dossard est spécifié, on impose le numéro.
                coureur.setDossard(dossard)
                nbreCoureursDansCourse = len(self[course])
                while nbreCoureursDansCourse < num - 1 : # on rajoute des coureurs vides via self["CoureursElimines"][course]
                    self[course].append(Coureur("","","",""))
                    self["CoureursElimines"][course].append(nbreCoureursDansCourse)
                    nbreCoureursDansCourse += 1
                self[course].append(coureur)
                #self.nombreDeCoureurs += 1
                self.evolutionDUnAuxEffectifsTotaux(coureur, evolution=1)
            else : # le dossard affecté n'est pas spécifié, on affecte le coureur créé au premier dossard libre.
                if self["CoureursElimines"][course] : # il est possible d'intercaler le coureur dans la liste existante suite à une suppression
                    premierIndiceLibre = self["CoureursElimines"][course].pop(0)
                    coureur.setDossard(str(premierIndiceLibre+1) + course) # on fixe le dossard du coureur
                    self[course][premierIndiceLibre]= coureur
                    #self.nombreDeCoureurs += 1
                    self.evolutionDUnAuxEffectifsTotaux(coureur, evolution=1)
                    return str(premierIndiceLibre+1) + course
                else : # aucun indice libre, on ajoute à la fin
                    coureur.setDossard(str(len(self[course])+1) + course) # on fixe le dossard du coureur
                    self[course].append(coureur)
                    #self.nombreDeCoureurs += 1
                    self.evolutionDUnAuxEffectifsTotaux(coureur, evolution=1)
                    return str(len(self[course])) + course
        else :
            print("Le coureur", coureur.nom, coureur.prenom,"existe déjà dans la base. On ne peut pas l'ajouter deux fois. Ne devrait jamais arriver.")
    def cles(self):
        L = list(self.keys())
        L.remove("CoureursElimines")
        return L
    def existe(self,element):
        """Retourne "" si le coureur n'existe pas et son dossard sinon"""
        #print("existe(self,",element,")")
        retour = ""
        if isinstance(element, Coureur):
            # c'est un coureur que l'on cherche à trouver
            for c in self.liste() :
                if c.nom.lower() == element.nom.lower() and c.prenom.lower() == element.prenom.lower() :
                    retour = c.dossard
                    break
        else :
            # c'est un dossard que l'on cherche
            doss = formateDossardNG(element)
            if doss[-1].isalpha():
                # nouvelle implémentation des dossards
                cle = doss[-1].upper()
                indice = int(doss[:-1])-1
            else :
                # compatibilité ascendante avec l'ancienne application et les dossards entièrement numériques / utilise l'entrée A.
                cle = "A"
                indice = int(doss)-1
            if cle in self.cles() and indice not in self["CoureursElimines"][cle] and indice < len(self[cle]):
                retour = self[cle][indice].dossard
        if retour :
            print("Le coureur ", element, "existe",retour)
        return retour
    def effacer(self,element) :
        try :
            if isinstance(element, Coureur):
                # c'est un coureur que l'on cherche à effacer
                for c in self.liste() :
                    if c.nom.lower() == element.nom.lower() and c.prenom.lower() == element.prenom.lower() :
                        doss = c.dossard
                        break
                self.effacer(doss)
            else :
                # c'est un dossard que l'on cherche
                doss = formateDossardNG(element)
                if doss[-1].isalpha():
                    # nouvelle implémentation des dossards
                    cle = doss[-1].upper()
                    indice = int(doss[:-1])-1
                else :
                    # compatibilité ascendante avec l'ancienne application et les dossards entièrement numériques / utilise l'entrée A.
                    cle = "A"
                    indice = int(doss)-1
                self.evolutionDUnAuxEffectifsTotaux(self.recuperer(doss), evolution=-1)
                self[cle][indice] = Coureur("","","","") # coureur vide mis à la place.
                #self.nombreDeCoureurs -= 1
                self["CoureursElimines"][cle].append(indice)
                self["CoureursElimines"][cle].sort() # on trie les indices pour que le prochain numéro réattribuée soit le plus petit.
        except :
            if isinstance(element, Coureur):
                print("Impossible d'effacer l'élément", element.nom, element.prenom, "des Coureurs actuels")
            else :
                print("Impossible d'effacer l'élément", element, "des Coureurs actuels")
    def afficher(self) :
        for course in self.cles() :
            print("Course",course)
            indicesLibres = self["CoureursElimines"][course]
            i = 0
            for coureur in self[course] :
                if not i in indicesLibres :
                    print("Indice " + str(i)+ course +":", coureur.dossard, coureur.nom, coureur.prenom, coureur.sexe)
                i += 1
    def reindexer(self,transcription) :
        """ réindexe les entrées de ce dictionnaire et tous les dossards qu'ils contiennent"""
        for ancienNom in transcription.keys() :
            if transcription[ancienNom] != ancienNom :
                self[transcription[ancienNom]] = self[ancienNom]
                del self[ancienNom]
                # on change les dossards de tous les coureurs de cette course
                for coureur in self[transcription[ancienNom]] :
                    coureur.setDossard(coureur.getDossard()[:-1] + transcription[ancienNom])
                    coureur.setCourse(transcription[ancienNom]) # si la course est définie manuellement, on impose son nouveau nom.
        # INUTILE CAR FAIT CI-AVANT nettoyage des noms vides vides dans ce dictionnaire
##        for nom in self.cles() :
##            if not self[nom] : # si c'est vide, on élimine.
##                del self[nom]



class Erreur():
    """ Une erreur de chronoHB"""
    def __init__(self, numero, courteDescription="", elementConcerne=""):
        self.numero = numero
        self.description = courteDescription
        if isinstance(elementConcerne,str) : # c'est un dossard.
            self.dossard = elementConcerne
            self.temps = 0.0
        elif isinstance(elementConcerne,float) : # c'est un temps.
            self.dossard = "0A"
            self.temps = elementConcerne
        else :
            self.dossard = "0A"
            self.temps = 0.0


class ErreursATraiter():
    """ Contient la liste des erreurs à traiter et des méthodes permettant de constater la "résolution" des erreurs précédemment constatées"""
    def __init__(self):
        self.vider()
    def add(self, err):
        self.controledesErreursPrecedentes(self, err)
        self.liste.append(err)
    def vider(self) :
        self.liste = []
    def controledesErreursPrecedentes(self) :
        # on peut contrôler que l'erreur nouvelle ne corrige pas une erreur précédente selon les cas repérés en conditions réelles
        # Exemple : pas sûr que cela soit utile. A voir
        True

### pour la partie import : les noms des classes doivent comporter deux caractères et ne pas finir par -F ou -G. => les modifier autoritairement sinon.
def naissanceValide(naissance) :
    try:
        #print("annee")
        annee = naissance[6:] # on permet les années sur 2 ou 4 chiffres. C'est datetime ci-dessous qui sera juge de la validité de la fin de chaine.
        #print("mois")
        mois = naissance[3:5]
        jour = naissance[0:2]
        correctDate = None
        newDate = datetime.datetime(int(annee),int(mois),int(jour))
        correctDate = True
        #print("La date de naissance fournie est valide :",jour,"/",mois,"/", annee)
    except :
        correctDate = False
        #print("La date de naissance fournie est INVALIDE :",naissance)
    return correctDate

def emailEstValide(email) :
    if email == "":
        return True
    else :
        regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
        return re.fullmatch(regex, str(email))

class Coureur():#persistent.Persistent):
    """Un Coureur"""
    def __init__(self, nom, prenom, sexe, dossard="", classe="", naissance="", etablissement="", etablissementNature="", absent=None, dispense=None, temps=0,\
                 commentaireArrivee="", VMA=0, aImprimer=False, scoreUNSS=1000000, course="", licence="", email="", email2=""):
        self.setDossard(dossard)
        self.nom = self.formateNomPrenom(nom)
        self.prenom = self.formateNomPrenom(prenom)
        self.setSexe(sexe)
        self.classe = str(classe)
        self.naissance = ""
        self.setNaissance(naissance) # traiter la chaine fournie et l'utiliser ou non.
        self.etablissement = etablissement
        self.etablissementNature = etablissementNature
        self.absent = bool(absent)
        self.dispense = bool(dispense)
        self.temps = float(temps)
        self.VMA = float(VMA)
        self.vitesse = 0
        self.rang = 0
        self.scoreUNSS = scoreUNSS 
        self.rangCat = 0
        self.rangSexe = 0
        self.commentaireArrivee = commentaireArrivee
        self.scoreUNSS = scoreUNSS
        self.aImprimer = aImprimer
        self.categorieAuto = True
        self.course = course
        self.setLicence(licence)
        self.email=""
        self.setEmail(email)
        self.email2=""
        self.setEmail2(email2)
        self.emailEnvoiEffectue = False
        self.emailEnvoiEffectue2 = False
        self.setEmailEnvoiEffectue(False)
        self.setEmailEnvoiEffectue2(False)
        self.emailNombreDEnvois = 0
        self.emailNombreDEnvois2 = 0
        self.tempsDerniereModif = 0
        self.categorie(CategorieDAge)
        self.__private_categorie = None
        # OBSOLETE : self.__private_categorie_manuelle = None ### devenue inutile suite à la distinction entre Catégorie et Course (version 1.7)
        self.actualiseCategorie()
    
    def reformateNomPrenom(self) :
        self.nom = self.formateNomPrenom(self.nom)
        self.prenom = self.formateNomPrenom(self.prenom)
    
    def formateNomPrenom(self, chaine) :
        chaineRetour = ""
        i = 0
        while i < len(chaine):
            if i == 0 or chaine[i - 1] in [" ","-"] :
                chaineRetour += chaine[i].upper()
            else :
                chaineRetour += chaine[i].lower()
            i += 1
        return chaineRetour
        
    def actualiseCategorie(self) :
        self.__private_categorie = None
        return self.categorie(CategorieDAge)



    def categorie(self, CategorieDAge=False):
        try : # compatibilité avec les vieilles sauvegardes restaurées
            self.etablissement
        except:
            self.etablissement = ""
            self.etablissementNature = ""
        try : # compatibilité ascendante avec les anciennes sauvegardes
            self.licence
        except :
            self.setLicence("")
        try :
            self.course
        except :
            self.course = ""
        #if not Parametres["CoursesManuelles"] :
        if self.__private_categorie == None :
            if CategorieDAge > 0 :
                if len(self.naissance) != 0 :
                    anneeNaissance = self.naissance[6:]
                    if CategorieDAge == 2 : ## UNSS
                         ### La catégorie d'athlétisme est utilisée sauf pour les élèvesà la limite entre collège et lycée
                         ###(un 3ème ayant redoublé est cadet : il coure en minimes / un minime en lycée ayant sauté une classe coure avec les cadets.)
                        cat = categorieAthletisme(anneeNaissance, etablissementNature = self.etablissementNature)
                        if self.etablissementNature == "CLG" and cat == "CA" : # le cadet a redoublé
                            cat = "MI"
                        elif self.etablissementNature and self.etablissementNature[0] == "L" and cat == "MI" : # le minime a sauté une classe.
                            cat = "CA"
                        self.__private_categorie = cat + "-" + self.sexe
                    else: ## catégories FFA
                        #print("calcul des catégories poussines, benjamins, junior, ... en fonction de la date de naissance codé. TESTE OK")
                        self.__private_categorie = categorieAthletisme(anneeNaissance) + "-" + self.sexe
            else :  ## catégories pour le cross du collège : initiale de la classe + "-" + sexe
                if len(self.classe) != 0 :
                    self.__private_categorie = self.classe[0] + "-" + self.sexe
        if not CoursesManuelles :  ### désormais, les catégories ne sont plus assimilées aux courses systématiquement mais seulement en mode non manuel.
            self.course = self.__private_categorie
        return self.__private_categorie

    def categorieSansSexe(self) :
        return self.categorie(Parametres["CategorieDAge"])[:2]
    def setLicence(self,licence):
        self.licence = str(licence)
        try :
            self.etablissementNoUNSS  = self.licence[:7]
        except :
            self.etablissementNoUNSS = ""
    def setEmail(self,email):
        if email :
            if emailEstValide(email) and self.email != email : # si valide et différent de l'actuel, on remplace l'email existant, sinon, on ne remplace pas la valeur actuelle.
                self.email = str(email)
                self.setEmailEnvoiEffectue(False)
                self.emailNombreDEnvois = 0
        else :
            self.email = ""
            self.setEmailEnvoiEffectue(False)
            self.emailNombreDEnvois = 0
    def setEmail2(self,email2):
        if email2 :
            if emailEstValide(email2) and self.email2 != email2 : # si valide et différent de l'actuel, on remplace l'email existant, sinon, on ne remplace pas la valeur actuelle.
                self.email2 = str(email2)
                self.setEmailEnvoiEffectue2(False)
                self.emailNombreDEnvois2 = 0
        else :
            self.email2 = ""
            self.setEmailEnvoiEffectue(False)
            self.emailNombreDEnvois2 = 0
    def setEmailEnvoiEffectue(self, val = True) :
        if self.dossard :
            self.emailEnvoiEffectue = bool(val)
            # print("emailEnvoiEffectue pour", self.nom, self.prenom, self.dossard, ":", self.emailEnvoiEffectue)
        # compatbilité ascendante avec vieilles sauvegardes
        try : 
            self.email
        except :
            self.email = ""
        try :
            self.emailNombreDEnvois
            if val :
                self.emailNombreDEnvois += 1 
        except : # cas d'import de vieilles sauvegardes n'ayant pas cette propriété.
            if bool(val) : # initialisation correcte en fonction de l'action demandée.
                self.emailNombreDEnvois = 1
            else :
                self.emailNombreDEnvois = 0
    def setEmailEnvoiEffectue2(self, val = True) :
        if self.dossard :
            self.emailEnvoiEffectue2 = bool(val)
            # print("emailEnvoiEffectue2 pour", self.nom, self.prenom, self.dossard, ":", self.emailEnvoiEffectue2)
        # compatbilité ascendante avec vieilles sauvegardes
        try : 
            self.email2
        except :
            self.email2 = ""
        try :
            self.emailNombreDEnvois2
            if val :
                self.emailNombreDEnvois2 += 1 
        except : # cas d'import de vieilles sauvegardes n'ayant pas cette propriété.
            if bool(val) : # initialisation correcte en fonction de l'action demandée.
                self.emailNombreDEnvois2 = 1
            else :
                self.emailNombreDEnvois2 = 0
    def categorieFFA(self) :
        return categorieAthletisme(self.naissance[6:])
    def scoreUNSSFormate(self) :
        if int(self.scoreUNSS) == self.scoreUNSS : # scoreUNSS est un entier au type float.
            retour = str(int(self.scoreUNSS))
        else :
            retour = str(round(self.scoreUNSS,1)).replace(".",",")
        return retour
    def nombreDeSecondesDepuisDerniereModif(self) :
        try :
            retour = int(time.time()-self.tempsDerniereModif)
        except : 
            self.tempsDerniereModif = time.time()
            retour = 0 # si pas d'heure de dernière modif, on l'initialise
        return retour
    def setCourse(self, c) :
        if CoursesManuelles and self.course != c : # si la course change, on renvoie l'email. Sinon, on ne fait rien.
            self.course = c
            self.setEmailEnvoiEffectue(False)
            self.setEmailEnvoiEffectue2(False)
            self.tempsDerniereModif = time.time()
        else :
            print("Mode courses automatiques : aucune actualisation de la course pour le coureur", self.nom, "vers le nom de course", c)
    def setScoreUNSS(self, nbreArriveesGroupement) :
        try :
            if self.etablissementNature and self.etablissementNature[0].upper() == "L" : # si la nature de l'établissement est non vide et si c'est un lycée
                self.scoreUNSS = self.rang * 100 / nbreArriveesGroupement
            else :
                self.scoreUNSS = self.rang
            self.nbreArriveesGroupement = nbreArriveesGroupement
        except :
            self.scoreUNSS = 1000000 # pour éviter qu'une erreur (non prévue) place le coureur dans les premiers.
            self.nbreArriveesGroupement = 0
##    def setCategorie(self, nouveauNom, CategorieDAge=False):
##        try :
##            self.__private_categorie_manuelle = str(nouveauNom)
##            self.categorieAuto = False
##        except:
##            print("nom de catégorie incorrect:", nouveauNom)
##    def setCategorieAuto(self):
##        self.categorieAuto = True
    def setEtablissement(self,etab, nature):
        print("on fixe l'établissement pour", self.nom, ":",etab)
        self.etablissement = etab
        self.etablissementNature = "CLG"
        if nature == "LG" or nature == "LP" :
            self.etablissementNature = nature
    def setDossard(self, dossard) :
        try :
            if dossard != "" :
                doss = formateDossardNG(dossard)
                if doss[-1].isalpha() :
                    int(doss[:-1])
                    self.dossard = doss # le dossard doit impérativement être au format "1A", "31B", etc...
                else :
                    int(doss)
                    self.dossard = doss + "A" # si rien n'est spécifié, c'est la course A.
            else :
                self.dossard = ""
        except :
            self.dossard = ""
    def getDossard(self, avecLettre=True) :
        if avecLettre :
            retour = self.dossard
        else :
            retour = self.dossard[:-1] # on élimine la lettre pour les affichages GUI ou autre, etc...
        return retour
    def setVMA(self, VMA) :
        try :
            self.VMA = float(VMA)
        except :
            self.VMA = 0
    def setNaissance(self, naissance) :
        if naissance != "" :
            chNaissance = str(naissance)[:10] # garder uniquement les 10 premiers caractères de la chaine.
            if naissanceValide(chNaissance) :
                self.naissance = chNaissance
            else :
                chNaissance = None
##            if len(chNaissance) > 8 :
##                self.naissance = time.strptime(chNaissance, "%d/%m/%Y") # année sur 4 chiffres
##            else :
##                self.naissance = time.strptime(chNaissance, "%d/%m/%y") # année sur 2 chiffres
##        else :
##            self.naissance = None
    def setCommentaire(self, commentaire):
        self.commentaireArrivee = str(commentaire)
    def setClasse(self, classe) :
        self.classe =classe
        self.__private_categorie = None # réinit
    def setSexe(self, sexe) :
        if str(sexe).upper() == "F" :
            self.sexe = "F"
        else :
            self.sexe = "G"
        self.__private_categorie = None # réinit
    def setAbsent(self, absent) :
        if absent :
            self.absent = True
        else :
            self.absent = False
    def setDispense(self, dispense) :
        if dispense :
            self.dispense = True
        else :
            self.dispense = False
    def setTemps(self, temps=0, distance=0):
        try :
            if self.temps != float(temps) : # si le temps change, on renvoie l'email, sinon, on ne fait rien.
                self.temps = float(temps)
                self.setEmailEnvoiEffectue(False)
                self.setEmailEnvoiEffectue2(False)
                self.tempsDerniereModif = time.time()
        except :
            self.temps = 0
        if self.temps >  0:
            self.vitesse = distance *3600 / self.temps
        else :
            self.vitesse = 0
    def tempsHMS(self) :
        # réglement FFA , arrondir à la seconde supérieure.
        if self.temps - int(self.temps) == 0 : # cas rarissime où il n'y a pas de partie décimale
            secondes = int(self.temps)
        else : 
            secondes = int(self.temps) + 1
        minutes = secondes // 60
        sec = str(int(secondes % 60))
        heures = str(int(minutes // 60))
        minu = str(int(minutes % 60))
        if len(heures) == 1 :
            heures = "0" + heures
        if len(minu) == 1 :
            minu = "0" + minu
        if len(sec) == 1 :
            sec = "0" + sec
        if heures == "00" :
            retour = minu + ":" + sec
        else :
            retour = heures + ":" + minu + ":" + sec
        return retour
    def tempsFormate(self) :
        #print(self.dossard,self.nom,self.temps)
        if self.temps > 0 : # ajouté suite au coureurVide utile pour affiché le temps potentiel du prochain coureur. 
            partieDecimale = str(int(((self.temps - int(self.temps))*100)))
            if len(partieDecimale) == 1 :
                partieDecimale = "0" + partieDecimale
            #print(self.nom,self.dossard,self.temps)
            if int(time.strftime("%j",time.gmtime(self.temps))) == 1 : # pas de jour à afficher. "Premier de l'année"
                if int(time.strftime("%H",time.gmtime(self.temps))) == 0 : # pas d'heure à afficher.
                    if int(time.strftime("%M",time.gmtime(self.temps))) == 0 : # pas de minute à afficher.
                        ch = time.strftime("%S s ",time.gmtime(self.temps)) + partieDecimale + "''"
                    else :
                        ch = time.strftime("%M min %S s ",time.gmtime(self.temps)) + partieDecimale + "''"
                else :
                    # valeur par excès à la seconde retenue par la FFA
                    if self.temps - int(self.temps) == 0 :# cas rarissime où il n'y a pas de partie décimale
                        ajout = 0
                    else :
                        ajout = 1 # on ajoute une seconde au self.temps tronqué à l'unité afin d'obtenir la valeur par excès à l'unité.
                    ch = time.strftime("%H h %M min %S s ",time.gmtime(int(self.temps)+ajout)) # + partieDecimale
            else :
                ch = str(int(time.strftime("%j",time.gmtime(self.temps)))-1) + " j " + time.strftime("%H h %M min %S s ",time.gmtime(self.temps))# + partieDecimale
        else :
            ch = "Pas de temps"
        return ch
    def vitesseFormatee(self) :
        if self.vitesse >=100 :
            ch = str(int(self.vitesse)) + " km/h"
        else :
            ch = str(round(self.vitesse, 1)).replace(".",",") + " km/h"
        return ch
    def vitesseFormateeAvecVMA(self) :
        if self.VMA and self.VMA > self.vitesse :
            supplVMA = " (" + str(int(self.vitesse/self.VMA*100)) + "% VMA)"
        else :
            supplVMA = ""
        return self.vitesseFormatee() + supplVMA
    def pourcentageVMA(self) :
        if self.VMA and self.vitesse and self.VMA > self.vitesse : # pas très gênant que le % de VMA soit supérieur à 100% dans le tableur mais gênant sur la TV
            pourcVMA = str(int(self.vitesse/self.VMA*100)) + "%"
        else :
            pourcVMA = "-"
        return pourcVMA
    def vitesseFormateeAvecVMAtex(self, retourALaLigne=False) :
        retour = self.vitesseFormateeAvecVMA().replace("%","\%")
        if retourALaLigne :
            retour = retour.replace("(","\\\\").replace(")","")# sur le diplôme, on supprime les parenthèses et on met le % de VMA à la ligne.
        return retour
    def setRang(self, rang) :
        if int(rang) != self.rang :
            self.rang = int(rang)
            self.setEmailEnvoiEffectue(False)
            self.setEmailEnvoiEffectue2(False)
            self.tempsDerniereModif = time.time()
    def setRangCat(self, rang) :
        if int(rang) != self.rangCat :
            self.rangCat = int(rang)
            self.setEmailEnvoiEffectue(False)
            self.setEmailEnvoiEffectue2(False)
            self.tempsDerniereModif = time.time()
    def setRangSexe(self, rang) :
        if int(rang) != self.rangSexe :
            self.rangSexe = int(rang)
            self.setEmailEnvoiEffectue(False)
            self.setEmailEnvoiEffectue2(False)
            self.tempsDerniereModif = time.time()
    def setAImprimer(self, valeur) :
        self.aImprimer = bool(valeur)
    def setNom(self, valeur) :
        if self.nom != str(valeur) : # si la valeur change, on envoie un diplome correctif
            self.nom = str(valeur)
            self.setEmailEnvoiEffectue(False)
            self.setEmailEnvoiEffectue2(False)
            self.tempsDerniereModif = time.time()
    def setPrenom(self, valeur) :
        if self.prenom != str(valeur) :
            self.prenom = str(valeur)
            self.setEmailEnvoiEffectue(False)
            self.setEmailEnvoiEffectue2(False)
            self.tempsDerniereModif = time.time()


class Course():#persistent.Persistent):
    """Une course"""
    def __init__(self, categorie, depart=False, temps=0):
        self.categorie=categorie
        if Parametres["CategorieDAge"] :
            self.label = categorie
        else :
            self.label = categorie[0] + "ème " + categorie[2]
        self.depart=depart
        self.temps=float(temps)
        self.description = categorie
        self.groupement = None
        self.resultats = []
        self.distance = 0
        self.tempsAuto = 0
        self.nomGroupement = categorie
        self.aRegenererPourImpression = False
##        self.equipesClasses = []
    def setARegenererPourImpression (self, val):
        self.aRegenererPourImpression = bool(val)
    def reset(self) :
        print("On annule le départ de",self.categorie,".")
        self.temps = 0
        self.depart = False
    def initNomGroupement(self, cat) :
        print("initNomGroupement categorie demandée:",cat) 
        self.nomGroupement = ancienGroupementAPartirDUneCategorie(cat).nomStandard
        return self.nomGroupement
    def setNomGroupement(self, nomDonne) :
        self.nomGroupement = str(nomDonne)
    def setTemps(self, temps=0, tempsAuto=False):
        if temps == 0 :
            self.temps = time.time()
        else :
            self.temps = float(temps)
        self.depart = True
        if tempsAuto :
            self.tempsAuto = self.temps
    def setTempsHMS(self, temps="00:00:00"):
        if temps == "00:00:00" :
            self.temps = time.time()
        else :
            if HMScorrect(temps) :
                partieDec = nbreCentiemes(self.temps)/100
                chaineTps = time.strftime("%d/%m/%Y", time.gmtime(self.temps))+ " " + temps
                print("Nouveau départ fixé à",chaineTps, "pour la catégorie", self.categorie)
                self.temps = time.mktime(time.strptime(chaineTps,  "%d/%m/%Y %H:%M:%S")) + partieDec
                #print("Nouvelle valeur en mémoire :", self.temps)
            else :
                print("Heure de départ incorrecte fournie par la boite de dialogue utilisateur:" + temps+".")
        self.depart = True
    def setDescription(self, description) :
        self.description = str(description)
    def setDistance(self, distance ) :
        self.distance = float(distance)
##    def incremente(self) :
##        self.effectif += 1
##    def decremente(self) :
##        self.effectif -= 1
##    def addEquipeClasse(equipe) :
##        self.equipesClasses.append(equipe)
##    def delEquipesClasses() :
##        self.equipesClasses.clear()
    def addResultat(self, coureur) :
        self.resultats.append(coureur)
    def delResultat(self, coureur) :
        i = 0
        while i < len(self.resultats) :
            if self.resultats[i].dossard == coureur.dossard :
                del self.resultats[i]
                break
            i += 1
    def top(self):
        # on empêche la modification de la donnée si elle existe déjà pour éviter les erreurs de manipulation dans l'interface.
        # Le bouton accessible utilisera top(), la modification manuelle plus laborieuse utilisera setTemps().
        # Utiliser setTemps() semble une démarche volontaire de correction.
        if self.temps == 0 :
            self.temps = time.time()
            self.depart = True
            self.tempsAuto = self.temps
    def duree(self):
        # durée de la course depuis le début
        if self.depart :
            duree = time.time() - self.temps
        else :
            duree = 0
        return duree
    def dureeFormatee(self):
        # durée de la course depuis le début FORMATEE pour affichage
        return formaterDuree(self.duree())
    def departFormate(self, tempsAuto=False, affichageHTML=False) :
        if affichageHTML :
            return formaterTempsPourHTML(self.temps)
        elif tempsAuto :
            return formaterTempsALaSeconde(self.tempsAuto)
        else :
            return formaterTempsALaSeconde(self.temps)
    def dateFormatee(self) :
        if self.temps :
            return formaterDatePourDiplome(self.temps)
        else :
            return "Départ non donné pour course "+self.description



def HMScorrect(ch) :
    retour = False
    if len(ch)==8 and ch[2] == ":" and ch[5] == ":" :
        try :
            int(ch[0:2])
            int(ch[3:5])
            int(ch[6:])
            retour = True
        except :
            pass
    return retour

class Groupement():
    """Un groupement de courses"""
    def __init__(self, nomDuGroupement, listeDesNomsDesCourses):
        self.nom = str(nomDuGroupement)
        self.listeDesCourses = listeDesNomsDesCourses
        self.manuel = False
        self.distance = 0
        self.actualiseNom()
        self.aRegenererPourImpression = False
        self.initEffectifs() # initialisation pour les nouvelles bases. Méthode permettant de mettre à niveau les anciennes.
##        self.equipesClasses = []
    def initEffectifs(self):
        self.nombreDeCoureursGTotal = 0
        self.nombreDeCoureursFTotal = 0
        self.nombreDeCoureursTotal = 0
        L1 = [ ["SE",0,0 ], ["ES",0,0 ], ["JU",0,0 ], ["CA",0,0 ], ["MI",0,0 ], ["BE",0,0 ], ["PO",0,0 ], ["EA",0,0 ], ["BB",0,0 ]]
        L2 = [["SE",0,0 ] , ['M0', 0,0], ['M1', 0,0], ['M2', 0,0], ['M3', 0,0], ['M4', 0,0], ['M5', 0,0], ['M6', 0,0], ['M7', 0,0], ['M8', 0,0], ['M9', 0,0], ['M10', 0,0]]
        self.nombreDeCoureursParCategorie = [L1, L2]
    def evolutionDUnAuxEffectifsTotaux(self, coureur, evolution=1):
        if not coureur.absent :
            self.nombreDeCoureursTotal += evolution
            if coureur.sexe == "F" :
                self.nombreDeCoureursFTotal += evolution
            else :
                self.nombreDeCoureursGTotal += evolution
            incrementeDecompteParCategoriesDAgeEtRetourneSonRang(coureur.categorieFFA() , self.nombreDeCoureursParCategorie, coureur.sexe, evolution=evolution)
    def getTotalParCategorie(self,catFFA, sexe):
        return getDecompteParCategoriesDAgeEtRetourneTotal(catFFA, self.nombreDeCoureursParCategorie, sexe)
    def setARegenererPourImpression (self, val):
        self.aRegenererPourImpression = bool(val)
    def setNombreDeCoureursTotal(self, nbreG, nbreF) :
        self.nombreDeCoureursGTotal = int(nbreG)
        self.nombreDeCoureursFTotal = int(nbreF)
        self.nombreDeCoureursTotal = int(nbreG) + int(nbreF)
    def setNom(self, nomChoisi):
        if len(nomChoisi) > 1 : # les noms à 1 caractère sont réservés aux challenges.
            print("nom choisi:",nomChoisi)
            self.nom = str(nomChoisi)
            self.manuel = True
    def setNomStandard(self, nomChoisi):
        if CoursesManuelles : # Cette méthode ne doit pas être utilisée en dehors des coursesManuelles et du processus de réindexation.
        # les noms standards doivent être fixes dans tous les autres cas.
            print("nomStandard choisi:",nomChoisi, "pour remplacer",self.nomStandard)
            self.nomStandard = str(nomChoisi)
    def setDistance(self, distance):
        self.distance = float(distance)
        ### il faut actualiser les distances de toutes les courses du groupement. Sinon, les calculs de vitesse tombent à l'eau.
        for nomCourse in self.listeDesCourses :
            Courses[nomCourse].setDistance(distance)
            #print(nomCourse,"se voit affecté la distance", distance)
    def actualiseNom(self) :
        self.nomStandard = ""
        if self.listeDesCourses :
            self.nomStandard = self.listeDesCourses[0]
            for nomCourse in self.listeDesCourses[1:] :
                self.nomStandard = self.nomStandard + " / " + str(nomCourse)
        if not self.manuel :
            self.nom = self.nomStandard
    def setListeDesCourses(self,liste):
        if CoursesManuelles : # utilisable uniquement pour la réindexation des coursesManuelles.
            # sinon, ne pas utiliser au risque de tout casser, d'où la protection posée ici.
            self.listeDesCourses = liste
    def addCourse(self, nomCourse):
        # on ajoute le nom de la course dans le groupement
        self.listeDesCourses.append(nomCourse)
        if not self.manuel :
            self.nom = self.nom + " / " + str(nomCourse)
        self.actualiseNom()
        # on actualise les propriétés nomGroupement de toutes les courses du groupement (pour éviter des centaines de parcours de listes toutes les 2 secondes)
        self.actualiseProprieteGroupementDesCourses()

    def removeCourse(self, nomCourse):
        self.listeDesCourses.remove(nomCourse)
        self.actualiseNom()
        self.actualiseProprieteGroupementDesCourses()

    def actualiseProprieteGroupementDesCourses(self):
        for c in self.listeDesCourses :
            Courses[c].setNomGroupement(self.nomStandard)
    def affichageInfoTerminal(self) :
        print("Groupement", self.nom, "de nom standard", self.nomStandard,"contenant les courses",self.listeDesCourses,"de distance",self.distance)


class Temps():#persistent.Persistent):
    """Un tempsCoureur sur la ligne d'arrivée, éventuellement affecté à un dossard
    tempsRequeteHTTP contient le temps (nbre de s depuis 1970) correspondant à l'heure de transfert
    vers le serveur HTTP : cela permet de corriger les décalages de réglage d'heures entre
    le(s) téléphone(s) qui chronomètrent et le serveur HTTP (où ce programme python tourne)"""
    def __init__(self, tempsCoureur, tempsClient, tempsServeur):
        self.tempsCoureur = float(tempsCoureur)
        self.tempsClient = float(tempsClient)
        self.tempsServeur = float(tempsServeur)
        avanceDuTelephone = self.tempsClient - self.tempsServeur
        self.tempsReel = (round(100*(self.tempsCoureur - avanceDuTelephone)))/100 # arrondir au centième car c'est le cas dans les requêtes web et l'interface GUI
    def tempsCoureurFormate(self, HMS = True) :
        if self.tempsReel :
            ch = formaterTemps(self.tempsCoureur, HMS)#time.strftime("%H:%M:%S:",time.gmtime(self.tempsCoureur)) + partieDecimale
        else :
            ch = "-"
        return ch
    def tempsReelFormate(self, HMS = True) :
        if self.tempsReel :
##            partieDecimale = str(round(((self.tempsReel - int(self.tempsReel))*100)))
##            if len(partieDecimale) == 1 :
##                partieDecimale = "0" + partieDecimale
            ch = formaterTemps(self.tempsReel, HMS)#time.strftime("%H:%M:%S:",time.gmtime(self.tempsReel)) + str(partieDecimale)
        else :
            ch = "-"
        return ch
    def tempsReelFormateDateHeure(self, sansCentieme = False) :
        if self.tempsReel :
            partieDecimale = str(round(((self.tempsReel - int(self.tempsReel))*100)))
            if len(partieDecimale) == 1 :
                partieDecimale = "0" + partieDecimale
            ch = time.strftime("%m/%d/%y-%H:%M:%S",time.localtime(self.tempsReel))
            if not sansCentieme :
                ch += ":" + str(partieDecimale)
        else :
            ch = "-"
        return ch
    def tempsCoureurFormateDateHeure(self) :
        if self.tempsCoureur :
            partieDecimale = str(round(((self.tempsCoureur - int(self.tempsCoureur))*100)))
            if len(partieDecimale) == 1 :
                partieDecimale = "0" + partieDecimale
            ch = time.strftime("%m/%d/%y-%H:%M:%S:",time.localtime(self.tempsCoureur)) + str(partieDecimale)
        else :
            ch = "-"
        return ch
    def tempsPlusUnCentieme (self) :
        tempsARetourner = Temps(self.tempsCoureur+0.01, self.tempsClient, self.tempsServeur)
        return tempsARetourner
##        self.dossardProvisoire = None
##    def setDossard(self, dossard):
##        if dossard != None :
##            self.dossard = int(dossard)
##            self.dossardProvisoire = self.dossard
##        else :
##            self.dossard = None
##    def setDossardProvisoire(self, dossardProvisoire):
##        """ n'a pas vocation à être utilisé par l'interface graphique :
##        Sert uniquement à l'optimisation du système afin de ne pas tout recalculer
##        en cas d'ajout ou de suppression de temps"""
##        self.dossardProvisoire = dossardProvisoire
def nbreCentiemes(nombre) :
    return round(((nombre - int(nombre))*100))

def formaterTemps(tps, HMS=True) :
    partieDecimale = str(round(((tps - int(tps))*100)))
    if len(partieDecimale) == 1 :
        partieDecimale = "0" + partieDecimale
    #if int(time.strftime("%j",time.gmtime(tps))) == 1 : # pas de jour à afficher. "Premier de l'année"
    if tps <= 0 :
        ch = "-"
    else :
        if int(time.strftime("%H",time.localtime(tps))) == 0 : # pas d'heure à afficher.
            #print(time.strftime("%M",time.gmtime(tps)))
            if int(time.strftime("%M",time.localtime(tps))) == 0 : # pas de minute à afficher.
                if HMS :
                    ch = time.strftime("%S s ",time.localtime(tps)) + partieDecimale + "''"
                else :
                    ch = time.strftime("%S:",time.localtime(tps)) + partieDecimale
            else :
                if HMS :
                    ch = time.strftime("%M min %S s ",time.localtime(tps)) + partieDecimale + "''"
                else :
                    ch = time.strftime("%M:%S:",time.localtime(tps)) + partieDecimale
        else :
            if HMS :
                ch = time.strftime("%H h %M min %S s",time.localtime(tps)) # + partieDecimale
            else :
                ch = time.strftime("%H:%M:%S:",time.localtime(tps)) + partieDecimale
    ##    else :
    ##        if HMS :
    ##            ch = str(int(time.strftime("%j",time.gmtime(tps)))-1) + " j " + time.strftime("%H h %M min %S s",time.gmtime(tps))# + partieDecimale
    ##        else :
    ##            ch = str(int(time.strftime("%j",time.gmtime(tps)))-1) + " j " + time.strftime("%H:%M:%S",time.gmtime(tps))# + partieDecimale
    return ch

def formaterDatePourDiplome(tps) :
    jour = time.strftime("%A",time.localtime(tps))
    return jour[0].upper() + jour[1:] + time.strftime(" %d %B %Y",time.localtime(tps))


def formaterTempsALaSeconde(tps) :
    if int(time.strftime("%H",time.localtime(tps))) == 0 : # pas d'heure à afficher.
        if int(time.strftime("%M",time.localtime(tps))) == 0 : # pas de minute à afficher.
            ch = time.strftime("00:00:%S",time.localtime(tps))
        else :
            ch = time.strftime("00:%M:%S",time.localtime(tps))
    else :
        ch = time.strftime("%H:%M:%S",time.localtime(tps))
    return ch

def formaterTempsALaSeconde(tps) :
    if int(time.strftime("%H",time.localtime(tps))) == 0 : # pas d'heure à afficher.
        if int(time.strftime("%M",time.localtime(tps))) == 0 : # pas de minute à afficher.
            ch = time.strftime("00:00:%S",time.localtime(tps))
        else :
            ch = time.strftime("00:%M:%S",time.localtime(tps))
    else :
        ch = time.strftime("%H:%M:%S",time.localtime(tps))
    return ch

def formaterTempsPourHTML(tps) :
    #print("test",eval(time.strftime("[%Y,%m,%d,%H,%M,%S]",time.localtime(tps)).replace(",0",",")))
    return eval(time.strftime("[%Y,%m,%d,%H,%M,%S]",time.gmtime(tps)).replace(",0",",")) + [nbreCentiemes(tps)*10]

def formaterDuree(tps, HMS=True) :
    partieDecimale = str(round(((tps - int(tps))*100)))
    if len(partieDecimale) == 1 :
        partieDecimale = "0" + partieDecimale
    #if int(time.strftime("%j",time.gmtime(tps))) == 1 : # pas de jour à afficher. "Premier de l'année"
    if int(time.strftime("%H",time.gmtime(tps))) == 0 : # pas d'heure à afficher.
        #print(time.strftime("%M",time.gmtime(tps)))
        if int(time.strftime("%M",time.gmtime(tps))) == 0 : # pas de minute à afficher.
            if HMS :
                ch = time.strftime("%S s ",time.gmtime(tps))
            else :
                ch = time.strftime("00:00:%S",time.gmtime(tps))
        else :
            if HMS :
                ch = time.strftime("%M min %S s ",time.gmtime(tps))
            else :
                ch = time.strftime("00:%M:%S",time.gmtime(tps))
    else :
        if HMS :
            ch = time.strftime("%H h %M min %S s",time.gmtime(tps)) # + partieDecimale
        else :
            ch = time.strftime("%H:%M:%S",time.gmtime(tps))
##    else :
##        if HMS :
##            ch = str(int(time.strftime("%j",time.gmtime(tps)))-1) + " j " + time.strftime("%H h %M min %S s",time.gmtime(tps))# + partieDecimale
##        else :
##            ch = str(int(time.strftime("%j",time.gmtime(tps)))-1) + " j " + time.strftime("%H:%M:%S",time.gmtime(tps))# + partieDecimale
    return ch

class EquipeClasse():
    """Un objet permettant de contenir les informations pour le challenge par classe"""
    def __init__(self, nom, listeCG, listeCF, ponderation=False):
        self.nom = nom
        self.listeCG = listeCG
        self.listeCF = listeCF
        self.score = 0
        for c in listeCG + listeCF :
            self.score += c.scoreUNSS
        self.ponderation = ponderation
        if ponderation :
            self.score = self.score * Parametres["nbreDeCoureursPrisEnCompte"]*2/(len(listeCG) + len(listeCF))
        #self.scoreNonPondere = score
        #self.complet = complet
        #self.listeDesRangs = listeDesRangs
##        # alimentation des listes ci-dessus avec les n (=nbreDeCoureursPrisEnCompte) meilleurs de la classe en question.
##        while (ng < nbreDeCoureursPrisEnCompte or nf < nbreDeCoureursPrisEnCompte) and i < len(listeOrdonneeParTempsDesDossardsDeLaClasse):
##            doss = listeOrdonneeParTempsDesDossardsDeLaClasse[i]
##            coureur = listeDesCoureurs[i-1]
##            if ng < nbreDeCoureursPrisEnCompte and coureur.sexe == "G" :
##                print("le coureur",coureur.prenom,"est en rang",coureur.rang," ng=",ng)
##                self.listeCG.append(coureur)
##                self.score += coureur.rang
##                ng += 1
##            if nf < nbreDeCoureursPrisEnCompte and coureur.sexe == "F" :
##                print("le coureur",coureur.prenom,"est en rang",coureur.rang," nf=",nf)
##                self.listeCF.append(coureur)
##                self.score += coureur.rang
##                nf += 1
##            i+=1
##        # correctif si le nombre nbreDeCoureursPrisEnCompte n'est pas atteint à l'arrivée.
##        if ng + nf < 2*nbreDeCoureursPrisEnCompte :
##            print("Application d'une pondération à la classe", self.nom, "pour cause d'un nombre insuffisant de coureurs à l'arrivée :",ng + nf)
##            self.score = self.score * 2*nbreDeCoureursPrisEnCompte / (ng + nf)
##        print(nom, listeOrdonneeParTempsDesDossardsDeLaClasse, listeDesCoureurs, nbreDeCoureursPrisEnCompte)
    def scoreFormate(self) :
        if int(self.score) == self.score :
            retour = str(self.score)
        else :
            retour = str(round(self.score,1)).replace(".",",")
        return retour
    def listeDesRangs(self) :
        listeRangs = []
        for c in self.listeCG + self.listeCF :
            listeRangs.append(c.scoreUNSS) # en collège, le score est le rang du coureur / en lycée, c'est une formule pour l'UNSS : 100 * place / nbre de coureurs
        return listeRangs.sort()
    def complet(self) :
        if Parametres["CategorieDAge"] == 2 :
            if len(self.listeCG) + len(self.listeCF) >= 5 :
                return True
            else :
                return False
        else : # cas Parametres["CategorieDAge"] == 0 (cross du collège)
            if len(self.listeCG) + len(self.listeCF) >= Parametres["nbreDeCoureursPrisEnCompte"]*2 :
                return True
            else :
                return False

# setup the database
def chargerDonnees() :
    global root,Coureurs,Courses,Groupements,ArriveeTemps,ArriveeTempsAffectes,ArriveeDossards,LignesIgnoreesSmartphone,LignesIgnoreesLocal,Parametres,\
           tempsDerniereRecuperationSmartphone,ligneDerniereRecuperationSmartphone,tempsDerniereRecuperationLocale,ligneDerniereRecuperationLocale,\
           CategorieDAge,CourseCommencee,positionDansArriveeTemps,positionDansArriveeDossards,nbreDeCoureursPrisEnCompte,ponderationAcceptee,\
           calculateAll,intituleCross,lieu,messageDefaut,cheminSauvegardeUSB,vitesseDefilement,tempsPause,sauvegarde, dictUIDPrecedents, noTransmission,\
           dossardModele,webcam,webcamSensibility,ligneTableauGUI,listeAffichageTV,CoursesManuelles,nbreDossardsAGenererPourCourseManuelles, genererQRcodesPourCourseManuelles,\
           genererListingQRcodes,genererListing,diplomeModele, diplomeDiffusionApresNMin, diplomeEmailExpediteur, diplomeMdpExpediteur, diplomeDiffusionAutomatique,\
           actualisationAutomatiqueDeLAffichageTV, FTPlogin, FTPmdp, FTPserveur, email,emailMDP,emailNombreDEnvoisMax,emailNombreDEnvoisDuJour
    noSauvegarde = 1
    sauvegarde="Courses"
    if os.path.exists(sauvegarde+".db") :
        with open(sauvegarde+".db","rb") as d :
            root = pickle.load(d)
            #print("on charge les données depuis le fichier",sauvegarde+".db")
        #d = shelve.open(sauvegarde)
        #retour = d['root']
    else :
        root = {}
    #print("Sauvegarde récupérée:", root)
    # get the data, creating an empty mapping if necessary
    if not "Coureurs" in root:
        #root["Coureurs"] = persistent.list.PersistentList()
        root["Coureurs"] = DictionnaireDeCoureurs()
    Coureurs=root["Coureurs"]
    tagConvertionEnCours = False
    if isinstance(Coureurs,list) : # traitement des anciennes sauvegardes afin de convertir la liste Coureurs en un dicitonnaire
        Coureurs = DictionnaireDeCoureurs(AncienneListeAImporter=Coureurs)
        tagConvertionEnCours = True
    if not "Courses" in root :
        #print("Courses n'est pas dans root : on le crée vide.")
        root["Courses"] = {}
    Courses=root["Courses"]
    if not "Groupements" in root :
        root["Groupements"] = []
        # compatibilité ascendante pour pouvoir importer de vieilles sauvegardes (avant création des groupements).
        # on récupère les noms de toutes les catégories présentes, on regarde si chacune est bien dans un groupement existant. A défaut, on le crée.
        L = []
        for cat in Courses :
            L.append(Courses[cat].categorie)
        for cat in L :
            estPresent = False
            for grpment in root["Groupements"] :
                if cat in grpment.listeDesCourses :
                    estPresent = True
                    break
            if not estPresent:
                print("Création du groupement à partir de la catégorie", cat)
                root["Groupements"].append(Groupement(cat,[cat])) # on crée les groupements de même nom que les catégories existantes.
    Groupements=root["Groupements"]
    if not "ArriveeTemps" in root :
        root["ArriveeTemps"] = []
    ArriveeTemps=root["ArriveeTemps"]
    if not "ArriveeTempsAffectes" in root :
        root["ArriveeTempsAffectes"] = []
    ArriveeTempsAffectes=root["ArriveeTempsAffectes"]
    if not "ArriveeDossards" in root :
        root["ArriveeDossards"] = []
    ArriveeDossards=root["ArriveeDossards"]
    if tagConvertionEnCours :
        ## on convertit une et une seule fois les dossards de ArriveeDossards et de ArriveeTempsAffectes
        i = 0
        while i < len(ArriveeDossards) :
            if not str(ArriveeDossards[i])[-1].isalpha() : # sécurité si lancé plusieurs fois
                ArriveeDossards[i] = str(ArriveeDossards[i]) + "A"
            i += 1
        i = 0
        while i < len(ArriveeTempsAffectes) :
            if str(ArriveeTempsAffectes[i]) != 0 and not str(ArriveeTempsAffectes[i])[-1].isalpha() :
                ArriveeTempsAffectes[i] = str(ArriveeTempsAffectes[i]) + "A"
            else :
                ArriveeTempsAffectes[i] = str(ArriveeTempsAffectes[i])
            i += 1

    if not "LignesIgnoreesSmartphone" in root :
        root["LignesIgnoreesSmartphone"] = []
    LignesIgnoreesSmartphone=root["LignesIgnoreesSmartphone"]
    if not "LignesIgnoreesLocal" in root :
        root["LignesIgnoreesLocal"] = []
    LignesIgnoreesLocal=root["LignesIgnoreesLocal"]
    if not "dictUIDPrecedents" in root :
        root["dictUIDPrecedents"] = {}
    dictUIDPrecedents=root["dictUIDPrecedents"]
    if not "ligneTableauGUI" in root :
        root["ligneTableauGUI"] = [1,0]
    ligneTableauGUI=root["ligneTableauGUI"]

    ### paramètres par défaut
    if not "Parametres" in root :
        root["Parametres"] = {}
    Parametres=root["Parametres"]
    if not "tempsDerniereRecuperationSmartphone" in Parametres :
        Parametres["tempsDerniereRecuperationSmartphone"]=0
    tempsDerniereRecuperationSmartphone = Parametres["tempsDerniereRecuperationSmartphone"]
    if not "ligneDerniereRecuperationSmartphone" in Parametres :
        Parametres["ligneDerniereRecuperationSmartphone"]=1
    ligneDerniereRecuperationSmartphone = Parametres["ligneDerniereRecuperationSmartphone"]
    if not "tempsDerniereRecuperationLocale" in Parametres :
        Parametres["tempsDerniereRecuperationLocale"]=0
    tempsDerniereRecuperationLocale = Parametres["tempsDerniereRecuperationLocale"]
    if not "ligneDerniereRecuperationLocale" in Parametres :
        Parametres["ligneDerniereRecuperationLocale"]=1
    ligneDerniereRecuperationLocale = Parametres["ligneDerniereRecuperationLocale"]
    if not "CategorieDAge" in Parametres :
        Parametres["CategorieDAge"]=0
    CategorieDAge=Parametres["CategorieDAge"]
    if not "CourseCommencee" in Parametres :
        Parametres["CourseCommencee"]=False
    CourseCommencee=Parametres["CourseCommencee"]
    if not "positionDansArriveeTemps" in Parametres :
        Parametres["positionDansArriveeTemps"]=0
    positionDansArriveeTemps=Parametres["positionDansArriveeTemps"]
    if not "positionDansArriveeDossards" in Parametres :
        Parametres["positionDansArriveeDossards"]=0
    positionDansArriveeDossards=Parametres["positionDansArriveeDossards"]
    if not "nbreDeCoureursPrisEnCompte" in Parametres :
        Parametres["nbreDeCoureursPrisEnCompte"]=3
    nbreDeCoureursPrisEnCompte=Parametres["nbreDeCoureursPrisEnCompte"]
    if not "ponderationAcceptee" in Parametres :
        Parametres["ponderationAcceptee"]=False
    ponderationAcceptee=Parametres["ponderationAcceptee"]
    if not "calculateAll" in Parametres :
        Parametres["calculateAll"]=False
    calculateAll=Parametres["calculateAll"]
    if not "intituleCross" in Parametres :
        Parametres["intituleCross"]="Cross du collège H. Bourrillon"
    intituleCross=Parametres["intituleCross"]
    if not "lieu" in Parametres :
        Parametres["lieu"]="Stade Mirandol"
    lieu=Parametres["lieu"]
    if not "messageDefaut" in Parametres :
        Parametres["messageDefaut"]="<prenom> de <classe>. Pour éla, merci beaucoup !"
    messageDefaut=Parametres["messageDefaut"]
    if not "cheminSauvegardeUSB" in Parametres :
        Parametres["cheminSauvegardeUSB"]="N:"
    cheminSauvegardeUSB=Parametres["cheminSauvegardeUSB"]
    if not "vitesseDefilement" in Parametres :
        Parametres["vitesseDefilement"]= "3"
    vitesseDefilement=Parametres["vitesseDefilement"]
    if not "tempsPause" in Parametres :
        Parametres["tempsPause"]= "8"
    tempsPause=Parametres["tempsPause"]
    if not "dossardModele" in Parametres :
        Parametres["dossardModele"]= "cross-HB"
    dossardModele=Parametres["dossardModele"]
    if not "webcam" in Parametres :
        Parametres["webcam"]= 0
    webcam=Parametres["webcam"]
    if not "webcamSensibility" in Parametres :
        Parametres["webcamSensibility"]= 60000
    webcamSensibility=Parametres["webcamSensibility"]
    if not "listeAffichageTV" in Parametres :
        Parametres["listeAffichageTV"] = []
    listeAffichageTV=Parametres["listeAffichageTV"]
    if not "CoursesManuelles" in Parametres :
        Parametres["CoursesManuelles"] = False
    CoursesManuelles=Parametres["CoursesManuelles"]
    if not "genererQRcodesPourCourseManuelles" in Parametres :
        Parametres["genererQRcodesPourCourseManuelles"] = True
    genererQRcodesPourCourseManuelles=Parametres["genererQRcodesPourCourseManuelles"]
    if not "nbreDossardsAGenererPourCourseManuelles" in Parametres :
        Parametres["nbreDossardsAGenererPourCourseManuelles"] = 120
    nbreDossardsAGenererPourCourseManuelles=Parametres["nbreDossardsAGenererPourCourseManuelles"]
    if not "genererListingQRcodes" in Parametres :
        Parametres["genererListingQRcodes"] = False
    genererListingQRcodes=Parametres["genererListingQRcodes"]
    if not "genererListing" in Parametres :
        Parametres["genererListing"] = True
    genererListing=Parametres["genererListing"]
    if not "diplomeModele" in Parametres :
        Parametres["diplomeModele"] = "Randon-Trail"
    diplomeModele=Parametres["diplomeModele"]
    if not "diplomeDiffusionApresNMin" in Parametres :
        Parametres["diplomeDiffusionApresNMin"] = 15
    diplomeDiffusionApresNMin=Parametres["diplomeDiffusionApresNMin"]
    if not "diplomeEmailExpediteur" in Parametres :
        Parametres["diplomeEmailExpediteur"] = "lax.olivier@gmail.com"
    diplomeEmailExpediteur=Parametres["diplomeEmailExpediteur"]
    if not "diplomeMdpExpediteur" in Parametres :
        Parametres["diplomeMdpExpediteur"] = ""
    diplomeMdpExpediteur=Parametres["diplomeMdpExpediteur"]
    if not "diplomeDiffusionAutomatique" in Parametres :
        Parametres["diplomeDiffusionAutomatique"] = 0
    diplomeDiffusionAutomatique=Parametres["diplomeDiffusionAutomatique"]
    if not "actualisationAutomatiqueDeLAffichageTV" in Parametres :
        Parametres["actualisationAutomatiqueDeLAffichageTV"] = False
    actualisationAutomatiqueDeLAffichageTV=Parametres["actualisationAutomatiqueDeLAffichageTV"]
    if not "FTPlogin" in Parametres :
        Parametres["FTPlogin"] = "olivier.lax@free.fr"
    FTPlogin=Parametres["FTPlogin"]
    if not "FTPmdp" in Parametres :
        Parametres["FTPmdp"] = "mdp"
    FTPmdp=Parametres["FTPmdp"]
    if not "FTPserveur" in Parametres :
        Parametres["FTPserveur"] = "ftp://mathlacroix.free.fr"
    FTPserveur=Parametres["FTPserveur"]
    if not "FTPdir" in Parametres :
        Parametres["FTPdir"] = "/public_html/chronoHB"
    FTPdir=Parametres["FTPdir"]
    if not "email" in Parametres :
        Parametres["email"] = "chronoHB@gmail.com;chronoHB2@gmail.com"
    email=Parametres["email"]
    if not "emailMDP" in Parametres :
        Parametres["emailMDP"] = "mdp"
    emailMDP=Parametres["emailMDP"]
    if not "emailNombreDEnvoisMax" in Parametres :
        Parametres["emailNombreDEnvoisMax"] = "2"
    emailNombreDEnvoisMax=Parametres["emailNombreDEnvoisMax"]
    if not "emailNombreDEnvoisDuJour" in Parametres :
        Parametres["emailNombreDEnvoisDuJour"] = {}
    emailNombreDEnvoisDuJour=Parametres["emailNombreDEnvoisDuJour"]
    ##transaction.commit()
    return globals()
    
chargerDonnees()

if os.name=="posix" :
    sep="/"
    compilateur = "/Library/TeX/texbin/pdflatex"
else :
    sep="\\"
    compilateur = 'start "" /I /wait /min /D .\\@dossier@\\tex .\\texlive\\2020\\bin\\win32\\pdflatex.exe -synctex=1 -no-shell-escape -interaction=nonstopmode -output-directory=.. '#

    ### commande fonctionnelle mais pas propre car le logo doit se trouver dans la distribution, dans win32
    ## start "" /I /wait /min /D ".\dossards\tex\" ".\texlive\2020\bin\win32\pdflatex.exe" -synctex=1 -interaction=nonstopmode -output-directory=".." "3-F.tex"

## SERVEUR en daemon
##import http.server
##import threading
##
##PORT = 8888
##server_address = ("", PORT)



##class TableauGUI() :
##    def __init__(self):
##        self.largeurCaracteres = 8
##        self.lignes = []
##        self.lignesActualisees = []
##        self.tempsReferences = []
##        self.indiceLignesDejaAffichees = 0
##        self.largeursColonnes = []
##        titres = ["No","Tps Téléphone","Doss.","Arrivée","Nom","Prénom","Dossard","Chrono","Cat.","Rang","Vit."]
##        self.donneesEditables = ["Tps Téléphone","Doss.","Arrivée"]
##        for t in titres :
##            if "no" in t.lower() :
##                self.largeursColonnes.append(len(t)*(self.largeurCaracteres+15))
##            else :
##                self.largeursColonnes.append(len(t)*self.largeurCaracteres)
##        #pour des colonnes de largeur nulle, je tentais d'ajouter le titre après. ECHEC titres[1] = "Heure Mesurée"
##        self.titres = tuple(titres)
##    def append(self, coureur, temps, dossardAffecte) :#, tempsSupplementaire = False):
##        #print(len(self.lignes)+1)
##        if len(coureur.nom)>=2 :
##            nom = str(coureur.nom)[0].upper()+ str(coureur.nom)[1:].lower()
##        else :
##            nom = str(coureur.nom).upper()
##        if not dossardAffecte :
##            dossardAffecte = "-"
##        if coureur.rang :
##            rang = coureur.rang
##        else :
##            rang = "-"
##        if coureur.vitesse :
##            vitesse = coureur.vitesseFormatee()
##        else :
##            vitesse = "-"
##        categorie = coureur.categorie(Parametres["CategorieDAge"])
##        if categorie == None :
##            categorie = "-"
##        if coureur.temps :
##            tempsDuCoureur = coureur.tempsFormate()
##        else :
##            tempsDuCoureur = "-"
##        if coureur.dossard != 0 :
##            dossard = coureur.dossard
##        else :
##            dossard = "-"
##        try :
##            index = self.tempsReferences.index(temps.tempsCoureur)
##            # cas des coureurs qui ont le même temps affecté à cause d'un décalage du nombre de saisies entre les deux smartphones.
##            trouve = False
##            while self.tempsReferences[index] == temps.tempsCoureur and not trouve :
##                print("dossard", self.lignes[index][6])
##                if self.lignes[index][6] == dossard : ## VERIFIER que le dossard est en position 6
##                    trouve = True
##                index += 1
##            if trouve :
##                index -= 1
##            else :
##                index = -1
##            print(trouve, "à l'index", index)
##        except :
##            index = -1
##        ligneAAjouter = []
##        if index >= 0 :
##            # mise à jour d'une ligne si ce n'est pas un temps en rab non encore affecté
##            ### if not tempsSupplementaire : # ATTENTION: si décommenté, les temps supplémentaires ne se mettront pas à jour indéfinement tant qu'un dossard ne leur sera pas affecté.
##            ligneAAjouter = [index+1, temps.tempsCoureurFormate(), dossardAffecte, temps.tempsReelFormate() , nom, coureur.prenom, dossard, tempsDuCoureur ,categorie, rang, vitesse]
##            self.lignes[index]= ligneAAjouter
##            print("mise à jour de la ligne", ligneAAjouter)
##            self.alimenteLignesActualisees(index)
##        else :
##            # ajout d'une ligne AU BON ENDROIT
##            index = len(self.lignes)
##            self.indiceLignesDejaAffichees = index
##            self.alimenteLignesActualisees(index)
##            ligneAAjouter = [index+1 , temps.tempsCoureurFormate(), dossardAffecte, temps.tempsReelFormate() , nom, coureur.prenom, coureur.dossard, coureur.tempsFormate(),categorie, rang, vitesse]
##            self.lignes.append(ligneAAjouter)
##            print("ajout de la ligne", ligneAAjouter)
##            self.tempsReferences.append(temps.tempsCoureur)
##        print("lignes actualisées", self.lignesActualisees)
##        # actualisation automatique de la largeur des colonnes : pertinent ?
##        i = 0
##        for element in ligneAAjouter :
##            if "no" in str(element).lower() :
##                newWidth = len(str(element))*(self.largeurCaracteres+15)
##            else :
##                newWidth = len(str(element))*self.largeurCaracteres
##            if newWidth > self.largeursColonnes[i] :
##                self.largeursColonnes[i] = newWidth
##            i += 1
##    def alimenteLignesActualisees(self, index):
##        if not index in self.lignesActualisees :
##            self.lignesActualisees.append(index)
##    def getlines(self) :
##        retour = []
##        retourIndices = list(self.lignesActualisees)
##        for i in self.lignesActualisees :
##            retour.append(self.lignes[i])
##        self.effaceLignesActualisees()
##        return retourIndices, retour
##    def effaceLignesActualisees(self) :
##        self.lignesActualisees.clear()
##    def reinit(self):
##        self.__init__()
####        print("lignes", self.lignes, "lignesAffichees", self.lignesDejaAffichees)
####        try :
####            while True :
####                self.lignesDejaAffichees.append(self.lignes[0])
####                retour.append(self.lignes[0])
####                del self.lignes[0]
####        except :
####            self.lignes = []
####            print("lignes", self.lignes, "lignesAffichees", self.lignesDejaAffichees)
####        self.lignesDejaAffichees = self.lignesDejaAffichees + self.lignes
####        retour = self.lignes
####        self.lignes.clear()
####        print("retour:", retour, "lignes", self.lignes)
####        return retour
####    def enTetes(self) :
####        return self.titres
####    def lignes(self) :
####        return self.donnees
####    def largeursColonnes(self) :
####        return largeursColonnesPourTreeViewEnFonctionDesDonnees # on peut compter 8 pixels par caractère.


#### export excel des résultats


def exportXLSX():
    fichier = 'impressions' + sep + '_resultats.xlsx'
    if os.path.exists(fichier) :
        os.remove(fichier)
    workbook = xlsxwriter.Workbook(fichier)
    worksheet = workbook.add_worksheet()
    ### constitution des champs à ajouter et de leurs contenus
    if CategorieDAge :
        CourseOuClasse = 'Course'
    else :
        CourseOuClasse = 'Classe'
    exportChampsOrdonnes = ['Dossard', 'Nom', 'Prénom', 'Sexe', 'Naissance', 'Catégorie FFA',\
                            CourseOuClasse, 'VMA', 'absent', 'dispense',\
                            'Rang', 'Temps (en s)', 'Temps (HMS)',\
                            'Vitesse (en km/h)', 'Pourcentage de VMA']
    exportProprietesOrdonnees = ['coureur.dossard', 'coureur.nom', 'coureur.prenom', 'coureur.sexe', 'coureur.naissance', 'coureur.categorieFFA()',\
                                 'groupementAPartirDeSonNom(coureur.course, nomStandard=True).nom if CategorieDAge else coureur.classe', 'coureur.VMA', '"oui" if coureur.absent else ""', '"oui" if coureur.dispense else ""',\
                                 'coureur.rang if coureur.rang else "-"','round(coureur.temps,2)', 'coureur.tempsHMS()',\
                                 'round(coureur.vitesse,1) if coureur.vitesse else "-"','coureur.pourcentageVMA()']
    if CategorieDAge == 2 : # cross UNSS, champs supplémentaires
        exportChampsOrdonnes += ['établissement', 'type','Licence',]
        exportProprietesOrdonnees += ['coureur.etablissement', 'coureur.etablissementNature', 'coureur.licence']
    ### remplissage 1ère ligne
    colonnesTempsFormate= []
    j = 0
    for champ in exportChampsOrdonnes :
        worksheet.write(chr(65+j) + "1", champ)
        if "HMS" in champ :
            colonnesTempsFormate.append(chr(65+j))
        j += 1
    i = 0
    ### remplissage du tableur avec les propriétés, coureur par coureur (ligne par ligne)
    L = Coureurs.liste()
    while i < len(L) :
        coureur = L[i]
        ligne = i + 2
        j = 0
        for propriete in exportProprietesOrdonnees :
            worksheet.write(chr(65+j)+ str(ligne), eval(propriete))
            j += 1
        i += 1
    ### formatage HMS
    fmt = workbook.add_format({'num_format':'hh:mm:ss'})
    for col in colonnesTempsFormate :
        worksheet.set_column(col + ':' + col, None, fmt)
    workbook.close()
    ### ouverture immédiate du tableur avec le logiciel par défaut sur l'ordinateur.
    path = os.getcwd()
    fichierAOuvrir = path + os.sep + fichier
    subprocess.Popen([fichierAOuvrir],shell=True)
    #subprocess.Popen(r'explorer /select,"' + )



#############################################################""

global TableauGUI
tableauGUI = []
### ajouté à la base données : initialisé avant dans la fonction dédiée.
### ligneTableauGUI = [1,0] # [noligne du tableau, noligneAStabiliser en deça ne pas actualiser la prochiane fois]

def reinitTableauGUI () :
    global tableauGUI
    tableauGUI.clear()

def alimenteTableauGUI (tableauGUI, coureur, temps, dossardAffecte, ligneAjoutee, derniereLigneStabilisee ):
    """ modifie le tableau tableauGUI des lignes à actualiser dans l'interface graphique """
    global ligneTableauGUI
    if derniereLigneStabilisee < ligneAjoutee :
##    if ligneAjoutee < len(tableauGUI) :
##        tableauGUI[ligneAjoutee - 1] = formateLigneGUI(coureur, temps, dossardAffecte)
##    else :
        tableauGUI.append(formateLigneGUI(coureur, temps, dossardAffecte, ligneAjoutee))
    ligneTableauGUI[0] = ligneTableauGUI[0]+1

def formateLigneGUI(coureur, temps, dossardAffecte, ligneAjoutee):
    #print(len(self.lignes)+1)
    if dossardAffecte == "0A" or dossardAffecte == "0" :
        dossardAffecte = "-"
    if coureur.rang :
        rang = coureur.rang
    else :
        rang = "-"
    if coureur.vitesse :
        vitesse = coureur.vitesseFormateeAvecVMA()
    else :
        vitesse = "-"
    categorie = coureur.categorie(Parametres["CategorieDAge"])
    if categorie == None :
        categorie = "-"
    if coureur.temps :
        if coureur.temps == -1 :
            tempsDuCoureur = "Course non partie"
        else :
            tempsDuCoureur = coureur.tempsFormate()
    else :
        if coureur.nom == "" :
            tempsDuCoureur = "-"
        else :
            tempsDuCoureur = "Nég. ou nul."
    if coureur.dossard != "0" :
        dossard = coureur.dossard
    else :
        dossard = "-"
    #Noligne = formateNoLigne(ligneAjoutee)
    #return [ligneAjoutee , temps.tempsCoureurFormate(), dossardAffecte, temps.tempsReelFormate() , nom, coureur.prenom, coureur.dossard, coureur.tempsFormate(),categorie, rang, vitesse]
    return [ligneAjoutee , temps, dossardAffecte, coureur.nom, coureur.prenom, coureur.dossard, coureur.classe, tempsDuCoureur,categorie, rang, vitesse]
##        self.lignes.append(ligneAAjouter)
##        print("ajout de la ligne", ligneAAjouter)
##        self.tempsReferences.append(temps.tempsCoureur)
##    print("lignes actualisées", self.lignesActualisees)
##    # actualisation automatique de la largeur des colonnes : pertinent ?
##    i = 0
##    for element in ligneAAjouter :
##        if "no" in str(element).lower() :
##            newWidth = len(str(element))*(self.largeurCaracteres+15)
##        else :
##            newWidth = len(str(element))*self.largeurCaracteres
##        if newWidth > self.largeursColonnes[i] :
##            self.largeursColonnes[i] = newWidth
##        i += 1
##    return [ ind , newlines ]




Resultats = {} # dictionnaire des résultats calculés qui sera regénéré à chaque lancement en fonction des données de root
ResultatsGroupements = {} # dictionnaire des résultats calculés qui sera regénéré à chaque lancement en fonction des données de root
# chaque entrée est :
# 1. un nom de course (key = "6-F", ...) : la valeur est une liste de coureurs dans l'ordre
# 2. un nom de classe (key = "65",...) : la valeur est une liste de coureurs dans l'ordre (à la fin, on aura ajouté les abandons, absents, dispensés.
# 3. un challenge par classe (key = "6", "5", ) : la valeur est une EquipeClasse (class à définir dont les médthoes permettront le calcul du barème)

##DonneesAAfficher = TableauGUI()
coureurVide = Coureur("", "", "")

### traitement des fichiers créés par le serveur web. On y accède en lecture et on ne l'efface jamais sauf si on reinitialise la course.
def traiterToutesDonnees():
    traiterDonneesSmartphone(True, False)
    traiterDonneesLocales(True,False)

def traiterDonneesSmartphone(DepuisLeDebut = False, ignorerErreurs = False):
    """Fonctionnement :  si le fichier de données smartphone a été modifié depuis le dernier traitement => agir.
        à la fin mémoriser heureDerniereRecuperationSmartphone
        retourne une liste d'instance de la class Erreur. La liste vide signifie que tout s'est bien importé
    """
    fichierDonneesSmartphone = "donneesSmartphone.txt"
    #print("Import depuis de le début :", DepuisLeDebut)
    if DepuisLeDebut :
        #root["ArriveeTemps"] = []
        #root["ArriveeTempsAffectes"] = []
        #root["ArriveeDossards"] = []
        Parametres["ligneDerniereRecuperationSmartphone"] = 1
        Parametres["tempsDerniereRecuperationSmartphone"] = 0
        Parametres["calculateAll"] = True
        dictUIDPrecedents.clear()
    retour = [] # si aucune ligne à traiter, on retourne []
    # ligneDerniereRecuperationSmartphone = Parametres["ligneDerniereRecuperationSmartphone"]
    if os.path.exists(fichierDonneesSmartphone) and derniereModifFichierDonnneesSmartphoneRecente(fichierDonneesSmartphone) :
        listeLigne = lignesAPartirDe(fichierDonneesSmartphone, Parametres["ligneDerniereRecuperationSmartphone"])
        i = 0
        #pasDErreur = True
        while i < len(listeLigne) : # plus d'arrêt à la première erreur : and pasDErreur :
            ligne = listeLigne[i]
            #print("Traitement de la ligne", Parametres["ligneDerniereRecuperationSmartphone"] , ":", ligne, end='')
            #print(ligne[-4:])
            if ligne[-4:] == "END\n" : # ligne DOIT ETRE complète (pour éviter les problèmes d'accès concurrant (le cas d'une lecture de ligne alors que l'écriture est non finie)
                codeErreur = decodeActionsRecupSmartphone(ligne, UIDPrecedents = dictUIDPrecedents)
                if codeErreur.numero :
                    # une erreur s'est produite
                    print("Code erreur :", codeErreur.numero)
                    print(ligne)
##                    if ignorerErreurs or Parametres["ligneDerniereRecuperationSmartphone"] in LignesIgnoreesSmartphone :
##                        print("Erreur ignorée")
##                        Parametres["ligneDerniereRecuperationSmartphone"] += 1
##                        Parametres["tempsDerniereRecuperationSmartphone"] = time.time()
                    #else :
                        #pasDErreur = False
##                else :
                    #print("Données importées pour la ligne :", Parametres["ligneDerniereRecuperationSmartphone"] )
                ### désormais, même s'il y a une erreur, on poursuit les imports.
                Parametres["ligneDerniereRecuperationSmartphone"] += 1
                Parametres["tempsDerniereRecuperationSmartphone"] = time.time()
                    ##transaction.commit()
            else :
                #pasDErreur = False
                print("Une ligne incomplète venant du smartphone : ne devrait pas se produire sauf en cas d'accès concurrant au fichier de données. On retente un import plus tard.")
            i += 1
            retour.append(codeErreur)
        #print("Erreurs retournées :",retour)
        if i == 0 : # si i est nul, c'est que le fichier a été parcouru en entier. Inutile de relancer de multiples sauvegardes.
            Parametres["tempsDerniereRecuperationSmartphone"] = time.time()
    else :
        Parametres["tempsDerniereRecuperationSmartphone"] = time.time()
    #    print("Fichier du smartphone déjà traité à cette heure")
##        retour = "RAS"
    return retour


    ########### copié collé de la fonction précédente en changeant juste les variables de mémorisation
def traiterDonneesLocales(DepuisLeDebut = False, ignorerErreurs = False):
    """traite le 2ème fichier de données :  celui généré par les requêtes effactuées localement directement sur l'interface GUI.
    Cela permettra de rejouer l'ensemble des actions depuis le début, puis éventuellement, plus tard, d'en annuler certaines.
    retourne une liste d'instance de la class Erreur. La liste vide signifie que tout s'est bien importé
    """
    fichierDonneesSmartphone = "donneesModifLocale.txt"
    #print("Import depuis de le début :", DepuisLeDebut)
    if DepuisLeDebut :
        #root["ArriveeTemps"] = []
        #root["ArriveeTempsAffectes"] = []
        #root["ArriveeDossards"] = []
        Parametres["ligneDerniereRecuperationLocale"] = 1
        Parametres["tempsDerniereRecuperationLocale"] = 0
        Parametres["calculateAll"] = True
        dictUIDPrecedents.clear()
    retour = []
    # ligneDerniereRecuperationSmartphone = Parametres["ligneDerniereRecuperationSmartphone"]
    if os.path.exists(fichierDonneesSmartphone) and derniereModifFichierDonnneesLocalesRecente(fichierDonneesSmartphone) :
        listeLigne = lignesAPartirDe(fichierDonneesSmartphone, Parametres["ligneDerniereRecuperationLocale"])
        i = 0
        #pasDErreur = True
        while i < len(listeLigne):# and pasDErreur :
            ligne = listeLigne[i]
            #print("Traitement de la ligne", Parametres["ligneDerniereRecuperationLocale"] , ":", ligne, end='')
            #print(ligne[-4:])
            if ligne[-4:] == "END\n" : # ligne DOIT ETRE complète (pour éviter les problèmes d'accès concurrant (le cas d'une lecture de ligne alors que l'écriture est non finie)
                codeErreur = decodeActionsRecupSmartphone(ligne, local=True, UIDPrecedents = dictUIDPrecedents)
                if codeErreur.numero :
                    # une erreur s'est produite
                    print("Code erreur : ", codeErreur.numero)
                    print(ligne)
##                    if ignorerErreurs or Parametres["ligneDerniereRecuperationLocale"] in LignesIgnoreesLocal :
##                        print("Erreur ignorée")
##                        Parametres["ligneDerniereRecuperationLocale"] += 1
##                        Parametres["tempsDerniereRecuperationLocale"] = time.time()
                    #else :
                     #   pasDErreur = False
                #else :
                    #print("Données correctement importées pour la ligne :", Parametres["ligneDerniereRecuperationLocale"] )
                # même si une erreur se produit, désormais, on poursuit les imports.
                Parametres["ligneDerniereRecuperationLocale"] += 1
                Parametres["tempsDerniereRecuperationLocale"] = time.time()
                    ##transaction.commit()
            else :
                #pasDErreur = False
                print("Une ligne incomplète venant du smartphone : ne devrait pas se produire sauf en cas d'accès concurrant au fichier de données. On retente un import plus tard.")
            i += 1
            retour.append(codeErreur)
        if i == 0 : # si i est nul, c'est que le fichier a été parcouru en entier. Inutile de relancer de multiples sauvegardes.
            Parametres["tempsDerniereRecuperationLocale"] = time.time()
    else :
        Parametres["tempsDerniereRecuperationLocale"] = time.time()
##        #print("Fichier des données locales déjà traité à cette heure")
##        retour = "RAS"
    return retour


def decodeActionsRecupSmartphone(ligne, local=False, UIDPrecedents = {}) :
    """ retourne une erreur transmise par une des fonctions mise en oeuvre ici."""
    #retour = Erreur(999) # a priori, on retourne une erreur. 10000 = erreur non répertoriée . Ne devrait pas se produire.
    listeAction = ligne.split(",")
    action = listeAction[1]
    dossard = str(listeAction[2])
    if dossard != "0" and dossard != "-1" and dossard != "0A" : # si le dossard est différent de 0 ou -1, il faudra regénérer un pdf d'une course.
        selectionnerCoursesEtGroupementsARegenererPourImpression(dossard)
    if listeAction[0] == "tps" :
        tpsCoureur = float(listeAction[3])
        tpsClient = float(listeAction[4])
        tpsServeur = float(listeAction[5])
        try :
            uid = int(listeAction[6])
        except :
            uid = 0
        try :
            noTransmission = int(listeAction[7])
        except :
            noTransmission = 0
        if uid not in UIDPrecedents :
            UIDPrecedents[uid]=[]
        #print("uid transmission",uid, "no ", noTransmission, "UIDPRECEDENTS", UIDPrecedents)
        if not noTransmission in UIDPrecedents[uid] :
            if action == "add" :
                retour = addArriveeTemps(tpsCoureur, tpsClient, tpsServeur, dossard)
            elif action =="del" :
                if local :
                # si la demande vient de l'interface GUI du serveur, c'est le tempsReel qu'il faut supprimer, la liste arriveeTemps est classée par rapport à celui-ci.
                    retour = delArriveeTemps(tpsCoureur, dossard)
                else :
                # si la demande vient d'un smartphone, on cherche à supprimer le tempsCoureur (celui mesuré sur le smartphone). La liste arriveeTemps n'est pas par ordre croissant du client.
                    retour = delArriveeTempsClient(tpsCoureur, dossard)
            elif action == "affecte" :
                Tps = Temps(tpsCoureur, tpsClient, tpsServeur)
                if local :
                # si la demande vient de l'interface GUI du serveur, c'est le tempsReel qu'il faut rechercher.
                    retour = affecteDossardArriveeTempsLocal(Tps.tempsReel, dossard)
                else :
                    # si la demande vient d'un smartphone, c'est le tempsCoureur qu'il faut rechercher.
                    retour = affecteDossardArriveeTemps(Tps.tempsCoureur, dossard)
            elif action == "desaffecte" :
                Tps = Temps(tpsCoureur, tpsClient, tpsServeur)
                if local :
                # si la demande vient de l'interface GUI du serveur, c'est le tempsReel qu'il faut rechercher.
                    retour = affecteDossardArriveeTempsLocal(Tps.tempsReel)
                else :
                    # si la demande vient d'un smartphone, c'est le tempsCoureur qu'il faut rechercher.
                    retour = affecteDossardArriveeTemps(Tps.tempsCoureur)
            else :
                print("Action venant du smartphone incorrecte", ligne)
                retour = Erreur(301)
            if uid and noTransmission :
                UIDPrecedents[uid].append(noTransmission)
        else :
            print("UID et noTransmission déjà utilisés : entrée ignorée. Probable problème de communication WIFI.\nLigne = ",ligne)
            retour = Erreur(451)
    elif listeAction[0] =="dossard" :
        dossardPrecedent = str(listeAction[3])
        try :
            uid = int(listeAction[4])
        except :
            uid = 0
        try :
            noTransmission = int(listeAction[5])
        except :
            noTransmission = 0
        if uid not in UIDPrecedents :
            UIDPrecedents[uid]=[]
        if not noTransmission in UIDPrecedents[uid] :
            if action == "add" :
                retour = addArriveeDossard(dossard, dossardPrecedent)
            elif action =="del" :
                retour = delArriveeDossard(dossard, dossardPrecedent)
            else :
                print("Action venant du smartphone incorrecte", ligne)
                retour = Erreur(301)
            if uid and noTransmission :
                UIDPrecedents[uid].append(noTransmission)
        else :
            print("UID et noTransmission déjà utilisés : entrée ignorée. Probable problème de communication WIFI.\nLigne = ",ligne)
            retour = Erreur(451)
    else :
        print("Type d'action venant du smartphone incorrecte", ligne)
        retour = Erreur(301)
    #print('Parametres["calculateAll"] après decodeAction... : ',Parametres["calculateAll"])
    return retour

def selectionnerCoursesEtGroupementsARegenererPourImpression(dossard) :
    cat = Coureurs.recuperer(dossard).course # les courses ne sont plus identifiées aux catégories : categorie(Parametres["CategorieDAge"])
    # on ajoute un flag pour la catégorie du coureur et son groupement indiquant que celles ci devront être regénérées pour les résultats en pdf.
    try :
        Courses[cat].setARegenererPourImpression(True)
        groupementAPartirDeSonNom(Courses[cat].nomGroupement, nomStandard = True).setARegenererPourImpression(True)
    except :
        print("ERREUR AVEC setARegenererPourImpression")
        try :
            print("nom groupement de la catégorie", cat, ":", Courses[cat].categorie)
        except :
            print("Le coureur au dossard",dossard, "n'a pas de course affectée")
        #print("Groupements" ,print(listNomsGroupements(nomStandard = True)))
        #print("Courses",Courses)

def effacerFichierDonnneesSmartphone() :
    print("Effacement des données venant des smartphones  effectué")
    file = "donneesSmartphone.txt"
    if os.path.exists(file) :
        os.remove(file)

def effacerFichierDonnneesLocales() :
    print("Effacement des modifications locales  effectué")
    file = "donneesModifLocale.txt"
    if os.path.exists(file) :
        os.remove(file)

def lignesAPartirDe(fichier, noLigne):
    """ retourne les lignes après la ligne noLigne (incluse)"""
    #print("Traitement des lignes suivantes :")
    L = []
    with open(fichier, 'r') as f:
        for ind, line in enumerate(f):
            if ind >= noLigne-1:
                #print(line, end='')
                L.append(line)
                #print("FONCTION A VERIFIER : doit retourner les lignes après la ligne n°", noLigne)
    return L

def derniereModifFichierDonnneesSmartphoneRecente(fichier):
    """ retourne true si le fichier a été complété par le serveur web depuis la dernière récupération."""
    #print( "Fichier modif :",os.path.getmtime(fichier), "Dernier Import :",Parametres["tempsDerniereRecuperationSmartphone"])
    retour = False
    if os.path.exists(fichier) :
        diff = os.path.getmtime(fichier) - Parametres["tempsDerniereRecuperationSmartphone"]
        if diff > 0 :
            retour = True
    return retour

def derniereModifFichierDonnneesLocalesRecente(fichier):
    """ retourne true si le fichier a été complété par le serveur web depuis la dernière récupération."""
    #print( "Fichier modif :",os.path.getmtime(fichier), "Dernier Import :",Parametres["tempsDerniereRecuperationLocale"])
    retour = False
    if os.path.exists(fichier) :
        diff = os.path.getmtime(fichier) - Parametres["tempsDerniereRecuperationLocale"]
        if diff > 0 :
            retour = True
    return retour

### fonctions de listes.
def listCourses():
    retour = []
##    if len(Courses)==0:
##        #print("There are no Courses.")
##        return retour
    for cat in Courses :
        #tests Courses[cat].top()
        #print(Courses[cat].categorie, Courses[cat].depart, Courses[cat].temps)
        retour.append(Courses[cat].categorie)
    return retour
    ##transaction.commit()

def listEtablissements():
    # print(Coureurs.afficher())
    retour = []
    if Coureurs.nombreDeCoureurs !=0:
        for coureur in Coureurs.liste() :
            # print(coureur)
            if coureur.etablissement not in retour :
                retour.append(coureur.etablissement)
        retour.sort()
    return retour

def listCategories(nomStandard=True):
    retour = []
    if Coureurs.nombreDeCoureurs != 0:
        for coureur in Coureurs.liste() :
            if nomStandard :
                nom = coureur.course
            else :
                nom = groupementAPartirDUneCategorie(coureur.course).nom
            if nom not in retour :
                retour.append(nom)
        retour.sort()
    return retour

def listClasses():
    retour = []
    if Coureurs.nombreDeCoureurs != 0:
        #print("There are classes.")
        for coureur in Coureurs.liste() :
            if coureur.classe not in retour :
                retour.append(coureur.classe)
        retour.sort()
    return retour

def listDossardsDUneClasse(classe):
    retour = []
    if Coureurs.nombreDeCoureurs != 0:
        for coureur in Coureurs.liste() :
            if coureur.classe == classe :
                retour.append(coureur.dossard)
    return retour

def listDossardsDUnGroupement(nom):
    retour = []
    if Coureurs.nombreDeCoureurs!=0:
        for coureur in Coureurs.liste() :
            if nomGroupementAPartirDUneCategorie(coureur.categorie(CategorieDAge)) == nom :
                retour.append(coureur.dossard)
    return retour

def listDossardsDUneCategorie(cat):
    retour = []
    if Coureurs.nombreDeCoureurs!=0:
        for coureur in Coureurs.liste() :
            if coureur.categorie(CategorieDAge) == cat :
                retour.append(coureur.dossard)
    return retour

def listCoureursDUneClasse(classe):
    retour = []
    if Coureurs.nombreDeCoureurs != 0:
        for coureur in Coureurs.liste()  :
            if coureur.classe == classe :
                retour.append(coureur)
    return triParNomPrenomCoureurs(retour)

def listCoureursDUneCourse(course, nomStandard=True):
    retour = []
    for coureur in Coureurs.liste()  :
        #print("fonction listCoureursDUneCourse")
        if nomStandard :
            nom = coureur.course
        else :
            nom = groupementAPartirDUneCategorie(coureur.course).nom
        if nom == course :
            retour.append(coureur)
    return triParNomPrenomCoureurs(retour)

def listCoureursDUneCategorie(categorie):
    retour = []
    for coureur in Coureurs.liste()  :
        if coureur.categorie(CategorieDAge) == categorie :
            retour.append(coureur)
    return triParNomPrenomCoureurs(retour)

def listCoureursDUnEtablissement(etablissement):
    retour = []
    for coureur in Coureurs :
        if coureur.etablissement == etablissement :
            retour.append(coureur)
    return triParNomPrenomCoureurs(retour)

##def listCoureursDUneCourse(course):
##    retour = []
##    if Coureurs.nombreDeCoureurs !=0:
##        for coureur in Coureurs.liste()  :
##            if coureur.categorie() == course :
##                retour.append(coureur)
##    return triParNomPrenomCoureurs(retour)

def listChallenges():
    ''' Un challenge existe uniquement pour les CategorieDAge==0 (cross du collège) ou 2 (UNSS)
    et si l'on trouve une course dont le nom standard est identique en G et en F'''
    listeCourses = []
    retour = []
    if len(Courses)!=0 and (Parametres["CategorieDAge"]== 0 or Parametres["CategorieDAge"]== 2) :
        # print("There are Courses.", Courses)
        # print("Coureurs", Coureurs.afficher())
        for cat in Courses :
            #tests Courses[cat].top()
            #print(Courses[cat].categorie, Courses[cat].depart, Courses[cat].temps)
            listeCourses.append(Courses[cat].categorie)
        for cat in listeCourses :
            if Parametres["CategorieDAge"]== 0 :
                NomDuChallenge = cat[0]
            else :
                NomDuChallenge = cat[:2] 
            #print("nom de challenge potentiel", NomDuChallenge)
            if NomDuChallenge + "-F" in listeCourses and NomDuChallenge + "-G" in listeCourses :
                ### en théorie, il faudrait créer le challenge même s'il n'y a que des filles cadettes et des garçons juniors. 
                ### Actuellement, c'est un "bug" qui n'apparaitra jamais car il y a toujours des coureurs en cadets et junior dans les deux sexes.
                if Parametres["CategorieDAge"]== 2 and \
                   NomDuChallenge in [ "M10","M9","M8","M7", "M6","M5", "M4","M3" ,"M2", "M1" ,"M0" , "SE" ,"ES", "JU" , "CA" ] : # cas du challenge UNSS lycée qui mélange tous les lycéens au dessus de Cadet !
                    # on ajoute les deux challenges LP et LG violemment.
                    if "LP" not in retour :
                        retour.append("LP")
                    NomDuChallenge = "LG"
                if NomDuChallenge not in retour :
                    retour.append(NomDuChallenge)
    return retour

def listCoursesEtChallenges():
    retour = []
    if len(Courses)!=0:
        for cat in Courses :
            #tests Courses[cat].top()
            #print(Courses[cat].categorie, Courses[cat].depart, Courses[cat].temps)
            retour.append(Courses[cat].categorie)
        if not Parametres["CategorieDAge"] :
            for cat in retour :
                NomDuChallenge = cat[0]
                if NomDuChallenge + "-F" in retour and NomDuChallenge + "-G" in retour and NomDuChallenge not in retour :
                    retour.append(NomDuChallenge)
    return retour
    ##transaction.commit()

##def listCoursesCommencees():
##    retour = []
##    if len(Courses)==0:
##        #print("There are no Courses.")
##        return retour
##    for cat in Courses :
##        #tests Courses[cat].top()
##        #print(Courses[cat].categorie, Courses[cat].depart, Courses[cat].temps)
##        if Courses[cat].temps != 0 :
##            retour.append(Courses[cat].categorie)
##    return retour
##
##def listCoursesNonCommencees():
##    retour = []
##    for cat in Courses :
##        if Courses[cat].temps == 0 :
##            retour.append(Courses[cat].categorie)
##    return retour

def listNomsGroupementsCommences(nomStandard = True):
    retour = []
    for groupement in Groupements :
        if groupement.listeDesCourses :
            nomDeLaPremiereCourseDuGroupement = groupement.listeDesCourses[0]
            if Courses[nomDeLaPremiereCourseDuGroupement].temps != 0 :
                if nomStandard :
                    retour.append(groupement.nomStandard)
                else :
                    retour.append(groupement.nom)
    return retour

def listNomsGroupementsNonCommences(nomStandard = True):
    #print("listNomsGroupementsNonCommences",Courses)
    retour = []
    for groupement in Groupements :
        if groupement.listeDesCourses :
            #print("groupement.listeDesCourses",groupement.listeDesCourses)
            nomDeLaPremiereCourseDuGroupement = groupement.listeDesCourses[0]
            #print("nomDeLaPremiereCourseDuGroupement",nomDeLaPremiereCourseDuGroupement)
            if Courses[nomDeLaPremiereCourseDuGroupement].temps == 0 :
                if nomStandard :
                    retour.append(groupement.nomStandard)
                else :
                    retour.append(groupement.nom)
    return retour

def listNomsGroupements(nomStandard = False, sansSlashNiEspace = False):
    retour = []
    for groupement in Groupements :
        if groupement.listeDesCourses :
            if nomStandard :
                nom = groupement.nomStandard
                if sansSlashNiEspace :
                    nom = nom.replace(" ","").replace("/","-")
            else :
                nom = groupement.nom
                if sansSlashNiEspace :
                    nom = nom.replace(" ","").replace("/","-")
            retour.append(nom)
    #print("liste des noms de groupements", retour)
    return retour

def listNomsGroupementsEtChallenges(nomStandard = False):
    retour = listNomsGroupements(nomStandard)
    retour += listChallenges()
    return retour


def listNomGroupements():
    retour = []
    for groupement in Groupements :
        retour.append(groupement.nom)
    return retour

def listGroupementsCommences():
    retour = []
    for groupement in Groupements :
        if groupement.listeDesCourses :
            nomDeLaPremiereCourseDuGroupement = groupement.listeDesCourses[0]
            if Courses[nomDeLaPremiereCourseDuGroupement].temps != 0 :
                retour.append(groupement)
    return retour

def listGroupementsNonCommences(nomStandard = False):
    retour = []
    for groupement in Groupements :
        if groupement.listeDesCourses :
            nomDeLaPremiereCourseDuGroupement = groupement.listeDesCourses[0]
            if Courses[nomDeLaPremiereCourseDuGroupement].temps == 0 :
                retour.append(groupement)
    return retour

def topDepart(listeDeGroupements):
    temps = time.time()
    if listeDeGroupements :
        for groupement in listeDeGroupements :
            for cat in groupement.listeDesCourses :
                Courses[cat].setTemps(temps, tempsAuto=True)
                print(Courses[cat].categorie, "est lancée :", Courses[cat].depart, ". Heure de départ :", Courses[cat].temps)

# pour corriger un départ depuis l'interface
def fixerDepart(nomGroupement,temps):
    for groupement in Groupements :
        if groupement.nom == nomGroupement :
            for categorie in groupement.listeDesCourses :
                Courses[categorie].setTempsHMS(temps)
                print(Courses[categorie].categorie, "est lancée :", Courses[categorie].depart, ". Heure de départ :", Courses[categorie].temps)


def generateListCoureursPourSmartphone() :
    """ on génère désormais un fichier CoureursA.txt, CoureursB.txt, etc... par course pour une recherche rapide à la n-ème ligne des coordonnées du coureur en fonction de son dossard.
        on ajoute Coureurs.txt qui contient tous les coureurs de toutes les courses pour les recherches par nom-prénom-catégorie-classe. Cela permet de n'avoir qu'un fichier à ouvrir pour le script CGI.
    """
    fichierDonneesSmartphoneAvecTousLesCoureurs = "Coureurs.txt"
    fichierDonneesSmartphone = "Coureurs"
    print("Catégorie d'age paramétrée : ",Parametres["CategorieDAge"])
    fComplet = open(fichierDonneesSmartphoneAvecTousLesCoureurs, 'w')
    for lettre in Coureurs.cles() :
        with open(fichierDonneesSmartphone + lettre + ".txt", 'w') as f :
            for coureur in Coureurs[lettre] :
                try :
                    #print("categorie",coureur.categorie(Parametres["CategorieDAge"]))
                    #print("description",Courses[coureur.categorie(Parametres["CategorieDAge"])].description
                    if CoursesManuelles :
##                        print("Nom : " , coureur.nom)
##                        print("course :",coureur.course)
##                        print("Description:",Courses[coureur.course].description)
##                        print("Groupements",listNomGroupements())
##                        nomStandard = Courses[coureur.course].description
##                        c = groupementAPartirDUneCategorie(nomStandard).nom
                        c = Courses[coureur.course].description
                    else :
                        c = Courses[coureur.categorie(Parametres["CategorieDAge"])].description

                    result = str(coureur.dossard) + "," + str(coureur.nom) + "," + str(coureur.prenom) +","+ str(coureur.classe) + "," +\
                             str(coureur.categorie(Parametres["CategorieDAge"])) + "," +\
                             str(c) + "," +\
                             str(coureur.commentaireArrivee).replace(",",";") + \
                             "," + str(coureur.etablissement)
                except :
                    result = str(coureur.dossard) + "," + str(coureur.nom) + "," + str(coureur.prenom) +","+ str(coureur.classe) + "," + \
                             "," + "," +str(coureur.commentaireArrivee).replace(",",";") + "," + str(coureur.etablissement)
                    #print("catégorie",coureur.categorie(Parametres["CategorieDAge"]))
                    #print("Courses.keys()", Courses.keys())
                    #print("course", Courses[coureur.course].description)
                    print("Coureur non pleinement ajouté à la liste pour les smartphones", str(coureur.dossard) + "," + str(coureur.nom) + "," + \
                          str(coureur.prenom) +","+ str(coureur.classe) + "," + str(coureur.categorie(Parametres["CategorieDAge"])) + "," + \
                          str(coureur.commentaireArrivee))
                result += "\n"
                f.write(result)
                fComplet.write(result)
        f.close()
    fComplet.close()

def generateQRcode(n) :
    osCWD = os.getcwd()
    chemin = "dossards" + os.sep + "QRcodes" + os.sep
    if not os.path.exists(chemin + str(n) + ".pdf") :
        #if n < 2 : # pour les tests
        print("création du QR-code" , n)
        with open("./modeles/QRcode.tex", 'r') as f :
            contenu = f.read()
        f.close()
        contenu = contenu.replace("@dossard@",str(n))
        TEXDIR = "dossards"+os.sep+"QRcodes"+os.sep+"tex"+os.sep
        creerDir(TEXDIR)
        with open(TEXDIR+str(n)+ ".tex", 'w') as f :
            f.write(contenu)
        f.close()
        compilateurComplete = compilateur.replace("@dossier@","dossards"+os.sep+"QRcodes")#.replace('-output-directory=".."','-output-directory=".."'+os.sep+"QRcodes")
        compiler(compilateurComplete, TEXDIR, str(n) + ".tex" , 1)
    ##else :
     #   print("le QR-code existe déjà")

def generateQRcodes() :
    print("Création des QR-codes nécessaires")
    i = 0
    L = Coureurs.liste()
    while i < len(L) :
        generateQRcode(L[i].dossard)
        i += 1
    print("Fin de la création des QR-codes.")
    
def generateQRcodesCoursesManuelles() :
    print("Création des QR-codes nécessaires pour les coursesManuelles")
    L = Coureurs.cles()
    for lettre in L :
        i = 1
        while i <= Parametres["nbreDossardsAGenererPourCourseManuelles"] :
            generateQRcode(str(i) + lettre)
            i += 1
    print("Fin de la création des QR-codes.")

def retourneDossardsNG(listeDeCoureurs, completeFichierParCategorie=False, imprimerLesAbsentsEtDispenses=True) :
    retour = ""
    # utilisation du modèle de dossard.
    modeleDosssard = "./modeles/dossards/" + dossardModele + ".tex"
    with open(modeleDosssard , 'r') as f :
        modele = f.read()
    f.close()
    ## génération du code tex pour le(s) dossard(s)
    for coureur in listeDeCoureurs :
        if imprimerLesAbsentsEtDispenses or (not coureur.dispense and not coureur.absent) :
            print("Création du dossard de ", coureur.nom, coureur.dossard, "pour la course" , coureur.course)
            groupementNom = groupementAPartirDeSonNom(coureur.course, nomStandard = True).nom
            chaineComplete = replaceDansDossardEnFonctionDesParametres(modele, coureur)
            if completeFichierParCategorie :
                groupementNomPourNomFichier = groupementNom.replace(" ","").replace("/","-")
                TEXDIR = "dossards"+os.sep+"tex"+os.sep
                creerDir(TEXDIR)
                with open(TEXDIR+ groupementNomPourNomFichier + ".tex", 'a',encoding="utf-8") as fileCat :
                    fileCat.write(chaineComplete+ "\n\n")
                fileCat.close()
            retour += chaineComplete
    return retour
    
def replaceDansDossardEnFonctionDesParametres(modele, coureur) :
    # gestion des logos personnalisés
    ## utilisation du bon logo.
    if os.path.exists("media/logo.png") :
        logoPersonnalise = "media/logo.png"
    elif os.path.exists("media/logo.jpg") :
        logoPersonnalise = "media/logo.jpg"
    else :
        logoPersonnalise = "media/logo-HB.png"
    if os.path.exists("media/logo-UNSS.png") :
        logoUNSSPersonnalise = "media/logo-UNSS.png"
    else :
        logoUNSSPersonnalise = "media/logo-UNSS48.png"
    # Création des chaines à afficher sous condition
    cat = "Catégorie : " + coureur.categorie(Parametres["CategorieDAge"])
    groupement = "Course : " + groupementAPartirDeSonNom(coureur.course, nomStandard = True).nom #nomGroupementAPartirDUneCategorie(cat,nomStandard=False)
    cl = "(" + coureur.classe + ")"
    etab = "(" + coureur.etablissement + ")"
    # A afficher dans tous les cas.
    modele = modele.replace("@nom@",coureur.nom.upper()).replace("@prenom@",coureur.prenom)\
                .replace("@dossard@",coureur.getDossard(avecLettre=False)).replace("@qrcode@",str(coureur.dossard))\
                .replace("@intituleCross@",Parametres["intituleCross"]).replace("@lieu@",Parametres["lieu"])\
                .replace("@logo@",logoPersonnalise).replace("@logoUNSS@",logoUNSSPersonnalise).replace("@lettreCourse@",coureur.course)
    if CategorieDAge == 0 : # cas du cross du collège : seuls les noms de classe sont importants
        retour = modele.replace("@classe@",cl).replace("@categorie@","")\
                       .replace("@groupement@","").replace("@etablissement@","")
    elif CategorieDAge == 1 :
        if CoursesManuelles : # cas de courses personnalisées : trail Randon
            retour = modele.replace("@classe@","").replace("@categorie@","")\
                       .replace("@groupement@",groupement).replace("@etablissement@","")
        else : # cas de courses par catégorie de la FFA
            retour = modele.replace("@classe@","").replace("@categorie@","")\
                       .replace("@groupement@",groupement).replace("@etablissement@","")
    else : # categorieDAge == 2 (cross UNSS)
        retour = modele.replace("@classe@","").replace("@categorie@","")\
                       .replace("@groupement@",groupement).replace("@etablissement@",etab)
    return retour
   
def getEnTetePersonnalise() :
    """ certains modèles de dossards nécessitent un en-tête personnalisé """
    enTetePersonnalise = "./modeles/en-tete-" + dossardModele + ".tex"
    print("Recherche d'un en-tête personnalisé :", enTetePersonnalise)
    if os.path.exists(enTetePersonnalise) :
        with open(enTetePersonnalise, 'r',encoding="utf-8") as f :
            entete = f.read()
        f.close()
    else :
        with open("./modeles/dossard-en-tete.tex", 'r',encoding="utf-8") as f :
            entete = f.read()
        f.close()
    return entete
    
def generateDossardsNG() :
    print("Utilisation de generateDossardsNG 2ème génération")
    generateQRcodes() # génère autant de QR-codes que nécessaire
    """ générer tous les dossards dans un fichier ET un fichier par catégorie => des impressions sur des papiers de couleurs différentes seraient pratiques"""
    # charger dans une chaine un modèle avec %nom% etc... , remplacer les variables dans la chaine et ajouter cela aux fichiers résultats.
    global CoureursParClasse
    entete = getEnTetePersonnalise()
    with open("./modeles/listing-en-tete.tex", 'r',encoding="utf-8") as f :
        enteteL = f.read()
    f.close()
    TEXDIR = "dossards"+os.sep+"tex"+os.sep
    creerDir(TEXDIR)
    ## effacer les tex existants
    liste_fichiers_tex_complete=glob.glob(TEXDIR+"**"+os.sep+'*.tex',recursive = True)
    liste_fichiers_pdf_complete=glob.glob("dossards"+os.sep+'*.pdf',recursive = False)
    for file in liste_fichiers_tex_complete + liste_fichiers_pdf_complete :
        os.remove(file)
    ## générer de nouveaux en-têtes.
    osCWD = os.getcwd()
    #os.chdir("dossards")
    listeCategories = listNomsGroupements(nomStandard=False,sansSlashNiEspace=True)
    listeCategories.append("0-tousLesDossards")
    for file  in listeCategories :
        print("Création du fichier",file + ".pdf")
        with open(TEXDIR+file+ ".tex", 'w',encoding="utf-8") as f :
            f.write(entete + "\n\n")
        f.close()
    ### création de tous les dossards + ceux par catégorie complétés.
    with open(TEXDIR+"0-tousLesDossards.tex", 'a',encoding="utf-8") as f :
        f.write(retourneDossardsNG(Coureurs.liste(), completeFichierParCategorie=True, imprimerLesAbsentsEtDispenses=False))
    f.close()
    ### ajout du enddocument à la fin de tous les fichiers de dossards générés
    for file  in listeCategories :
        with open(TEXDIR+file+ ".tex", 'a',encoding="utf-8") as f :
            f.write("\n\\end{document}")
        f.close()
    ### création d'un listing sous forme de tableau
    print("Générer listing tableau",genererListing)
    if genererListing :
        with open(TEXDIR+"0-listing.tex", 'a',encoding="utf-8") as fL :
            fL.write(enteteL + "\n\n")
            L = list(CoureursParClasse.keys())
            L.sort()
            for nomClasse in L :
                alimenteListingPourClasse(nomClasse, fL, listingSimple=True)
                if nomClasse != L[:-1] :
                    fL.write("\\newpage\n\n")
            fL.write("\\end{document}")
        fL.close()
    ### création du listing de QR-codes.
    print("Générer listing tableau",genererListingQRcodes)
    if genererListingQRcodes :
        with open(TEXDIR+"0-listing-QRCodes.tex", 'w',encoding="utf-8") as fL :
            fL.write(enteteL + "\n\n")
            L = list(CoureursParClasse.keys())
            L.sort()
            for nomClasse in L :
                alimenteListingPourClasse(nomClasse, fL)
                if nomClasse != L[:-1] :
                    fL.write("\\newpage\n\n")
            fL.write("\\end{document}")
        fL.close()
    #### création des QR-codes pour imprimer à part (cross de Rieutort)
    # if CoursesManuelles and genererQRcodesPourCourseManuelles :
        # with open("./modeles/qrcodes-en-tete.tex", 'r',encoding="utf-8") as f :
            # enteteQR = f.read()
        # f.close()
        # generateQRcodesCoursesManuelles()
        # fichier  = "0-QR-codes-pour-ajout-sur-dossards-existants"
        # ## création d'un fichier de QR-codes pour impression - plastifiage - agrafage sur d'autres dossards existants.
        # with open(TEXDIR+ fichier + ".tex", 'a',encoding="utf-8") as fL :
            # fL.write(enteteQR + "\n\n")
            # L = Coureurs.cles()
            # Coureurs.afficher()
            # print("Affichage des Coureurs pour comprendre")
            # print("Clés",Coureurs.cles())
            # for nomCourse in L :
                # alimenteListingPourCourse(nomCourse, fL)
                # if nomCourse != L[:-1] :
                    # fL.write("\n\\newpage\n\n")
            # fL.write("\\end{document}")
        # fL.close()
    ### compilation de tous les fichier sprésents 
    compilerTousLesTex(TEXDIR, "dossards")
    #print(listeCategories)
    # pour chaque fichier dans listeCategories , ajouter le end document.
    # for file in listeCategories :
        # with open(TEXDIR+file + ".tex", 'a',encoding="utf-8") as f :
            # f.write("\\end{document}")
        # f.close()
        # # il faut compiler tous les fichiers de la liste.
        # #print(file)
        # compilateurComplete = compilateur.replace("@dossier@","dossards")
        # compilerDossards(compilateurComplete, ".", file + ".tex" , 1)


def generateDossardsAImprimer() :
    """ générer tous les dossards non encore imprimés (créés manuellement) dans un fichier pdf spécifique.
        Retourne la liste des numéros de dossards qui ont été ajoutés dans le pdf à imprimer."""
    # charger dans une chaine un modèle avec %nom% etc... , remplacer les variables dans la chaine et ajouter cela aux fichiers résultats.
    #print("Utilisation de generateDossardsAImprimer")
    retour=[]
    entete = getEnTetePersonnalise()
    TEXDIR = "dossards"+os.sep+"tex"+os.sep
    creerDir(TEXDIR)
    # utilisation du modèle de dossard.
##    if Parametres["CategorieDAge"] :
##        modeleDosssard = "./modeles/dossard-modele.tex"
##    else :
##        modeleDosssard = "./modeles/dossard-modele-classe.tex"
##    modeleDosssard = "./modeles/" + dossardModele
##    with open(modeleDosssard, 'r') as f :
##        modele = f.read()
##    f.close()
    ##  génère la liste des coureurs concernés par l'impression
    listeAImprimer = []
    for coureur in Coureurs.liste() :
        if not coureur.dispense and not coureur.absent and coureur.aImprimer : # si le coureur a été créé manuellement et n'a pas été imprimé.
            listeAImprimer.append(coureur)
            retour.append(coureur.dossard)
    ## générer de nouveaux en-têtes.
    osCWD = os.getcwd()
    if listeAImprimer :
        with open(TEXDIR+"A-imprimer.tex", 'w',encoding="utf-8") as f :
            f.write(entete + "\n\n")
            f.write(retourneDossardsNG(listeAImprimer, completeFichierParCategorie=False, imprimerLesAbsentsEtDispenses=False))
            f.write("\\end{document}")
        f.close()
        compilateurComplete = compilateur.replace("@dossier@","dossards")
        print(compilerDossards(compilateurComplete, ".", "A-imprimer.tex" , 1))
    else :
        print("Il n'y a aucun dossard à imprimer qui ne l'ait pas déjà été.")
    return retour

def generateDossard(coureur) :
    """ générer un dossard dans un fichier et l'ouvrir dans le lecteur pdf par défaut"""
    # charger dans une chaine un modèle avec %nom% etc... , remplacer les variables dans la chaine et ajouter cela aux fichiers résultats.
    print("Utilisation de generateDossard pour le coureur", coureur.nom, coureur.prenom, coureur.dossard)
    entete = getEnTetePersonnalise()
    TEXDIR = "dossards"+os.sep+"tex"+os.sep
    creerDir(TEXDIR)
    # utilisation du modèle de dossard.
##    modeleDosssard = "./modeles/" + dossardModele
##    with open(modeleDosssard, 'r') as f :
##        modele = f.read()
##    f.close()
    ## générer de nouveaux en-têtes.
    osCWD = os.getcwd()
    #os.chdir("dossards")
    file = coureur.nom.replace(" ","-") + "-" + coureur.prenom.replace(" ","-")
    with open(TEXDIR+file+ ".tex", 'w',encoding="utf-8") as f :
        f.write(entete + "\n\n")
        f.write(retourneDossardsNG([coureur], completeFichierParCategorie=False, imprimerLesAbsentsEtDispenses=True)+ "\n\n")
        f.write("\\end{document}")
    f.close()
    generateQRcode(coureur.dossard)
    compilateurComplete = compilateur.replace("@dossier@","dossards")
    print(compilerDossards(compilateurComplete, ".", file + ".tex" , 1))
    fichierAOuvrir = "dossards" + sep + file+".pdf"
    # Finalement, on retourne le nom du fichier.
    # celui-ci sera imprimé directement sous windows et ouvert dans le lecteur pdf par défaut sous les OS unix
    # subprocess.Popen([fichierAOuvrir],shell=True)
    #open(fichierAOuvrir)
    return fichierAOuvrir

def alimenteListingPourClasse(nomClasse, file, listingSimple = False):
    nomAffiche = nomClasse
    if Parametres["CategorieDAge"] == 2 : # UNSS
        denomination = ""
    elif Parametres["CategorieDAge"] == 1 : # catégories FFA
        if CoursesManuelles :
            denomination = "Course"
            nomAffiche = groupementAPartirDUneCategorie(nomClasse).nom
        else :
            denomination = "Catégorie"
    else : # cross du collège
        denomination = "Classe"
    debutTab = """{}\\hfill {}
%\\fcolorbox{black}{gray!30}{
%\\begin{minipage}{0.9\\textwidth}
\\Huge
\\begin{center}
\\textbf{""" + denomination + " " + nomAffiche + """}
\\end{center}
%\\end{minipage}
%}
%{}\\hfill {}

\\normalsize
"""

    if listingSimple :
        debutTab += """
    \\begin{longtable}{| p{0.3\\textwidth} | p{0.3\\textwidth} | p{0.3\\textwidth} |}
    \\hline
    """
        file.write(debutTab)
        #print("CoureursParClasseOrdonnes[nomClasse]",CoureursParClasseOrdonnes[nomClasse])
        for element in CoureursParClasseOrdonnes[nomClasse] :
            file.write(alimenteLignePourListingClasse(element))
    else :
        debutTab += """
    \\begin{longtable}{| p{0.23\\textwidth} | p{0.23\\textwidth} | p{0.23\\textwidth} | p{0.23\\textwidth}|}
    \\hline
    """
        file.write(debutTab)
        ligne = 1
        while ligne <= (len(CoureursParClasseOrdonnes[nomClasse])//4 + 1) :
            imin = (ligne - 1)* 4
            imax = ligne * 4
            if imax >= len(CoureursParClasseOrdonnes[nomClasse]):
                listeDeQuatreCoureursMax = CoureursParClasseOrdonnes[nomClasse][imin:]
            else :
                listeDeQuatreCoureursMax = CoureursParClasseOrdonnes[nomClasse][imin:imax]
            file.write(alimenteLignePourListingQRCodesClasse(listeDeQuatreCoureursMax))
            ligne += 1
    file.write("\n\\end{longtable}\n\n")
    


    
def alimenteListingPourCourse(nomCourse, file):
    if Parametres["CategorieDAge"] :
        denomination = "Catégorie"
    else :
        denomination = "Classe"
    debutTab = """
\\begin{longtable}{p{10cm} p{10cm}}
\\endhead
"""
    file.write(debutTab)
    partie1 = """\\begin{minipage}{\\linewidth} {}

\\smallskip
{}\\hfill {} {\\footnotesize """
    partie2 = """ } {}\\hfill {}
    
\\medskip
{}\\hfill {} \\includegraphics[width=8cm]{QRcodes/"""
    partie3 = """.pdf} {}\\hfill {}
    
\\vspace{0.8cm}
\\bigskip
\\vfill
{}
\\end{minipage}"""
    doss = 1
    while doss <= Parametres["nbreDossardsAGenererPourCourseManuelles"] :
        #file.write("\\includegraphics[width=\\linewidth]{QRcodes/" + str(doss) + nomCourse + ".pdf}")
        file.write(partie1 + str(doss) + nomCourse + partie2 +  str(doss) + nomCourse + partie3 )
        if doss % 2 == 0 :
            file.write("\\\\\n")
            if doss % 6 == 0 :
                file.write("\n\\newpage\n")
        else :
            file.write(" & ")
        doss += 1
    file.write("\n\\end{longtable}\n\n")
    

def alimenteLignePourListingClasse(coureur) :
    contenuLigne = "@nom@  &  @prenom@ & @dossard@ \\hline \n"
    return contenuLigne.replace("@dossard@", coureur.dossard).replace("@nom@", coureur.nom).replace("@prenom@",coureur.prenom)


def alimenteLignePourListingQRCodesClasse(listeDeQuatreCoureursMax) :
    retour = ""
    i = 0
    #test = []
    if listeDeQuatreCoureursMax : # si la liste est non vide, on complète le tableau.
        while i < 4 :
            if i < len(listeDeQuatreCoureursMax) :
                nom = listeDeQuatreCoureursMax[i].nom
                if len(nom) >= 11 :
                    nom = nom[:10] # nom coupé à 10 caractères maximum
                prenom = listeDeQuatreCoureursMax[i].prenom
                dossard = listeDeQuatreCoureursMax[i].dossard
                #test.append(nom +  prenom + str(dossard))
                retour += alimenteCellulePourListingClasse(nom, prenom, dossard)
            i += 1
            if i == 4 :
                retour += " \\\\\n \\hline"
            else :
                retour += " & \n"
    #print(test)
    return retour

def alimenteCellulePourListingClasse(nom, prenom, dossard) :
    # ancienne ligne où l'on compilait le QR-code : {}\\hfill {} \\qrcode[height=0.5\\linewidth]{@dossard@} {}\\hfill {}
    contenuCellule = """\\begin{minipage}{\\linewidth}
{}

\\smallskip
{}\\hfill {} {\\footnotesize @dossard@ @prenom@ @nom@ } {}\\hfill {}

\\medskip
{}\\hfill {} \\includegraphics[width=0.5\\linewidth]{QRcodes/@dossard@.pdf} {}\\hfill {}

\\bigskip
\\end{minipage} """
    return contenuCellule.replace("@dossard@", str(dossard)).replace("@nom@", nom).replace("@prenom@",prenom)

CoureursParClasse = {}
CoureursParClasseOrdonnes = {}

def CoureursParClasseUpdate():
    global CoureursParClasse, CoureursParClasseOrdonnes
    CoureursParClasse.clear()
    CoureursParClasseOrdonnes.clear()
    if CategorieDAge == 2 : # cas UNSS
        for c in Coureurs.liste() :
            if not c.etablissement in CoureursParClasse.keys() :
                CoureursParClasse[c.etablissement]=[]
            CoureursParClasse[c.etablissement].append(c)
    elif CategorieDAge == 1 : # cas catégories d'age
        if CoursesManuelles :
            for c in Coureurs.liste() :
                if not c.course in CoureursParClasse.keys() :
                    CoureursParClasse[c.course]=[]
                CoureursParClasse[c.course].append(c)
        else :
            for c in Coureurs.liste() :
                if not c.categorie(True) in CoureursParClasse.keys() :
                    CoureursParClasse[c.categorie(True)]=[]
                CoureursParClasse[c.categorie(True)].append(c)
    else : # cas cross de collège
        for c in Coureurs.liste() :
            if not c.classe in CoureursParClasse.keys() :
                CoureursParClasse[c.classe]=[]
            CoureursParClasse[c.classe].append(c)
    # tri par ordre alphabétique
    #CoureursParClasseOrdonnes = dict(CoureursParClasse)
    for k in CoureursParClasse :
        CoureursParClasseOrdonnes[k] = triParNomPrenomCoureurs(CoureursParClasse[k])
        #print("CoureursParClasseOrdonnes[k]",CoureursParClasseOrdonnes[k])



def moyenneDesTemps(listeDesTemps) :
    """A coder : argument une liste de temps en secondes
    retour : un temps moyen en secondes."""
    #print("Liste des temps",listeDesTemps)
    if listeDesTemps :
        tpsEnSecondes = sum(listeDesTemps)/len(listeDesTemps)
        #tps = Temps(tpsEnSecondes,0,0)
        retour = formaterDuree(tpsEnSecondes)
    else :
        retour = ""
    return retour

def medianeDesTemps(listeDesTemps) :
    """A coder : argument une liste de temps en secondes
    retour : un temps median en secondes."""
    if listeDesTemps :
        if len(listeDesTemps)%2 == 0 :
            # pair
            i = len(listeDesTemps)//2
            tpsEnSecondes = (listeDesTemps[i-1]+listeDesTemps[i])/2
        else :
            # impair
            i = len(listeDesTemps)//2
            tpsEnSecondes = listeDesTemps[i]
        #tps = Temps(tpsEnSecondes,0,0)
        retour = formaterDuree(tpsEnSecondes)#tps.tempsReelFormate()
    else :
        retour = ""
    return retour

def pourcentage(eff, effTot):
    if effTot :
        retour = str(round(100*eff/effTot))+"\\%"
    else :
        retour = "0\\%"
    return retour

def testTMPStats():
    with open("./modeles/impression-en-tete.tex", 'r') as f :
        entete = f.read()
    f.close()
    print(creerFichierClasse("36",entete))

def supprimerFichier(file):
    try :
        os.remove(file)
        if DEBUG :
            print("suppression du fichier",file)
    except :
        texte = "Impossible de supprimer le fichier "+file+" car il est ouvert dans un autre programme.\nIl ne sera pas regénéré."
        showinfo("ERREUR !",texte)
        if DEBUG :
            print(texte)

def selectPlusRecent(dossier,formatDuNom):
    fichierSelectionne = None
    datePrecedente = 0
    fichierSelectionne = ""
    for file in glob.glob(dossier + os.sep + formatDuNom,recursive = False) :
        if datePrecedente < os.path.getmtime(file) :
            datePrecedente = os.path.getmtime(file)
            fichierSelectionne = file
    return fichierSelectionne

def nettoyerTousLesFichiersGeneres():
    ## effacer les dossards et les fichiers les ayant générés.
    DOSSDIR = "dossards"+os.sep
    L1 =glob.glob(DOSSDIR+"tex"+os.sep+'*.tex',recursive = True)
    L2 =glob.glob(DOSSDIR+'*.pdf',recursive = False)
    ## effacer les tex existants d'impressions
    IMPDIR = "impressions"+os.sep
    L3 =glob.glob(IMPDIR+"tex"+os.sep+'*.tex',recursive = False)
    L4 =glob.glob(IMPDIR+'*.pdf',recursive = False)
    listeTotale = L1 + L2 + L3 + L4
    for file in listeTotale :
        try :
            supprimerFichier(file)
        except :
            print("Impossible de supprimer:", file)
    ## effacer les videos.
    if os.path.exists("videos") :
        shutil.rmtree("videos")
    ## effacer les bases de donnees superflues : toutes sauf la dernière.
    L = []
    for schema in ["*.db","*_DS.txt","*_ML.txt"] :
        fichierRecent = selectPlusRecent("db",schema)
        if fichierRecent :
            L.append(fichierRecent)
    ## effacer les fichiers de base de données automatiquement générés + fichiers de LOG
    L5 =glob.glob('db'+os.sep+'*.db',recursive = False)
    L6 =glob.glob('db'+os.sep+'*.txt',recursive = False)
    L7 =glob.glob('logs'+os.sep+'*.txt',recursive = False)
    for file in L5 + L6 :
        if not file in L :
            try :
                supprimerFichier(file)
            except :
                print("Impossible de supprimer:", file)


def generateImpressions() :
    """ générer tous les fichiers tex des impressions possibles et les compiler """
    #print("Resultats avant impressions", ResultatsPourImpressions)
    #print("Courses",Courses)
    StatsEffectifs = True ## à basculer dans les paramètres
    ContenuLignesCategories = ""
    ContenuLignesGroupements = ""
    for DIR in [ "impressions", "impressions"+os.sep+"tex" ]:
        if not os.path.exists(DIR) :
            os.makedirs(DIR)
    # charger dans une chaine un modèle avec %nom% etc... , remplacer les variables dans la chaine et ajouter cela aux fichiers résultats.
    with open("./modeles/impression-en-tete.tex", 'r',encoding="utf-8") as f :
        entete = f.read()
    f.close()
    with open("./modeles/impression-en-teteC.tex", 'r',encoding="utf-8") as f :
        enteteC = f.read()
    f.close()
    with open("./modeles/impression-en-teteS.tex", 'r',encoding="utf-8") as f :
        if Parametres["CategorieDAge"] == 2 :
            enteteS = f.read().replace("@categorie@","établissement")
        elif Parametres["CategorieDAge"] == 1 :
            enteteS = f.read().replace("@categorie@","catégorie")
        else :
            enteteS = f.read().replace("@categorie@","classe")
    f.close()
    with open("./modeles/impression-en-teteSGpments.tex", 'r',encoding="utf-8") as f :
        enteteSGpments = f.read()#.replace("@date",dateDuJour())
    f.close()
    with open("./modeles/stats-ligne.tex", 'r',encoding="utf-8") as f :
        ligneStats = f.read()
    f.close()
    TEXDIR = "impressions"+os.sep+"tex"+os.sep
    ## effacer les tex existants
    liste_fichiers_tex_complete=glob.glob(TEXDIR+"**"+os.sep+'*.tex',recursive = True)
    for file in liste_fichiers_tex_complete :
        os.remove(file) # on supprime tous les .tex
    liste_fichiers_pdf_complete=glob.glob("impressions"+os.sep+"**"+os.sep+'*.pdf',recursive = True)
    for file in liste_fichiers_pdf_complete : # on selectionne les pdf à supprimer
        nomFichierPdfDecoupe = os.path.basename(file)[:-4].split("_")
        try :
            nomFichierPdfDecoupe.remove("")
        except :
            pass
        #print(nomFichierPdfDecoupe, Courses)
        if nomFichierPdfDecoupe[0] == "Classe" :
            catG = nomFichierPdfDecoupe[1][0] + "-G"
            catF = nomFichierPdfDecoupe[1][0] + "-F"
            aSupprimer = Courses[catG].aRegenererPourImpression or Courses[catF].aRegenererPourImpression
        else :
            aSupprimer = False
        try :
            casOuOnSupprime = (aSupprimer) \
                          or (nomFichierPdfDecoupe[0] == "Categorie" and Courses[nomFichierPdfDecoupe[1]].aRegenererPourImpression) \
                          or (nomFichierPdfDecoupe[0] == "Course" and groupementAPartirDeSonNom(Courses[nomFichierPdfDecoupe[1]].nomGroupement, nomStandard = True).aRegenererPourImpression) \
                          or nomFichierPdfDecoupe[0] == "Challenge" or nomFichierPdfDecoupe[0]=="statistiques"
        except :
            casOuOnSupprime = True
        if not casOuOnSupprime : #pas a virer car aucun changement , on conserve.
            print("on conserve le fichier ", file, " car aucun resultat n'est survenu depuis la dernière génération.")
        else : # si categorie ou groupement disparu ou à regénérer, on vire
            # pour l'instant, on vire également les challenges par classe (problème mineur). Optimisation peu importante vu le travail.
            supprimerFichier(file)
    ## générer de nouveaux en-têtes.
    #osCWD = os.getcwd()
    #os.chdir("impressions")
    ### (pas urgent) générer le tex des statistiques ?
    print("Création du fichier de statistiques")
    fstats = open(TEXDIR+"_statistiques.tex", 'w', encoding="utf-8")
    fstats.write(enteteS)

    ### générer les tex pour chaque classe + alimenter les statistiques de chacune
    nbreArriveesTotal = 0
    nbreDispensesTotal = 0
    nbreAbsentsTotal = 0
    nbreAbandonsTotal = 0
    if Parametres["CategorieDAge"] == 2 :
        denomination = "Eleves"
    elif Parametres["CategorieDAge"] == 0 :
        denomination = "Classe"
    else :
        denomination = "Categorie"
    if not CoursesManuelles : # dans le cas de courses manuelles, cela n'a pas de sens de faire des moyennes de temps sur des courses de longueurs différentes
        # on ne fait ces statistiques que pour des courses identiques pour une même catégorie (d'âge ou de niveau pour un cross de collège)
        enTeteDesStatistiquesParCategories = """\textbf{Statistiques par @categorie@ :}

\begin{center}
\begin{tabular}{|*{11}{c|}}
\hline

\multicolumn{1}{|c|}{\multirow{2}{*}{\textbf{@categorie@ }}}
  & 
  \multicolumn{2}{|c|}{\textbf{Arrivés} } 
  & 
  \multicolumn{2}{|c|}{\textbf{Dispensés} }
  & 
  \multicolumn{2}{|c|}{\textbf{Absents} }
  & 
  \multicolumn{2}{|c|}{\textbf{Abandons} }
  & \multicolumn{1}{|c|}{\multirow{2}{*}{\textbf{Moyenne}}} 
  & \multicolumn{1}{|c|}{\multirow{2}{*}{\textbf{Médiane}}} \\
  \cline{2-9}
\multicolumn{1}{|c|}{} & F & G & F & G &F & G &F & G & \multicolumn{1}{|c|}{} & \multicolumn{1}{|c|}{}
 \\
\hline"""
        fstats.write(enTeteDesStatistiquesParCategories)
        for classe in Resultats :  # doute lors d'une fusion manuelle de deux branches. Est ce "for classe in ResultatsPourImpressions:" ?
            #print(classe,"est traité pour création tex", Resultats[classe])
            # si cross du collège, on ne met que les classes dans les statistiques. Si categorieDAge, on met toutes les catégories présentes.
            if (Parametres["CategorieDAge"] or (len(classe) != 1 and classe[-2:] != "-F" and classe[-2:] != "-G")) :
                #print("Création du fichier de "+classe)
                contenu, ArrDispAbsAbandon = creerFichierClasse(classe,entete, False)
                nomFichier = classe.replace(" ","_").replace("__","_")
                if ArrDispAbsAbandon[8] :
                    if not os.path.exists("impressions"+os.sep+denomination +"_"+nomFichier+ ".pdf") :
                        with open(TEXDIR+ denomination +"_"+nomFichier+ ".tex", 'w',encoding="utf-8") as f :
                            f.write(contenu)
                            f.write("\n\\end{longtable}\\end{center}\\end{document}")
                        f.close()
                    # alimentation des statistiques
                    listeDesTempsDeLaClasse = ArrDispAbsAbandon[8]
                    effTot = sum(ArrDispAbsAbandon[:-1])
                    effTotG = ArrDispAbsAbandon[1]+ArrDispAbsAbandon[3]+ArrDispAbsAbandon[5]+ArrDispAbsAbandon[7]
                    effTotF = ArrDispAbsAbandon[0]+ArrDispAbsAbandon[2]+ArrDispAbsAbandon[4]+ArrDispAbsAbandon[6]
                    moyenne = moyenneDesTemps(listeDesTempsDeLaClasse)
                    mediane = medianeDesTemps(listeDesTempsDeLaClasse)
                    ### Statistiques en effectifs par défaut : voir si envie d'avoir des statistiques en % plus tard : tout est prêt dans le else ###
                    if StatsEffectifs :
                        if effTotF :
                            FArr = str(ArrDispAbsAbandon[0]) + "{\\scriptsize /" + str(effTotF) + "}"
                        else :
                            FArr = "{-}"
                        if effTotG :
                            GArr = str(ArrDispAbsAbandon[1]) + "{\\scriptsize /" + str(effTotG) + "}"
                        else :
                            GArr = "{-}"
                        if effTotF :
                            FD = str(ArrDispAbsAbandon[2]) + "{\\scriptsize /" + str( effTotF) + "}"
                        else :
                            FD = "{-}"
                        if effTotG :
                            GD = str(ArrDispAbsAbandon[3]) + "{\\scriptsize /" + str( effTotG) + "}"
                        else :
                            GD = "{-}"
                        if effTotF :
                            FAba = str(ArrDispAbsAbandon[4]) + "{\\scriptsize /" + str( effTotF) + "}"
                        else :
                            FAba = "{-}"
                        if effTotG :
                            GAba = str(ArrDispAbsAbandon[5]) + "{\\scriptsize /" + str( effTotG) + "}"
                        else :
                            GAba = "{-}"
                        if effTotF :
                            FAbs = str(ArrDispAbsAbandon[6]) + "{\\scriptsize /" + str( effTotF) + "}"
                        else :
                            FAbs = "{-}"
                        if effTotG :
                            GAbs = str(ArrDispAbsAbandon[7]) + "{\\scriptsize /" + str( effTotG) + "}"
                        else :
                            GAbs = "{-}"
                    else :
                        FArr = pourcentage(ArrDispAbsAbandon[0], effTotF)
                        GArr = pourcentage(ArrDispAbsAbandon[1], effTotG)
                        FD = pourcentage(ArrDispAbsAbandon[2], effTotF)
                        GD = pourcentage(ArrDispAbsAbandon[3], effTotG)
                        FAba = pourcentage(ArrDispAbsAbandon[4], effTotF)
                        GAba = pourcentage(ArrDispAbsAbandon[5], effTotG)
                        FAbs = pourcentage(ArrDispAbsAbandon[6], effTotF)
                        GAbs = pourcentage(ArrDispAbsAbandon[7], effTotG)
                    # nbreArriveesTotal += ArrDispAbsAbandon[0] + ArrDispAbsAbandon[1]
                    # nbreDispensesTotal += ArrDispAbsAbandon[2] + ArrDispAbsAbandon[3]
                    # nbreAbandonsTotal += ArrDispAbsAbandon[6] + ArrDispAbsAbandon[7]
                    # nbreAbsentsTotal += ArrDispAbsAbandon[4] + ArrDispAbsAbandon[5]
                    #print(classe,FArr,GArr,FD,GD,FAba,GAba,FAbs,GAbs,moyenne,mediane)
                    ContenuLignesCategories += ligneStats.replace("@classe",classe).replace("@FArr",FArr)\
                                 .replace("@GArr",GArr).replace("@FD",FD)\
                                 .replace("@GD",GD).replace("@FAba",FAba)\
                                 .replace("@GAba",GAba).replace("@FAbs",FAbs)\
                                 .replace("@GAbs",GAbs).replace("@moy",moyenne)\
                                 .replace("@med",mediane)
                    ### la catégorie n'a plus à être regénérée sauf modification
                    if denomination != "Classe" :
                        try :
                            Courses[classe].setARegenererPourImpression(False)
                        except :
                            True
                        # si la classe a été générée, la catégorie également via les groupements ci-dessous

    for classe in ResultatsGroupementsPourImpressions :
        #print(classe,"est traité pour création tex", Resultats[classe])
        # si cross du collège, on ne met que les classes dans les statistiques. Si categorieDAge, on met toutes les catégories présentes.
        #if Parametres["CategorieDAge"] or (len(classe) != 1 and classe[-2:] != "-F" and classe[-2:] != "-G") :
        if CoursesManuelles :
            nomCourse = groupementAPartirDUneCategorie(classe).nom
        else :
            nomCourse = classe
        print("Création du fichier de "+classe + " : " + nomCourse)

        # on gardera finalement les groupements pour ne pas afficher les abandons, etc...
        if (not estChallenge(classe)) :# and len(groupementAPartirDeSonNom(classe, nomStandard = True).listeDesCourses) > 1) : # si c'est un groupement ET qu'il comporte plus d'une catégorie, on génère un fichier dédié.

            contenu, ArrDispAbsAbandon = creerFichierClasse(classe,entete, True)
            nomFichier = nomCourse.replace(" ","_").replace("-/-","_").replace("/","_").replace("\\","_").replace("___","_")
            if ArrDispAbsAbandon[8] :
                if not os.path.exists("impressions"+os.sep+"Course_"+nomFichier+ ".pdf") :
                    with open(TEXDIR+"Course_"+nomFichier+ ".tex", 'w',encoding="utf-8") as f :
                        f.write(contenu)
                        f.write("\n\\end{longtable}\\end{center}\\end{document}")
                    f.close()
                # alimentation des statistiques
                listeDesTempsDeLaClasse = ArrDispAbsAbandon[8]
                effTot = sum(ArrDispAbsAbandon[:-1])
                effTotG = ArrDispAbsAbandon[1]+ArrDispAbsAbandon[3]+ArrDispAbsAbandon[5]+ArrDispAbsAbandon[7]
                effTotF = ArrDispAbsAbandon[0]+ArrDispAbsAbandon[2]+ArrDispAbsAbandon[4]+ArrDispAbsAbandon[6]
                #print("listeDesTempsDeLaClasse",classe,":", listeDesTempsDeLaClasse)
                moyenne = moyenneDesTemps(listeDesTempsDeLaClasse)
                mediane = medianeDesTemps(listeDesTempsDeLaClasse)
                ### Statistiques en effectifs par défaut : voir si envie d'avoir des statistiques en % plus tard : tout est prêt dans le else ###
                if StatsEffectifs :
                    if effTotF :
                        FArr = str(ArrDispAbsAbandon[0]) + "{\\scriptsize /" + str(effTotF) + "}"
                    else :
                        FArr = "{-}"
                    if effTotG :
                        GArr = str(ArrDispAbsAbandon[1]) + "{\\scriptsize /" + str(effTotG) + "}"
                    else :
                        GArr = "{-}"
                    if effTotF :
                        FD = str(ArrDispAbsAbandon[2]) + "{\\scriptsize /" + str( effTotF) + "}"
                    else :
                        FD = "{-}"
                    if effTotG :
                        GD = str(ArrDispAbsAbandon[3]) + "{\\scriptsize /" + str( effTotG) + "}"
                    else :
                        GD = "{-}"
                    if effTotF :
                        FAba = str(ArrDispAbsAbandon[4]) + "{\\scriptsize /" + str( effTotF) + "}"
                    else :
                        FAba = "{-}"
                    if effTotG :
                        GAba = str(ArrDispAbsAbandon[5]) + "{\\scriptsize /" + str( effTotG) + "}"
                    else :
                        GAba = "{-}"
                    if effTotF :
                        FAbs = str(ArrDispAbsAbandon[6]) + "{\\scriptsize /" + str( effTotF) + "}"
                    else :
                        FAbs = "{-}"
                    if effTotG :
                        GAbs = str(ArrDispAbsAbandon[7]) + "{\\scriptsize /" + str( effTotG) + "}"
                    else :
                        GAbs = "{-}"
                else :
                    FArr = pourcentage(ArrDispAbsAbandon[0], effTotF)
                    GArr = pourcentage(ArrDispAbsAbandon[1], effTotG)
                    FD = pourcentage(ArrDispAbsAbandon[2], effTotF)
                    GD = pourcentage(ArrDispAbsAbandon[3], effTotG)
                    FAba = pourcentage(ArrDispAbsAbandon[4], effTotF)
                    GAba = pourcentage(ArrDispAbsAbandon[5], effTotG)
                    FAbs = pourcentage(ArrDispAbsAbandon[6], effTotF)
                    GAbs = pourcentage(ArrDispAbsAbandon[7], effTotG)
                nbreArriveesTotal += ArrDispAbsAbandon[0] + ArrDispAbsAbandon[1]
                nbreDispensesTotal += ArrDispAbsAbandon[2] + ArrDispAbsAbandon[3]
                nbreAbandonsTotal += ArrDispAbsAbandon[6] + ArrDispAbsAbandon[7]
                nbreAbsentsTotal += ArrDispAbsAbandon[4] + ArrDispAbsAbandon[5]
                #print(classe,FArr,GArr,FD,GD,FAba,GAba,FAbs,GAbs,moyenne,mediane)
                #if estNomDeGroupement(classe) :
                ContenuLignesGroupements += ligneStats.replace("@classe",groupementAPartirDeSonNom(classe).nom).replace("@FArr",FArr)\
                             .replace("@GArr",GArr).replace("@FD",FD)\
                             .replace("@GD",GD).replace("@FAba",FAba)\
                             .replace("@GAba",GAba).replace("@FAbs",FAbs)\
                             .replace("@GAbs",GAbs).replace("@moy",moyenne)\
                             .replace("@med",mediane)
                ### le groupement et toutes les catégories incluses n'ont plus à être regénérés sauf modification ultérieure
                groupementAPartirDeSonNom(classe).setARegenererPourImpression(False)
                for c in groupementAPartirDeSonNom(classe).listeDesCourses :
                    Courses[c].setARegenererPourImpression(False)
            #else :
            #    ContenuLignesCategories += ligneStats.replace("@classe",classe).replace("@FArr",FArr)\
             #            .replace("@GArr",GArr).replace("@FD",FD)\
              #           .replace("@GD",GD).replace("@FAba",FAba)\
               #          .replace("@GAba",GAba).replace("@FAbs",FAbs)\
                #         .replace("@GAbs",GAbs).replace("@moy",moyenne)\
                 #        .replace("@med",mediane)

        # on ferme le fichier de statistiques des classes
    fstats.write(ContenuLignesCategories)
    fstats.write("\n\\end{tabular}\\end{center}\n ")

    fstats.write(enteteSGpments)
    fstats.write(ContenuLignesGroupements)
    fstats.write("\n\\end{tabular}\\end{center}\n\n\\bigskip")
    fstats.write("\n\n\\textbf{Nombre total d'arrivées : }" + str(nbreArriveesTotal))
    if not CoursesManuelles :
        fstats.write("\n\n\\textbf{Nombre total de dispensés : }" + str(nbreDispensesTotal))
        fstats.write("\n\n\\textbf{Nombre total d'abandons : }" + str(nbreAbandonsTotal))
        fstats.write("\n\n\\textbf{Nombre total d'absents : }" + str(nbreAbsentsTotal))
    fstats.write(absentsDispensesAbandonsEnTex())
    fstats.write("\\end{document}")
    fstats.close()
    ### générer les tex pour chaque challenge
    if Parametres["CategorieDAge"]==0 or Parametres["CategorieDAge"]==2 :
        listeChallenges = listChallenges()
        #print("liste des challenges", listeChallenges)
        for challenge  in listeChallenges :
            #print(ResultatsGroupements[challenge])
            try :
                if ResultatsGroupements[challenge] : # il y a des classes qui ont atteint le nombre d'arrivées suffisantes.
                    print("Création du fichier du challenge", challenge)
                    with open(TEXDIR+"Challenge_"+challenge+ ".tex", 'w',encoding="utf-8") as f :
                        f.write(creerFichierChallenge(challenge,enteteC))
                        f.write("\n\\end{longtable}\\end{center}\\end{document}")
                    f.close()
            except :
                print("Aucun résultat pour le challenge", challenge)

    ### générer les tex pour chaque groupement.
    ### VOIR SI C'EST ENCORE UTILE CI-DESSOUS
##    listeGroupements = listGroupementsCommences()
##    for groupement  in listeGroupements :
##        print("Groupement en cours de génération ",groupement.nom)
##        # ajout d'une sécurité si aucune arrivée dans une course
##        aCreer = False
##        #groupement = groupementAPartirDeSonNom(nomGroupement)
##        if groupement :
##            for nomCategorie in groupement.listeDesCourses :
##                if nomCategorie in ResultatsGroupements.keys() :
##                    aCreer = True
##                    break
##            if aCreer :
##                nomFichier = groupement.nom.replace(" ","-")
##                with open(TEXDIR+nomFichier+ ".tex", 'w') as f :
##                    f.write(creerFichierCategories(groupement,entete))
##                    f.write("\n\\end{longtable}\\end{center}\\end{document}")
##                f.close()

    # pour chaque fichier dans impressions , compiler.
    compilerTousLesTex(TEXDIR, "impressions")


def compilerTousLesTex(TEXDIR, PDFDIR):
    """compile tous les fichiers tex présents dans le dossier TEXDIR spécifié en paramètre. Compile dans PDFDIR."""
    liste_fichiers_tex_complete=glob.glob(TEXDIR+'*.tex',recursive = False)
    for file in liste_fichiers_tex_complete :
        # il faut compiler tous les fichiers de la liste.
        fichierACompiler = os.path.basename(file)
        compilateurComplete = compilateur.replace("@dossier@",PDFDIR)
        #print("on compile: ---"+compilateurComplete + "--- "+ TEXDIR + "---" + fichierACompiler )
        print(compiler(compilateurComplete, TEXDIR, fichierACompiler , 1))
    #os.chdir(osCWD)
    print("Tous les PDF de " + TEXDIR + " ont été générés dans " + PDFDIR + ".")

def listeDesCategoriesDUnGroupement(nomGroupement):
    retour = []
    for groupement in Groupements :
        if groupement.nomStandard == nomGroupement :
            retour = groupement.listeDesCourses
            break
    return retour

def groupementAPartirDeSonNom(nomGroupement, nomStandard=True):
    """ retourne un objet groupement à partir de son nom"""
    retour = None
    ### j'aurais dû choisir un dictionnaire mais, en l'état, je ne change pas un ensemble qui fonctionne.
    ### il me faudrait changer tous les appels aux groupements partout dans le code pour optimiser.
    for groupement in Groupements :
        if nomStandard :
            if groupement.nomStandard == nomGroupement :
                retour = groupement
                break
        else :
            if groupement.nom == nomGroupement :
                retour = groupement
                break
    return retour

def ancienGroupementAPartirDUneCategorie(categorie):
    """ retourne un objet groupement à partir d'un nom de catégorie
    Non optimisé. Cette version est maintenue pour permettre l'import d'anciennes sauvegardes
    où les Courses n'avaient pas la propriété nomGroupement. Cette fonction, utilisée une seule fois, permet de la construire."""
    ### cette recherche avait du sens avant la création de la propriété nomGroupement de l'object Course. Tenu à jour, cela permet une optimisation
    retour = None
    for groupement in Groupements :
        for cat in groupement.listeDesCourses :
            if cat == categorie :
                retour = groupement
                break
    #print("categorie cherchée :", categorie)
    #print("Groupement trouvé :", groupement.nom, groupement.listeDesCourses)
    #print(categorie, "est dans",retour.affichageInfoTerminal())
    return retour

def nomGroupementAPartirDUneCategorie(categorie, nomStandard = True):
    """ retourne un str nom du groupement à partir d'un nom de catégorie"""
    return groupementAPartirDUneCategorie(categorie).nomStandard
    # try :
        # retour = Courses[categorie].nomGroupement ### compatibilité avec les anciennes sauvegardes sans cette propriété.
    # except :
        # try :
            # retour = Courses[categorie].initNomGroupement(categorie)
        # except:
            # retour = ""
    # #print("categorie",Courses[categorie].categorie ,retour)
    # return retour

def groupementAPartirDUneCategorie(categorie):
    """ retourne un objet groupement à partir d'un nom de catégorie ou de course"""
    ### cette recherche avait du sens avant la création de la propriété nomGroupement de l'object Course. Tenu à jour, cela permet une optimisation
    # retour = None
    # for groupement in Groupements :
        # for cat in groupement.listeDesCourses :
            # if cat == categorie :
                # retour = groupement
                # break
    #print("categorie cherchée :", categorie)
    #print("Groupement trouvé :", groupement.nom, groupement.listeDesCourses)
    #print(categorie, "est dans",retour.affichageInfoTerminal())
    try :
        #print("Croupements",Groupements)
        #print("nomGroupement",Courses[categorie].nomGroupement)
        retour = Groupements[findIndex(Courses[categorie].nomGroupement, Groupements)] ### compatibilité avec les anciennes sauvegardes sans cette propriété.
    except :
        retour = Groupements[findIndex(Courses[categorie].initNomGroupement(categorie), Groupements)]# Courses[categorie].initNomGroupement(categorie)
    return retour

def findIndex(nom, L):
    '''Find index of nom in liste de groupements L where it is present'''
    i = 0
    for el in L :
        #print("on cherche", nom, "et on trouve", el.nomStandard)
        if el.nomStandard == nom :
            break
        i += 1
    return i

def nettoieGroupements() :
    """ supprime les listes vides de Groupements, créées par l'interface si un utilisateur décide d'effectuer des regroupements non incrémentés à partir de 1."""
    i = len(Groupements) - 1
    while i >= 0 :
        if not Groupements[i].listeDesCourses :
            del Groupements[i]
        i -= 1

def updateNomGroupement(nomStandard, nomChoisi) :
    groupementAPartirDeSonNom(nomStandard).setNom(nomChoisi)

def updateGroupements(categorie, placeInitiale, placeFinale):
    print("Mise à jour de Groupements : ",categorie,"déplacée du groupement ", placeInitiale, "vers le",placeFinale)
    # print("Groupements initial :")
    # for grp in Groupements :
        # print(grp.listeDesCourses)
        # print(grp.nom)
    if placeInitiale != placeFinale :
        Groupements[placeInitiale-1].removeCourse(categorie)
        if placeFinale - 1 == len(Groupements) :
            Groupements.append(Groupement(categorie,[categorie]))
        else :
            Groupements[placeFinale-1].addCourse(categorie)
        nettoieGroupements()
        #print("Groupements final :")
        #for grp in Groupements :
        #    print(grp.nom,grp.listeDesCourses)

def supprimeCourseDuGroupementEtNettoieGroupements(course):
    """ Utilisé uniquement lors des courses manuelles où des courses sont identiques aux groupements. Les courses sont créées manuellement
    et peuvent exister sans coureur dedans, ce qui n'arrive jamais sinon."""
    i = 0
    for g in Groupements :
        if course in g.listeDesCourses :
            g.removeCourse(course)
            del Groupements[i]
        i += 1

def absentsDispensesAbandonsEnTex() :
    Labs = []
    Ldisp = []
    Labandon = []
    #print(Resultats)
    for c in Coureurs.liste() :
        if c.absent :
            Labs.append(c)
            #print(c.nom, c.prenom, "était absent.")
        elif c.dispense :
            Ldisp.append(c)
            #print(c.nom, c.prenom, "est dispensé")
        elif c.rang == 0 and Courses[c.course].depart : # si la course a été lancé et le coureur n'est pas arrivé.
            Labandon.append(c)
            #print(c.nom, c.prenom, "a abandonné")
    print("nbre abs", len(Labs), "  nbre disp",len(Ldisp), "  nbre abandons", len(Labandon))
    retour = """\n\n\\newpage\n
{} \\hfill  {\LARGE ABSENTS}  \\hfill {}\n

\\begin{multicols}{3}
\\begin{itemize}
"""
    for c in Labs:
        retour += "\\item[$\\bullet$]" + c.nom + " " + c.prenom + " (" 
        if Parametres["CategorieDAge"] == 2 : 
            retour += c.etablissement
        elif Parametres["CategorieDAge"] == 1 : 
            retour += c.categorie(CategorieDAge)
        else :
            retour += c.classe 
        retour += ")"
    retour +="\\end{itemize}\\end{multicols}\n\n"
    retour += """\\bigskip

{} \\hfill  {\LARGE DISPENSES}  \\hfill {}

\\begin{multicols}{3}
\\begin{itemize}
"""
    for c in Ldisp:
        retour += "\\item[$\\bullet$]" + c.nom + " " + c.prenom + " (" 
        if Parametres["CategorieDAge"] == 2 : 
            retour += c.etablissement
        elif Parametres["CategorieDAge"] == 1 : 
            retour += c.categorie(CategorieDAge)
        else :
            retour += c.classe 
        retour += ")"
    retour +="\\end{itemize}\\end{multicols}\n\n"
    retour += """\\bigskip

{} \\hfill  {\LARGE ABANDONS (ceux sans temps, ni absents, ni dispensés)}  \\hfill {}

\\begin{multicols}{3}
\\begin{itemize}
"""
    for c in Labandon:
        retour += "\\item[$\\bullet$]" + c.nom + " " + c.prenom + " (" 
        if Parametres["CategorieDAge"] == 2 : 
            retour += c.etablissement
        elif Parametres["CategorieDAge"] == 1 : 
            retour += c.categorie(CategorieDAge)
        else :
            retour += c.classe 
        retour += ")"
    retour +="\\end{itemize}\\end{multicols}\n\n"
    return retour

def syscmd(cmd, encoding=''):
    """
    Runs a command on the system, waits for the command to finish, and then
    returns the text output of the command. If the command produces no text
    output, the command's return code will be returned instead.
    """
    p = subprocess.Popen(cmd, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        close_fds=True)
    p.wait()
    output = p.stdout.read()
    if len(output) > 1:
        if encoding: return output.decode(encoding)
        else: return output
    return p.returncode

def compiler(compilateur, chemin, fichierACompiler, nombre) :
    """ compilation latex avec le compilateur de son choix et suppression des fichiers de compilation"""
    #print("arguments :" , compilateur, chemin, fichierACompiler, nombre)
    retour = ""
    #with open("out.txt", mode='w') as f :
    for i in range(nombre) :
        #os.chdir(chemin)
        compilateurComplete = compilateur.replace("@dossier@",chemin)
        cmd = compilateurComplete +  fichierACompiler
        #subprocess.run(cmd, stdout=f)
        print("Exécution de", cmd)
        syscmd(cmd)
        #os.system(cmd)
        retour ="Compilation du fichier " + fichierACompiler + " effectuée."
        #"Compilation avec " + compilateurComplete + " du fichier généré " + fichierACompiler + " effectuée."
    # suppression des fichiers de compilation.
    aSupprimer = ""
    ### uniquement pour ce programme car la compilation s'effectue toujours dans le dossier .. Nettoyage du dossier chemin et de sa racine pour gérer tous les cas.
    for ch in [ chemin , chemin + os.sep + ".." ] :
        for ext in [".aux" , ".log", ".out" , ".synctex.gz", ".hd"] :
            if os.path.exists(os.path.join(ch,fichierACompiler[:-4] + ext)) :
                #retour += "Suppression de " + os.path.join(ch,fichierACompiler[:-4] + ext)
                os.remove(os.path.join(ch,fichierACompiler[:-4] + ext))#  + fichierACompiler[:-4] + ".log " + fichierACompiler[:-4] + ".out " + fichierACompiler[:-4] + ".synctex.gz "
    # affichage dans le lecteur de pdf par défaut.
    #cmd = "open " + fichierACompiler[:-4] + ".pdf"
    #os.system(cmd)
    #subprocess.call(cmd, stdout=f)
    #f.close()
    return retour

def compilerDossards(compilateur, chemin, fichierACompiler, nombre) :
    # TEMPORAIRE EN ATTENDANT DE TROUVER POURQUOI LA COMPILATION PLANTE AVEC syscmd() (qui a le mérite d'éviter toute sortie à l'écran.
    """ compilation latex avec le compilateur de son choix et suppression des fichiers de compilation"""
    for DIR in [ "dossards", "dossards"+os.sep+"tex" ]:
        if not os.path.exists(DIR) :
            os.makedirs(DIR)
    #print("arguments :" , compilateur, chemin, fichierACompiler, nombre)
    retour = ""
    #with open("out.txt", mode='w') as f :
    for i in range(nombre) :
        #os.chdir(chemin)
        cmd = compilateur + fichierACompiler
##        if DEBUG :
##            cmd = cmd + "& pause"
        #subprocess.Popen(cmd, shell=False)
        print(cmd)
        syscmd(cmd)
        ### subprocess.call(("cmd.exe", cmd), shell=False)
        ##### FONCTIONNEL mais visible :
        #os.system(cmd)
        retour = "Compilation avec " + compilateur + " du fichier généré " + fichierACompiler + " effectuée."
    # suppression des fichiers de compilation.
    aSupprimer = ""
    for ext in [".aux" , ".log", ".out" , ".synctex.gz"] :
        fichierCherche = os.path.join("dossards",fichierACompiler[:-4] + ext)
        #print(fichierCherche, "cherché")
        if os.path.exists(fichierCherche) :
            #print("Suppression de ",fichierCherche)
            os.remove(fichierCherche)#  + fichierACompiler[:-4] + ".log " + fichierACompiler[:-4] + ".out " + fichierACompiler[:-4] + ".synctex.gz "
    return retour

def dossardPrecedentDansArriveeDossards(dossard):
    """ retourne le dossard précédent l'argument dans arriveeDossards
    retourne 0 si dossard est le premier de la liste
    retourne -1 si le dossard n'existe pas dans arriveeDossards"""
    if len(ArriveeDossards)==0:
        print("Pas de dossards arrivés.")
        return "-1"
    try :
        n = ArriveeDossards.index(formateDossardNG(dossard))
        if n == 0 :
            return "0"
        else :
            return formateDossardNG(ArriveeDossards[n-1])
    except:
        return "-1"

def dossardSuivantDansArriveeDossards(dossard):
    """ retourne le dossard suivant l'argument dans arriveeDossards
    retourne 0 si dossard est le dernier de la liste
    retourne -1 si le dossard n'existe pas dans arriveeDossards"""
    if len(ArriveeDossards)==0:
        print("Pas de dossards arrivés.")
        return "-1"
    try :
        n = ArriveeDossards.index(formateDossardNG(dossard))
        if n+1 == len(ArriveeDossards) :
            return "0"
        else :
            return formateDossardNG(ArriveeDossards[n+1])
    except:
        return "-1"

# def listCoureurs():
    # if len(Coureurs)==0:
        # print("There are no Coureurs.")
        # return
    # for coureur in Coureurs :
        # print(coureur.dossard, coureur.nom, coureur.prenom, coureur.sexe, coureur.naissance, coureur.classe, coureur.absent, \
              # coureur.dispense, coureur.tempsFormate(), coureur.rang, coureur.commentaireArrivee, coureur.vitesseFormateeAvecVMA())

def listArriveeDossards():
    if len(ArriveeDossards)==0:
        print("Pas de dossards arrivés.")
        return
    print(ArriveeDossards)

def listArriveeTemps() :
    if len(ArriveeTemps)==0:
        print("Pas de temps saisi.")
        return
    i = 0
    for tps in ArriveeTemps:
        #tests tps.setDossardProvisoire(None)
        if ArriveeTempsAffectes[i] == "0" :
            print(str(i) + "." , tps.tempsCoureur , "s (tempsReel = ", tps.tempsReel ,")")
        else :
            print(str(i) + "." , tps.tempsCoureur , "s (tempsReel = ", tps.tempsReel ,")- affecté manuellement au dossard", ArriveeTempsAffectes[i] )
        i += 1
    ##transaction.commit()


### Fonctions métiers pour gérer les temps et résultats des courses.


def formateTemps(temps):
    partieDecimale = str(round(((temps - int(temps))*100)))
    if len(partieDecimale) == 1 :
        partieDecimale = "0"+partieDecimale
    ch = time.strftime("%M min %S s ",time.localtime(temps)) + partieDecimale
    return ch

#print(formateTemps(124.715563659667969))

def generateResultatsChallenge(nom,listeOrdonneeParTempsDesDossardsDeLaClasse,nbreDeCoureursPrisEnCompte):
    score = 0
    nf = 0
    ng = 0
    i = 0
    listeCG = []
    listeCF = []
    #listeDesRangs = []
    nbreDeCoureursPrisEnCompte = int(nbreDeCoureursPrisEnCompte)
    while (ng < nbreDeCoureursPrisEnCompte or nf < nbreDeCoureursPrisEnCompte) and i < len(listeOrdonneeParTempsDesDossardsDeLaClasse):
        doss = listeOrdonneeParTempsDesDossardsDeLaClasse[i]
        coureur = Coureurs.recuperer(doss)
        if coureur.temps != 0 :
            if ng < nbreDeCoureursPrisEnCompte and coureur.sexe == "G" :
                #print("le coureur",coureur.prenom,"est en rang",coureur.rang," ng=",ng)
                listeCG.append(coureur)
                score += coureur.rang
                ng += 1
            if nf < nbreDeCoureursPrisEnCompte and coureur.sexe == "F" :
                #print("le coureur",coureur.prenom,"est en rang",coureur.rang," nf=",nf)
                listeCF.append(coureur)
                score += coureur.rang
                nf += 1
            #listeDesRangs.append(coureur.rang)
        i+=1
##    scoreNonPondere = score
##    if ng + nf < 2*nbreDeCoureursPrisEnCompte :
##        # correctif si le nombre nbreDeCoureursPrisEnCompte n'est pas atteint à l'arrivée.
##        complet = False
##        if Parametres["ponderationAcceptee"] :
##            print("Application d'une pondération", 2*nbreDeCoureursPrisEnCompte, "/", ng + nf ," à la classe", nom, "pour cause d'un nombre insuffisant de coureurs à l'arrivée :",ng + nf)
##            score = score * 2*nbreDeCoureursPrisEnCompte / (ng + nf)
##    else :
##        complet = True
    #print(nom, score, listeCG, listeCF)
    return EquipeClasse(nom, listeCG, listeCF, Parametres["ponderationAcceptee"])

def generateResultatsChallengeUNSS(nom,listeOrdonneeParScoreDesDossardsDeLaClasse):
    #print("nom cat-etab",nom,listeOrdonneeParScoreDesDossardsDeLaClasse)
    nbreDeCoureursNecessairesParEquipe = 5
    tousLesGars = []
    toutesLesFilles = []
    listeFSelect = []
    listeGSelect = []
    unCoureurCategorieLimiteDejaSelectionne = False
    complet = False
    listeDesEquipes = []
    # on sépare les filles des garçons dans listeCG et listeCF.
    if listeOrdonneeParScoreDesDossardsDeLaClasse :
        # paramétrage en fonction du type de challenge : collège, LGT ou LP
        doss = listeOrdonneeParScoreDesDossardsDeLaClasse[0]
        coureur = Coureurs.recuperer(doss)
        # si le premier coureur est en lycée pro, tous le sont (normalement). Sinon, le problème est en amont.
        if coureur.etablissementNature and coureur.etablissementNature.upper() == "LP" :
            nbreDeFillesNecessairesParEquipe = 0
            nbreDeGarsNecessairesParEquipe = 0
            nbreMaxdUnSexe = 5
            categoriesLimitees = []# [ "M10","M9","M8","M7", "M6","M5", "M4","M3" ,"M2", "M1" ,"M0" , "SE" ,"ES"]
        else :
            nbreDeFillesNecessairesParEquipe = 2
            nbreDeGarsNecessairesParEquipe = 2
            if coureur.etablissementNature and coureur.etablissementNature[0].upper() == "L" : # le challenge est en lycée GT
                nbreMaxdUnSexe = 3
                categoriesLimitees = [ "M10","M9","M8","M7", "M6","M5", "M4","M3" ,"M2", "M1" ,"M0" , "SE" ,"ES"]
            else : # sinon, on est en collège, la catégorie limitée est alors les cadets et supérieurs.
                nbreMaxdUnSexe = 3
                categoriesLimitees = [ "M10","M9","M8","M7", "M6","M5", "M4","M3" ,"M2", "M1" ,"M0" , "SE" ,"ES", "JU" , "CA" ]
        i = 0
        ### parcours de la liste ordonnée fournie pour la séparer par sexe
        while i < len(listeOrdonneeParScoreDesDossardsDeLaClasse):
            doss = listeOrdonneeParScoreDesDossardsDeLaClasse[i]
            coureur = Coureurs.recuperer(doss)
            if coureur.temps > 0 :
                if coureur.sexe == "F" :
                    toutesLesFilles.append(coureur)
                else : # c'est un gars !
                    tousLesGars.append(coureur)
            i+=1
         # tant qu'il y a au moins un coureur, on parcourt les deux listes en choisissant le meilleur garçon ou la meilleur fille (en terme de scoreUNSS)
        i=0
        while toutesLesFilles or tousLesGars :
             ## on doit trouver le prohcain coureur à selectionner : celui qui a le scoreeUNSS le plus bas.
             ## ETAPE DE SELECTION
             coureurSelectionne = None
             if len(listeGSelect) >= nbreMaxdUnSexe and toutesLesFilles :
                 # il manque uniquement des filles et il y a des filles
                 # on prend prioritairement dans la reserveF (ceux qui sont d'une catégorie plus élevée) et sinon dans la listeCF.
                 ind = indicePremierCoureurAutoriseUNSS(toutesLesFilles, categoriesLimitees, unCoureurCategorieLimiteDejaSelectionne)
                 if ind != None :
                    coureurSelectionne = toutesLesFilles.pop(ind)
             elif len(listeFSelect) >= nbreMaxdUnSexe and tousLesGars :
                 # il manque uniquement des garçons et il y a des garçons
                 # on prend prioritairement dans la reserveG (ceux qui sont d'une catégorie plus élevée) et sinon dans la listeCG.
                 ind = indicePremierCoureurAutoriseUNSS(tousLesGars, categoriesLimitees, unCoureurCategorieLimiteDejaSelectionne)
                 if ind != None :
                    coureurSelectionne = tousLesGars.pop(ind)
             else :
                 # on est libre de choisir un gars ou une fille.
                 indiceMeilleurFilleAutorisee = indicePremierCoureurAutoriseUNSS(toutesLesFilles, categoriesLimitees, unCoureurCategorieLimiteDejaSelectionne)
                 indiceMeilleurGarsAutorise = indicePremierCoureurAutoriseUNSS(tousLesGars, categoriesLimitees, unCoureurCategorieLimiteDejaSelectionne)
                 if indiceMeilleurFilleAutorisee != None and indiceMeilleurGarsAutorise != None : # aucun des deux coureurs cherché n'est None.
                     # il reste des filles et des garçons : on prend le meilleur suivant dans l'un des deux sexes.
                     if toutesLesFilles[indiceMeilleurFilleAutorisee].scoreUNSS <  tousLesGars[indiceMeilleurGarsAutorise].scoreUNSS :
                         ## on prélève la meilleure fille à l'équipe
                         coureurSelectionne = toutesLesFilles.pop(indiceMeilleurFilleAutorisee)
                     else :
                         coureurSelectionne = tousLesGars.pop(indiceMeilleurGarsAutorise)
                         ## on affecte le meilleur gars dans l'équipe
                 else :
                     if indiceMeilleurFilleAutorisee != None : # il reste au moins une fille sélectionnable mais plus de gars
                         coureurSelectionne = toutesLesFilles.pop(indiceMeilleurFilleAutorisee)
                     elif indiceMeilleurGarsAutorise != None : # L'inverse du test ci-dessus.
                         coureurSelectionne = tousLesGars.pop(indiceMeilleurGarsAutorise)
              
             # les deux listes ne contiennent plus de coureur autorisé (les deux indices sont None. Il est nécessaire de stopper la boucle !            
             if coureurSelectionne == None :
                # si l'une des étapes amène à sélectionner None, c'est qu'il n'y a plus de coureur qui convienne pour compléter
                # l'équipe en cours, qui restera donc incomplète.
                 break

             #print(coureurSelectionne.nom, "sélectionné alors que unCoureurCategorieLimiteDejaSelectionne = ",unCoureurCategorieLimiteDejaSelectionne) 
            ## ETAPE AVANT DE BOUCLER (le coureurSelectionne à ce stade est différent de None.
            # le coureur actuellement ajouté est sous-classé : un seul par équipe autorisé.
             #if DEBUG and coureurSelectionne.categorieSansSexe() != coureurSelectionne.categorieFFA() :
             #    print(coureurSelectionne.nom, "est inscrit en", coureurSelectionne.categorieSansSexe(),"mais de catégorie réelle", coureurSelectionne.categorieFFA())
             if coureurSelectionne.categorieFFA() in categoriesLimitees : # on regarde la catégorie réelle du coureur et non celle générée par la méthode coureur.categorie(...)
                 #print(coureurSelectionne.nom, " est de catégorie réelle", coureurSelectionne.categorieFFA()," et court avec des élèves de catégorie inférieure. On empêche la présence d'un autre dans ce cas pour l'équipe actuelle")
                 unCoureurCategorieLimiteDejaSelectionne = True
            # on alimente l'équipe avec celui sélectionné.
             if coureurSelectionne.sexe == "F" :
                 listeFSelect.append(coureurSelectionne)
             else :
                 listeGSelect.append(coureurSelectionne)
             if len(listeGSelect) + len(listeFSelect) >= nbreDeCoureursNecessairesParEquipe and len(listeGSelect) >= nbreDeGarsNecessairesParEquipe and len(listeFSelect)  >= nbreDeFillesNecessairesParEquipe :
                # si une équipe est complète, on réinitialise les variables.
                 #print("Création d'une équipe pour le challenge",nom, [cour.nom for cour in listeGSelect], [cour.nom for cour in listeFSelect])
                 listeDesEquipes.append(EquipeClasse(nom, listeGSelect, listeFSelect))
                 unCoureurCategorieLimiteDejaSelectionne = False
                 listeGSelect = []
                 listeFSelect = []
    return listeDesEquipes


def indicePremierCoureurAutoriseUNSS(listeDeCoureurs, categoriesInterdites, unCoureurCategorieLimiteDejaSelectionne) :
    ''' la liste de coureurs fournie, une liste de catégories interdites , est ce que cette liste doit être prise en compte ou pas via le boolean unCoureurCategorieLimiteDejaSelectionne
    Retourne l'indice du premier coureur autorisé de la listeDeCoureurs fournie.
    Peut retourner None si aucun coureur n'est trouvé '''
    if unCoureurCategorieLimiteDejaSelectionne :
        retour = None
        i = 0
        for c in listeDeCoureurs :
            if not c.categorieFFA() in categoriesInterdites :
                retour = i
                break
            i += 1
    else :
        if listeDeCoureurs :
            retour = 0
        else :
            retour = None
    return retour

def getDecompteParCategoriesDAgeEtRetourneTotal(catFFA , DecompteParCategoriesDAge, sexe) :
    # en théorie inutile puisque toutes les catégories sont présentes. rangDansCategorie = 0
    present = False
    for L in DecompteParCategoriesDAge : # on fait le test pour les petites catégories puis pour les vétérans.
        # on considère que les séniors doivent être meilleurs que toutes les autres catégories (au dessus et en dessous).
        ## tester d'abord si la catégorie existe dans la liste.
        if not present :
            for el in L :
                if el[0] == catFFA:
                    present = True
                    if sexe == "F" :
                        return el[2]
                    else :
                        return el[1] 
                    break
    if not present and Parametres["CategorieDAge"] : # pour le cross du collège, les catégories FFA sont inutiles.
        print("ANORMAL : on devrait toujours trouver un nombre de coureurs par catégorie")


    
ResultatsPourImpressions = {}
ResultatsGroupementsPourImpressions = {}

def genereResultatsCoursesEtClasses(premiereExecution = False) :
    """ procédure mettant à jour le dictionnaire Résultats et le rang de chaque coureur dans sa course"""
    global tableauGUI
    #Tdebut = time.time()
    retour = calculeTousLesTemps(premiereExecution)
    #print("calculeTousLesTemps : ",retour)
    Resultats.clear()
    ResultatsPourImpressions.clear()
    ResultatsGroupements.clear()
    ResultatsGroupementsPourImpressions.clear()
    #listeDesClasses = []
    # ETAPE 1 : on alimente Resultats avec les coureurs des classes (cross du collège), des catégories d'athlétisme, ou des catégories par établissement (UNSS)
    for coureur in Coureurs.liste() :
        doss = coureur.dossard
        cat = coureur.categorie(Parametres["CategorieDAge"])
        if CoursesManuelles :
            groupement = coureur.course
        else :
            groupement = nomGroupementAPartirDUneCategorie(cat)
        #print(groupement)
        classe = coureur.classe
        etab = coureur.etablissement
        ### ajout du coureur au groupement pour résultat du groupement.
        if groupement not in ResultatsGroupements :
            ResultatsGroupements[groupement] = []
            ResultatsGroupementsPourImpressions[groupement] = []
        if coureur.temps > 0  : #si pas d'erreur, on l'ajoute not coureur.absent and not coureur.dispense and coureur.temps != -1 and coureur.temps != 0 :
            ResultatsGroupements[groupement].append(doss)
        ResultatsGroupementsPourImpressions[groupement].append(doss)
##        else :
##            coureur.setRang(0)

        ### ajout du coureur dans sa classe ou sa catégorie d'age.
        if Parametres["CategorieDAge"] == 0 :
            if classe not in Resultats :
                ResultatsPourImpressions[classe] = []
                Resultats[classe] = []
            #if classe not in listeDesClasses : ### raison d'être de cette liste à trouver ! Encore utile ? Je ne pense pas : commenté.
            #    listeDesClasses.append(classe)
            if coureur.temps != -1 :
                # les coureurs au temps -1 sont abs, disp ou abandons donc on doit les mettre  not coureur.absent and not coureur.dispense and coureur.temps != 0 :
                Resultats[classe].append(doss)
            ResultatsPourImpressions[classe].append(doss)
        elif Parametres["CategorieDAge"] == 1 :
            if cat not in Resultats :
                Resultats[cat] = []
                ResultatsPourImpressions[cat] = []
            if coureur.temps != -1 : # les coureurs au temps nul sont abs, disp ou abandons donc on doit les mettre  not coureur.absent and not coureur.dispense and coureur.temps != -1 and coureur.temps != 0 :
                Resultats[cat].append(doss)
            ResultatsPourImpressions[cat].append(doss)
        elif Parametres["CategorieDAge"] == 2 :
            if coureur.etablissementNature and coureur.etablissementNature[0].upper() == "L" : # cas du challenge UNSS lycée qui mélange tous les lycéens !
                ### cas particulier du challenge UNSS : on ajoute les résultats des LP et des LG même s'il n'y a aucun coureur.
                ### Trop galère de tout changer sachant que les challenges ne correspondent pas à des catégories d'athlétisme du logiciel
                if "LP-"  + str(etab) not in Resultats :
                    Resultats["LP-" + str(etab)] = []
                    ResultatsPourImpressions["LP-" + str(etab)] = []
                if "LG-" + str(etab) not in Resultats :
                    Resultats["LG-" + str(etab)] = []
                    ResultatsPourImpressions["LG-" + str(etab)] = []
                if coureur.etablissementNature.upper() == "LP" : # c'est un lycée pro.
                    nomDuResultat = "LP-" + str(etab)
                else :
                    nomDuResultat = "LG-" + str(etab)
            else : # on est en collège : le nom du résultat est constitué de la cat sur deux caractères
                nomDuResultat = str(coureur.categorieSansSexe()) + "-" + str(etab) # le challenge sera calculé entre les résultats de MI-Bourrillon, MI-Gevaudan, etc... Basé sur l'initiale de la catégorie.
            if nomDuResultat not in Resultats :
                Resultats[nomDuResultat] = []
                ResultatsPourImpressions[nomDuResultat] = []
            if coureur.temps > 0 :
                # les coureurs au temps nul sont abs, disp ou abandons donc on doit les mettre  not coureur.absent and not coureur.dispense and coureur.temps != 0 :
                Resultats[nomDuResultat].append(doss)
                #print("Dossard ajouté ", doss,"dans Resultats[",nomDuResultat,"]")
                #print(Resultats)
            ResultatsPourImpressions[nomDuResultat].append(doss)
            #else :
            #    print("Dossard au temps négatif ignoré", doss)
        # Finalement, on ne parcourt qu'une liste ci-dessus (tout le début commenté) et on trie tout ensuite. Sûrement plus rapide.
    ## ETAPE 2 : on alimente ResultatsGroupements, on affecte les rangs aux coureurs en fonction de leur rang d'arrivée dans le Groupement.
    #### A SEPARER SOUS FORME D'UNE FONCTION EXECUTEE DANS PLUSIEURS THREADS=> gain de temps pour les tris sur plusieurs coeurs
    ### on traite les rangs dans les Groupements
    #keyList = []
    for nom in ResultatsGroupements :
        groupementAPartirDeSonNom(nom,nomStandard = True).initEffectifs()
        # on considère que la meilleure catégorie est SENIOR.
        L1 = [ ["SE",0,0 ], ["ES",0,0 ], ["JU",0,0 ], ["CA",0,0 ], ["MI",0,0 ], ["BE",0,0 ], ["PO",0,0 ], ["EA",0,0 ], ["BB",0,0 ]]
        L2 = [["SE",0,0 ] , ['M0', 0,0], ['M1', 0,0], ['M2', 0,0], ['M3', 0,0], ['M4', 0,0], ['M5', 0,0], ['M6', 0,0], ['M7', 0,0], ['M8', 0,0], ['M9', 0,0], ['M10', 0,0]]
        DecompteParCategoriesDAge = [L1, L2]
        # rang par sexes
        RangSexe = [0,0]
        #keyList.append(nom)
        ResultatsGroupements[nom] = triParTemps(ResultatsGroupements[nom])
        # on affecte son rang à chaque coureur dans sa Course (et son score UNSS)
        #print("course ",nom,":",Resultats[nom])
        ### inutile car obligatoire vu ce qui précède : if estUnGroupement(nom) :
            #print(nom, "est une course ou un groupement",Resultats[nom])
        i = 0
        nbreArriveesGroupement = len(ResultatsGroupements[nom])
        #print("Groupement", nom, "NbreArrivéesTotal", nbreArriveesGroupement, ResultatsGroupements[nom])
        while i < nbreArriveesGroupement :
            doss = ResultatsGroupements[nom][i]
            coureur = Coureurs.recuperer(doss)
            groupementAPartirDeSonNom(nom,nomStandard = True).evolutionDUnAuxEffectifsTotaux(coureur)
            #print("coureur",coureur.nom,"(",doss,")",coureur.tempsFormate(),coureur.temps)
            if coureur.temps > 0 :
            ### si le coureur doit apparaître dans le tableau des résultats, on lui affecte un rang
                coureur.setRang(i+1)
                if coureur.sexe == "F" :
                    iSexe = 1 # rang dans la liste incrémentée.
                else :
                    iSexe = 0 # rang dans la liste incrémentée.
                RangSexe[iSexe] += 1
                ### cas du score UNSS si c'est un lycée : on affecte le score de la formule de calcul
                if Parametres["CategorieDAge"] == 2 :
                    coureur.setScoreUNSS(nbreArriveesGroupement) # on fournit le rang et le nombre total de coureurs arrivés dans le groupement.
                if Parametres["CategorieDAge"] : # cas où les catégories d'athlétisme sont utilisées (valeur 1 ou 2)
                    catFFA = coureur.categorieFFA()
                    coureur.setRangCat(incrementeDecompteParCategoriesDAgeEtRetourneSonRang(catFFA , DecompteParCategoriesDAge, coureur.sexe))
                    coureur.setRangSexe(RangSexe[iSexe])
            else : # inutile car les seuls coureurs dans Resultats sont ceux ayant un rang légitime vu le filtrage 10 lignes au dessus :
            # avec "if not coureur.absent and not coureur.dispense and coureur.temps != -1 and coureur.temps != 0"
                coureur.setRang(0)
                coureur.setRangCat(0)
                coureur.setRangSexe(0)
            #print("dossard",doss,"coureur",coureur.nom,coureur.tempsFormate(),coureur.rang)
            i += 1
        # on définit combien de coureurs appartiennent à un groupement.
        ### inutile car fait au fur et à mesure par groupementAPartirDeSonNom(nom,nomStandard = True).setNombreDeCoureursTotal(RangSexe[0], RangSexe[1])
    ### ETAPE 3 : On traite les rangs dans les classes ou cat-établissment (pour l'UNSS), on trie les coureurs d'une même catégorie et d'un même établissement par score.
    keyList = []
    for nom in Resultats :
        keyList.append(nom)
        Resultats[nom] = triParTemps(Resultats[nom])
        # # on affecte son rang à chaque coureur dans sa Course.
        # #print("course ",nom,":",Resultats[nom])
        # ### inutile car obligatoire vu ce qui précède : if estUneCourse(nom) :
        # i = 0
        # while i < len(Resultats[nom]) :
            # doss = Resultats[nom][i]
            # coureur = Coureurs[doss-1]
            # #print("coureur",coureur.nom,"(",doss,")",coureur.tempsFormate(),coureur.temps)
            # if coureur.temps > 0 :
            # ### si le coureur doit apparaître dans le tableau des résultats, on lui affecte un rang
                # coureur.setRangCat(i+1)
            # else : # inutile car les seuls coureurs dans Resultats sont ceux ayant un rang légitime vu le filtrage 10 lignes au dessus :
            # # avec "if not coureur.absent and not coureur.dispense and coureur.temps != -1 and coureur.temps != 0"
                # coureur.setRangCat(0)
            # #print("dossard",doss,"coureur",coureur.nom,coureur.tempsFormate(),coureur.rang)
            # i += 1
    #### POINT DE RENCONTRE DE TOUS LES THREADS (pas d'accès concurrant ni pour les tris, ni pour le rang de chaque coureur qui ne coure que dans une course..
    ## ETAPE 4 : on calcule les résultats du challenge par classe après que les deux catégories G et F soient triées => obligation de séparer.
    #print("ResultatsGroupements avant calcul des challenges :",ResultatsGroupements)
    if Parametres["CategorieDAge"] == 0 or Parametres["CategorieDAge"] == 2 : # challenge uniquement pour le cross du collège et pour l'UNSS
        L = []
        #print(keyList)
        for nom in keyList :
            Resultats[nom] = triParScoreUNSS(Resultats[nom])
            ### inutile car obligatoire désormais : if estUneClasse(nom) :
            if Parametres["CategorieDAge"] == 0 :
                challenge = nom[0]
            else :
                challenge = nom[:2]
            # création du challenge pour ce niveau, si inexistant
            if challenge not in ResultatsGroupements :
                ResultatsGroupements[challenge] = []
                L.append(challenge)
            # cas du cross du collège
            if Parametres["CategorieDAge"] == 0 :
                # on alimente le challenge avec une EquipeClasse
                equ = generateResultatsChallenge(nom, Resultats[nom], Parametres["nbreDeCoureursPrisEnCompte"])
                if Parametres["ponderationAcceptee"] :
                    ResultatsGroupements[challenge].append(equ)
                elif equ.complet :
                    ResultatsGroupements[challenge].append(equ)
            elif Parametres["CategorieDAge"] == 2 :
                Lequ = generateResultatsChallengeUNSS(nom, Resultats[nom])
                # le codage de generateResultatsChallengeUNSS ne génère que des équipes complètes.
                for equ in Lequ :
                    ResultatsGroupements[challenge].append(equ)
        # on trie chaque challenge par score.
        for challenge in L :
            ResultatsGroupements[challenge]=triParScore(ResultatsGroupements[challenge])
    return retour

def estUneCourseOuUnGroupement(nom):
    if nom in Courses :
        retour = True
    elif groupementAPartirDeSonNom(nom, nomStandard=False) != None :
        retour = True
    else :
        retour = False
    #print(nom,retour)
    return retour

def estUneCourse(nom):
    if nom in Courses :
        retour = True
    else :
        retour = False
    return retour

def estUnGroupement(nom):
    if groupementAPartirDeSonNom(nom, nomStandard=False) != None :
        retour = True
    else :
        retour = False
    return retour

def estUneClasse(nom):
    """ Comme le nom des classes est libre, estUneClasse est vraie si n'est pas une course ou un groupement ET n'est pas un challenge (1 caractère)"""
    return len(nom) > 1 and (not estUneCourseOuUnGroupement(nom))

def estSuperieurNPC(d1, d2):
    if d1.nom == d2.nom :
        # si les noms sont égaux, on compare les prénoms
        return d1.prenom > d2.prenom
    else :
        return d1.nom > d2.nom

def estSuperieurNP(d1, d2):
    if Coureurs.recuperer(d1).nom == Coureurs.recuperer(d2).nom :
        # si les noms sont égaux, on compare les prénoms
        return Coureurs.recuperer(d1).prenom > Coureurs.recuperer(d2).prenom
    else :
        return Coureurs.recuperer(d1).nom > Coureurs.recuperer(d2).nom

def estSuperieur(d1, d2):
    if Coureurs.recuperer(d1).temps == 0 or Coureurs.recuperer(d1).temps == -1 :
        # si le temps est nul, c'est que la ligne d'arrivée n'a pas été franchie. si -1 c'est que le départ n'a pas été donné
        return True
    elif Coureurs.recuperer(d2).temps == 0 or Coureurs.recuperer(d2).temps == -1:
        return False
    else :
        return Coureurs.recuperer(d1).temps > Coureurs.recuperer(d2).temps

def estSuperieurS(E1, E2):
    ''' on trie par EquipeClasse.score puis en fonction de listeDesRangs()'''
    if E1.score == 0 :
        # si le score est nul, c'est que le challenge n'est pas jouable pour cette équipe.
        return True
    elif E2.score == 0 :
        return False
    else :
        if E1.score == E2.score :
            try :
                retour = E1.listeDesRangs() > E2.listeDesRangs()
            except :
                retour = True # comparer deux objets None signifie qu'il n'y a pas de départage possible. Ne devrait pas arriver si on départage des équipeClasse complètes.
            return retour # E1.listeDesRangs() > E2.listeDesRangs()
        else :
            return E1.score > E2.score
            
def estSuperieurSUNSS(E1, E2):
    ''' on trie par coureur.scoreUNSS puis en fonction de coureur.rang'''
    if Coureurs.recuperer(E1).scoreUNSS == Coureurs.recuperer(E2).scoreUNSS :
        try :
            retour = Coureurs.recuperer(E1).rang > Coureurs.recuperer(E2).rang
        except :
            retour = True #pas de départage possible.
        return retour 
    else :
        return Coureurs.recuperer(E1).scoreUNSS > Coureurs.recuperer(E2).scoreUNSS

def triParNomPrenomCoureurs(listeDeCoureurs):
    return trifusionNPC(listeDeCoureurs)

def triParNomPrenom(listeDeDossard):
    return trifusionNP(listeDeDossard)

def triParTemps(listeDeDossard):
    return trifusion(listeDeDossard)

def triParScoreUNSS(listeDeCoureurs):
    return trifusionSUNSS(listeDeCoureurs)

def triParScore(listeDEquipes):
    return trifusionS(listeDEquipes)

def trifusionSUNSS(T) :
    if len(T)<=1 : return T
    T1=[T[x] for x in range(len(T)//2)]
    T2=[T[x] for x in range(len(T)//2,len(T))]
    return fusionSUNSS(trifusionSUNSS(T1),trifusionSUNSS(T2))

def trifusionS(T) :
    if len(T)<=1 : return T
    T1=[T[x] for x in range(len(T)//2)]
    T2=[T[x] for x in range(len(T)//2,len(T))]
    return fusionS(trifusionS(T1),trifusionS(T2))

def trifusionNP(T) :
    if len(T)<=1 : return T
    T1=[T[x] for x in range(len(T)//2)]
    T2=[T[x] for x in range(len(T)//2,len(T))]
    return fusionNP(trifusionNP(T1),trifusionNP(T2))

def trifusionNPC(T) :
    if len(T)<=1 : return T
    T1=[T[x] for x in range(len(T)//2)]
    T2=[T[x] for x in range(len(T)//2,len(T))]
    return fusionNPC(trifusionNPC(T1),trifusionNPC(T2))

def trifusion(T) :
    if len(T)<=1 : return T
    T1=[T[x] for x in range(len(T)//2)]
    T2=[T[x] for x in range(len(T)//2,len(T))]
    return fusion(trifusion(T1),trifusion(T2))

def fusionSUNSS(T1,T2) :
    if T1==[] :return T2
    if T2==[] :return T1
    if estSuperieurSUNSS(T2[0], T1[0]) :
        return [T1[0]]+fusionSUNSS(T1[1 :],T2)
    else :
        return [T2[0]]+fusionSUNSS(T1,T2[1 :])

def fusionS(T1,T2) :
    if T1==[] :return T2
    if T2==[] :return T1
    if estSuperieurS(T2[0], T1[0]) :
        return [T1[0]]+fusionS(T1[1 :],T2)
    else :
        return [T2[0]]+fusionS(T1,T2[1 :])

def fusionNP(T1,T2) :
    if T1==[] :return T2
    if T2==[] :return T1
    if estSuperieurNP(T2[0], T1[0]) :
        return [T1[0]]+fusionNP(T1[1 :],T2)
    else :
        return [T2[0]]+fusionNP(T1,T2[1 :])

def fusionNPC(T1,T2) :
    if T1==[] :return T2
    if T2==[] :return T1
    if estSuperieurNPC(T2[0], T1[0]) :
        return [T1[0]]+fusionNPC(T1[1 :],T2)
    else :
        return [T2[0]]+fusionNPC(T1,T2[1 :])

def fusion(T1,T2) :
    if T1==[] :return T2
    if T2==[] :return T1
    if estSuperieur(T2[0], T1[0]) :
        return [T1[0]]+fusion(T1[1 :],T2)
    else :
        return [T2[0]]+fusion(T1,T2[1 :])

def tempsClientIsNotInArriveeTemps(newTps) :
    """ retourne True si le tempsClient n'est pas présent dans ArriveeTemps."""
    retour = True
    i = len(ArriveeTemps)
    if i > 0 :
        tpsDejaPresent = ArriveeTemps[i-1]
        while i > 0 and tpsDejaPresent.tempsReel <= newTps.tempsReel :
            #print("comparaison de ",tpsDejaPresent.tempsCoureur, " et ",  newTps.tempsCoureur )
            if tpsDejaPresent.tempsReel == newTps.tempsReel :
                print("Temps déjà ajouté à l'arrivée. Temps Coureur sur téléphone :", newTps.tempsCoureur, ". Temps réel sur serveur:", newTps.tempsReel,"(non ajouté)." )
                retour= False
                #break
            #if  tpsDejaPresent.tempsReel + 60 < newTps.tempsReel :
    ##        if  tpsDejaPresent.tempsReel < newTps.tempsReel :
    ##            print("Temps supérieur de plus de 60 s par rapport à celui examiné. On stoppe la recherche.")
    ##            break
            i -= 1
            tpsDejaPresent = ArriveeTemps[i-1]
    return retour

def dupliqueTemps(tps):
    """ argument : une instance de la classe Temps.
        Retourne un temps disponible dans la liste des temps arrivés (le même temps que celui fourni plus quelques centièmes (en tempsReel calculé)."""
    nouveauTps = tps
    if tempsClientIsNotInArriveeTemps(nouveauTps) :
        #print("Un nouveau temps disponible a été trouvé. On le retourne :",nouveauTps.tempsReelFormateDateHeure())
        return nouveauTps
    else :
        print("Le temps",nouveauTps.tempsReelFormateDateHeure(),"existe déjà. On ajoute une nouvelle milliseconde tant que l'on ne trouve pas un temps disponible.")
        nouveauTps = tps.tempsPlusUnCentieme()
        return dupliqueTemps(nouveauTps)


def addArriveeTemps(tempsCoureur, tempsClient, tempsServeur, dossard="0A") :
    """ ajoute un temps dans la liste par ordre croissant pour que ArriveeTemps reste toujours croissante
    par rapport au tempsReel (heure d'arrivée du coureur sur le serveur (pour gérer plusieurs clients en décalage horaire).
    Doit prendre garde au temps mesuré sur le smartphone pour qu'un temps ne soit pas importé deux fois."""
    CodeRetour = Erreur(201)
    newTps = dupliqueTemps(Temps(tempsCoureur, tempsClient, tempsServeur))
##    if tempsClientIsNotInArriveeTemps(newTps) :
    doss = formateDossardNG(dossard)
    n = len(ArriveeTemps)
    if n == 0 : # si c'est le premier temps, on l'ajoute.
        ArriveeTemps.append(newTps)
        ArriveeTempsAffectes.append(doss)
        CodeRetour = Erreur(0)
    else :
        while n > 0 :
            tpsDejaPresent = ArriveeTemps[n-1]
            if tpsDejaPresent.tempsReel < newTps.tempsReel:
                ArriveeTemps.insert(n,newTps)
                ArriveeTempsAffectes.insert(n, doss)
                CodeRetour = Erreur(0)
                break
            #print("RECALCUL COMPLET DU TABLEAU AFFICHE")
            #DonneesAAfficher.reinit() # on regénère le tableau GUI
            Parametres["calculateAll"] = True # le recalcul de tous les chronos s'effectue uniquement si le temps n'est pas ajouté à la toute fin.
            ##transaction.commit()
            n -= 1
        # traitement à part du cas où l'on insère un temps en tout début de liste.
        if n == 0 and tpsDejaPresent.tempsReel > newTps.tempsReel :
            ArriveeTemps.insert(0,newTps)
            ArriveeTempsAffectes.insert(0,doss)
            CodeRetour = Erreur(0)
    if dossard != "0" :# si on affecte un dossard, cela peut avoir des conséquences sur des associations temps-coureur antérieures. On recalculera tout.
        Parametres["calculateAll"] = True
##    else :
##        CodeRetour = "Le temps à ajouter est déjà dans la liste des temps d'arrivée. Situation impossible avec les smartphones et l'interface normalement."
    #print("Insertion du temps", newTps.tempsReel, "(temps local sur le serveur) pour le dossard", doss)
    return CodeRetour

##def delDossardAffecteArriveeTemps(tempsCoureur) :
##    return affecteDossardArriveeTemps(tempsCoureur)

def affecteDossardArriveeTemps(tempsCoureur, dossard="0A") :
    """ affecte un dossard à un temps déjà présent dans les données en effectuant une recherche sur le tempsCoureur
    (utile uniquement pour les requêtes venant des smartphones.
    Appelé avec le temps seul, efface le dossard affecté. """
    doss = formateDossardNG(dossard)
    n = len(ArriveeTemps)
    message = "Temps coureur cherché : " + str(tempsCoureur) + " pour affectation du dossard " + doss + "."
    CodeRetour = Erreur(311, message , elementConcerne=tempsCoureur )
    if n != 0 :
        while n > 0 :
            tpsPresent = ArriveeTemps[n-1]
            if tpsPresent.tempsCoureur == tempsCoureur :
                ArriveeTempsAffectes[n-1] = doss
                #print("Dossard", doss, "affecté au temps sélectionné")
                CodeRetour = Erreur(0, message , elementConcerne=tempsCoureur )
                break
            n -= 1
    Parametres["calculateAll"] = True
    #transaction.commit()
    return CodeRetour

def affecteDossardArriveeTempsLocal(tempsReel, dossard="0A") :
    """ affecte un dossard à un temps déjà présent dans les données en effectuant une recherche sur le tempsReel.
    (utile pour une affectation du temps depuis l'interface graphique du serveur)
    Appelé avec le temps seul, efface le dossard affecté. """
    doss = formateDossardNG(dossard)
    n = len(ArriveeTemps)
    temps = Temps(tempsReel, 0, 0)
    message = "Temps cherché : " + str(temps.tempsReel) + " soit " + temps.tempsReelFormate() + " pour affectation du dossard " + doss + "."
    CodeRetour = Erreur(312, message , elementConcerne=tempsReel )
    print(message)
    if n != 0 :
        while n > 0 :
            tpsPresent = ArriveeTemps[n-1]
            #print("tpsPresent en position", n-1, ":",tpsPresent.tempsReel, " soit ", tpsPresent.tempsReelFormate())
            if tpsPresent.tempsReel == tempsReel :
                ArriveeTempsAffectes[n-1] = doss
                print("Dossard", doss, "affecté au temps sélectionné", temps.tempsReelFormate())
                CodeRetour = Erreur(0, message, elementConcerne=tempsReel )
                break
            n -= 1
    Parametres["calculateAll"] = True
    #transaction.commit()
    return CodeRetour


def delArriveeTempsClient(tempsCoureur, dossard="0A") :
    """ les temps sont classés par tempsReel (sur le serveur) par ordre croissant
    supprime UN tempsCoureur (mesuré côté clients donc pas forcément dans l'ordre croissant) dans la liste éventuellement associé au dossard précisé (ou non)."""
    dossard = formateDossardNG(dossard)
    tempsASupprimer = float(tempsCoureur)
    n = len(ArriveeTemps)
    if n > 0 :
        tpsDejaPresent = ArriveeTemps[n-1]
        codeRetour = "1"
        while n > 0 : # suppression du parcours conditionnel vu que la liste n'est pas triée par tempsCoureur : and tpsDejaPresent.tempsCoureur >= tempsASupprimer :
            tpsDejaPresent = ArriveeTemps[n-1]
            if dossard == "0A" :
                if tpsDejaPresent.tempsCoureur == tempsASupprimer :
                    del ArriveeTemps[n-1]
                    del ArriveeTempsAffectes[n-1]
                    codeRetour = ""
                    break
            else :
                doss = str(dossard)
                if tpsDejaPresent.tempsCoureur == tempsASupprimer and doss == ArriveeTempsAffectes[n-1] :
                    del ArriveeTemps[n-1]
                    del ArriveeTempsAffectes[n-1]
                    codeRetour = ""
                    break
            n -= 1
        if codeRetour == "" and not Parametres["calculateAll"] :
            print("on regénère tout le tableau")
            #DonneesAAfficher.reinit() # on regénère le tableau GUI
            Parametres["calculateAll"] = True
            #transaction.commit()
        #else :
            #codeRetour = ""
            #print("La suppression demandée ne peut pas être effectuée car le temps " + str(tempsCoureur) + " pour le dossard " + str(dossard) + " a déjà été supprimé.")
    return Erreur(0) # la suppression constitue une correction d'erreur et ne doit donc jamais en signaler une # codeRetour

def delArriveeTemps(tempsCoureur, dossard="0A") :
    """ supprime UN temps dans la liste par ordre croissant éventuellement associé au dossard précisé (ou non)."""
    codeRetour = ""
    dossard = formateDossardNG(dossard)
    tempsASupprimer = float(tempsCoureur)
    n = len(ArriveeTemps)
    tpsDejaPresent = ArriveeTemps[n-1]
    codeRetour = "1"
    while n > 0 and tpsDejaPresent.tempsReel >= tempsASupprimer :
        tpsDejaPresent = ArriveeTemps[n-1]
        if dossard == "0A" :
            if tpsDejaPresent.tempsReel == tempsASupprimer :
                del ArriveeTemps[n-1]
                del ArriveeTempsAffectes[n-1]
                codeRetour = ""
                break
        else :
            doss = str(dossard)
            if tpsDejaPresent.tempsReel == tempsASupprimer and doss == ArriveeTempsAffectes[n-1] :
                del ArriveeTemps[n-1]
                del ArriveeTempsAffectes[n-1]
                codeRetour = ""
                break
        n -= 1
    if codeRetour == "" :
        Parametres["calculateAll"] = True # le prochain calcul se fera sur l'intégralité des données.
    else :
        #codeRetour = ""
        print("La suppression demandée ne peut pas être effectuée : ",tempsCoureur,",", dossard,". Le temps a été supprimé depuis un autre autre appareil.")
    return Erreur(0) # la suppression constitue une correction d'erreur ne doit donc jamais créer une erreur . codeRetour # retour "" ou False si tout va bien.

##def dissocieArriveeTemps(tempsCoureur) :
##    """ supprime UN temps dans la liste par ordre croissant éventuellement associé au dossard précisé (ou non)."""
##    codeRetour = 130
##    try :
##        i = ArriveeTemps.index(tempsCoureur)
##        ArriveeTempsAffectes[i] = 0
##        codeRetour = 0
##        print("suppression effectuée")
##    except :
##        print("suppression du temps impossible")
##    return codeRetour


def coureurExists(nom, prenom) :
    return Coureurs.existe(Coureur(nom, prenom, "G"))
    # retour = 0
    # i=0
    # while i < len (Coureurs) and not retour :
        # if Coureurs[i].nom.lower() == nom.lower() and Coureurs[i].prenom.lower() == prenom.lower() :
            # retour = Coureurs[i].dossard
        # i += 1
    # return retour

def ajoutEstIlValide(nom, prenom, sexe, classe, naissance, etablissement, etablissementNature, course) :
    etablissementNatureValide = etablissementNature.upper() == "CLG" or etablissementNature.upper() == "LG" or etablissementNature.upper() == "LP"
    return nom and prenom and sexe and \
           ((Parametres["CategorieDAge"] == 0 and classe) \
             or (Parametres["CategorieDAge"] == 1 and not Parametres["CoursesManuelles"] and naissanceValide(naissance)) \
             or (Parametres["CategorieDAge"] == 1 and Parametres["CoursesManuelles"] and naissanceValide(naissance) and len(course)) \
             or (Parametres["CategorieDAge"] == 2 and naissanceValide(naissance) and etablissement and etablissementNatureValide))
             # si infos indispensables dans tous les cas
             # 0 - cas du cross du collège. On a besoin uniquement de la classe.
             # 1 - cas de courses organisées en fonction des catégories de la FFA
             # 2 - cas de courses UNSS (organisées en fonction des catégories de la FFA et des établissements)

def addCoureur(nom, prenom, sexe, classe='', naissance="", etablissement = "", etablissementNature = "", absent=None, dispense=None, temps=0,\
                commentaireArrivee="", VMA="0", aImprimer = False, licence = "", course="", dossard="", email="", email2="",\
                CoureursParClasseUpdateActif = True) : #, courseDonneeSousSonNomStandard = False):
    try :
        # print("addCoureur", nom, prenom, sexe, classe, naissance,  absent, dispense, temps, commentaireArrivee, VMA, course)
        vma = float(VMA)
    except :
        vma = "0"
    # si les données fournies sont valides ET 
    # si le dossard existe ou (si le dossard n'existe pas et qu'on le trouve par ses noms-prénoms), alors on modifie le coureur tel que spécifié.
    # sinon on crée le coureur
    if ajoutEstIlValide(nom, prenom,sexe, classe, naissance, etablissement, etablissementNature, course) :
        dossardTrouve = coureurExists(nom, prenom)
        dossardNonSpecifieEtLesNomsPrenomsExistent = (dossard == "" and dossardTrouve != "")
        dossardSpecifieEtDejaOccupe = (dossard != "" and Coureurs.existe(dossard))
        if dossardSpecifieEtDejaOccupe or dossardNonSpecifieEtLesNomsPrenomsExistent :
            if dossard == "" :
                dossard = dossardTrouve
            # # on empêche de créer plusieurs fois le même coureur avec des dossards différents.
            # if (dossard != "" and dossardTrouve == dossard) or  :
            print("On met à jour les caractéristiques du coureur au dossard", dossard)            
            ### on actualise des propriétés du coureur.
            auMoinsUnChangement = False
            coureur = Coureurs.recuperer(dossard)
            # si les paramètres sont identiques à l'existant, on ne fait rien et on ne référence pas cette actualisation pour l'interface graphique.
            #print("Actualisation de ", Coureurs[dossard-1].nom, Coureurs[dossard-1].prenom, "(", dossard, "): status, VMA, commentaire à l'arrivée.")
            if coureur.sexe != sexe :
                coureur.setSexe(sexe)
                print("sexe changé de",coureur.sexe,"en",sexe)
                auMoinsUnChangement = True
            if CoursesManuelles :
##                if courseDonneeSousSonNomStandard :
##                    lettreCourse = course
##                    course = 
##                else :
                lettreCourse = lettreCourseEnModeCoursesManuelles(course)#, avecCreation=False)
                print("course",course, "et lettreCourse" , lettreCourse, "coureur.course",coureur.course)
##                nomStandard = estDansGroupementsEnModeManuel(course)
##                if not nomStandard :
##                    nomStandard = addCourse(course)
                if lettreCourse != coureur.course :
                    addCourse(course, lettreCourse = lettreCourse) # création si besoin de la course
                    coureur.setCourse(lettreCourse)
                    print("lettre course changée")
                    auMoinsUnChangement = True
            if nom.upper() != coureur.nom.upper() :
                coureur.setNom(nom)
                print("nom changé de ", coureur.nom, "en", nom, "(", dossard, ")")
                auMoinsUnChangement = True
            if prenom.lower() != coureur.prenom.lower() :
                coureur.setPrenom(prenom)
                print("prenom changé")
                auMoinsUnChangement = True
            # print("email",email, "emailDeux", email2, "coureur.email",coureur.email, "coureur.email2",coureur.email2)
            if email != coureur.email :
                coureur.setEmail(email)
                print("email changé")
                auMoinsUnChangement = True
            if email2 != coureur.email2 :
                coureur.setEmail2(email2)
                print("email2 changé")
                auMoinsUnChangement = True
            if dispense != None and coureur.dispense != dispense :
                # print("coureur",coureur.nom)
                coureur.setDispense(dispense)
                print("dispense changée")
                auMoinsUnChangement = True
            if absent != None and coureur.absent != absent :
                coureur.setAbsent(absent)
                print("absent changé")
                auMoinsUnChangement = True
            if coureur.commentaireArrivee != commentaireArrivee :
                coureur.setCommentaire(commentaireArrivee)
                print("commentaire changé")
                auMoinsUnChangement = True
            if coureur.classe != classe :
                coureur.setClasse(classe)
                print("classe changée")
                auMoinsUnChangement = True
            if coureur.licence != licence :
                coureur.setLicence(licence)
                print("licence changée")
                auMoinsUnChangement = True
            if coureur.VMA != vma :
                coureur.setVMA(vma)
                print("VMA changée")
                auMoinsUnChangement = True
            if coureur.naissance != naissance :
                coureur.setNaissance(naissance)
                print("naissance changée")
                auMoinsUnChangement = True
            if coureur.etablissement != etablissement or coureur.etablissementNature != etablissementNature :
                coureur.setEtablissement(etablissement,etablissementNature)
                print("etablissement changé")
                auMoinsUnChangement = True
            if auMoinsUnChangement :
                if not CoursesManuelles :
                    addCourse(Coureurs.recuperer(dossard).categorie(Parametres["CategorieDAge"])) 
                    # pour toutes les courses automatiques, on doit actualiser la course si besoin.
                print("Coureur actualisé", dossard, nom, prenom, sexe, classe, naissance, etablissement, etablissementNature, absent, dispense,\
                      commentaireArrivee, " (catégorie :", Coureurs.recuperer(dossard).categorie(Parametres["CategorieDAge"]),\
                      "Course manuelle:",course,")")
                retour, d = [0,1,0,0], dossard
            else :
                retour, d = [0,0,0,1], dossard
            # else :
                # print("Le coureur existe déjà avec le dossard", dossardTrouve, "alors que le dossard proposé est", dossard)
                # retour,d  = [0,0,1,0], ""
        else :
            ### on crée le coureur (il n'a pas encore de numéro de dossard)
            if CoursesManuelles :
            ####    on cherche si la course proposée existe dans son nom et on trouve la lettre correspondante. 
            ####    Si non, on la crée et on récupère la lettre.
            ####    lettreCourse = ...
                lettre = ""
                if dossard : 
                    lettre = dossard[-1]
                # lettreCourse = addCourse(course, lettreCourse = lettre) # crée la course si besoin et surtout, retourne sa lettre à partir de son nom. #lettreCourseEnModeCoursesManuelles(course)
                # print("lettreCourse",lettreCourse)
                # récupération du nom standard de la course
                # nomStandard = estDansGroupementsEnModeManuel(course)
                # ne devrait jamais arriver puisque addCourse() a été exécuté ci-dessus if not nomStandard :
                    #nomStandard = addCourse(course)
            else :
                lettreCourse = "A"
                # addCourse(Coureurs.recuperer(dossard).categorie(Parametres["CategorieDAge"]))
            if dossard != "" : # si un numéro de dossard est proposé à l'import de nouveaux coureurs, on teste si il est libre.
                # si oui, alors, on crée le coureur.
                if (not Coureurs.existe(dossard)) and lettreCourse == formateDossardNG(dossard)[-1] : # le dossard est libre et la lettre du dossard correcte.
                    # création avec le dossard imposé.
                    dossard = Coureurs.ajouter(Coureur( nom, prenom, sexe, classe=classe, naissance=naissance, etablissement=etablissement,\
                                                etablissementNature=etablissementNature, absent=absent,\
                                                dispense=dispense, temps=temps, commentaireArrivee=commentaireArrivee, VMA=vma,\
                                                aImprimer=aImprimer,\
                                                course=lettreCourse, email=email, email2=email2), dossard = dossard)
                else :
                    print("ERREUR : le dossard ", dossard, "n'est pas libre et ne correspond pas au coureur en cours de création (ou à la lettre de course :", lettreCourse,").")
            else :
                dossard = Coureurs.ajouter(Coureur( nom, prenom, sexe, classe=classe, naissance=naissance, etablissement=etablissement,\
                                                etablissementNature=etablissementNature, absent=absent,\
                                                dispense=dispense, temps=temps, commentaireArrivee=commentaireArrivee, VMA=vma,\
                                                aImprimer=aImprimer,\
                                                course=lettreCourse, email=email, email2=email2), course = lettreCourse)
            # on crée la course après le coureur pour disposer de la catégorie quand elle est calculée par l'objet Coureur.
            if CoursesManuelles :
                lettreCourse = addCourse(course, lettreCourse = lettre) # crée la course si besoin et surtout, retourne sa lettre à partir de son nom. #lettreCourseEnModeCoursesManuelles(course)
                print("lettreCourse",lettreCourse)
            else :
                addCourse(Coureurs.recuperer(dossard).categorie(Parametres["CategorieDAge"]))
            ##print("dossard récupéré:",dossard)
            ##transaction.commit()
            print("Coureur", dossard,"ajouté", nom, prenom, sexe, classe, naissance, etablissement, etablissementNature, "dans course",\
                  lettreCourse, " (",course,")")
            #print(" (catégorie :", Coureurs.recuperer(dossard).categorie(Parametres["CategorieDAge"]),")", "Course :", course)
            ## Coureurs[-1].setCourse(addCourse(course))
            retour, d = [1,0,0,0], dossard
    else :
        print("Il manque un paramètre obligatoire (valide) pour créer le coureur. nom=",nom," ; prénom=",prenom," ; sexe=",sexe," ; classe=",classe," ; naissance=",naissance," ; établissement=",etablissement," ; établissementType=", etablissementNature)
        retour,d  = [0,0,1,0], ""
    if CoureursParClasseUpdateActif :
        CoureursParClasseUpdate()
    return retour, d
##    except :
##        print("Impossible d'ajouter " + nom + " " + prenom + " avec les paramètres fournis : VMA invalide,...")
##        print(nom, prenom, sexe, classe, naissance, etablissement, etablissementNature, absent, dispense, temps, commentaireArrivee, VMA, aImprimer)


def lettreCourseEnModeCoursesManuelles(course):#, avecCreation=True) :
    """ retourne la lettre correspondant au nom d'une course en mode CoursesManuelles
        si la course n'existe pas, retourne la prochaine lettre disponible dans l'ordre alphabétique"""
    trouve = False
    a = 1
    for g in Groupements :
        if g.nom == course :
            trouve = True
            break
        a += 1
    retour = chr(a+64)
    if not trouve :
         # print("Recherche de la première lettre non encore utilisée")
        for a in range(1,27) :
            if not chr(a+64) in Courses.keys() :
                break
        retour = chr(a+64)
        print("La première lettre non encore utilisée est", retour)
    return retour



def estDansGroupementsEnModeManuel(course):
    retour = ""
    for g in Groupements :
        if g.nomStandard == course :
            retour = g.nomStandard
            break
    return retour
    

def addCourse(course, lettreCourse="") :
    """ addCourse prend en argument le nom personnalisé de la course (obligatoire)
    En mode CoursesManuelles, la lettre que l'on veut attribuer à cette course peut être fourni (facultatif)
    Comportement :
    en mode automatique, crée la course si elle n'existe pas. Crée le groupement du même nom (par défaut)
    en mode coursesManuelles, crée la course et le groupement mais peut également imposer la lettre, si fournie.
    Dans ce cas, les courses intermédiaires sont créées avec le nom standard,
    et le nom de la course avec la lettre est actualisé si le nom en cours est standard."""
##    # si CoursesManuelles, les courses portent le nom "A" comme entrée dans Courses.
##    # on doit trouver si une course existante c a pour propriété c.nom == categorie
##    # si ce n'est pas le cas, on crée la course et le groupement correspondant à l'identique en affectant le nom personnalisé avec la méthode adhoc
    if course :
        if CoursesManuelles :
            if not lettreCourse :
                lettreCourse = lettreCourseEnModeCoursesManuelles(course)
                print("La lettre attribuée manuellement à la course est :",lettreCourse)
            #print("lettreCourse:",lettreCourse,"Courses => ", Courses.keys())
            if not lettreCourse in Courses :
                print("Création de la course manuelle", lettreCourse, "avec le nom :", course)
                Groupements.append(Groupement(lettreCourse,[lettreCourse]))
                Groupements[-1].setNom(course)
                c = Course(course)
                Courses.update({lettreCourse : c})
            return lettreCourse
        else :
            # compatibilité ascendante pour créer les groupements pour des courses qui existeraient déjà dans de vieilles bases de données.
            estPresent = False
            for grpment in Groupements :
                if course in grpment.listeDesCourses :
                    estPresent = True
                    break
            if not estPresent:
                print("Création du groupement dans addCourse", course)
                Groupements.append(Groupement(course,[course]))
                #print("Groupements = ",[i.nom for i in Groupements])
            # création de la course si elle n'existe pas.
            #print(course, " est dans ", Courses,"?")
            if course not in Courses :
                print("Création de la course", course)
                c = Course(course)
                Courses.update({course : c})
            return course
            #print("cat",Courses[course].categorie)
    else :
        print("ERREUR qui ne devrait pas survenir, la course transmise vaut :", course)

def estDansGroupementsEnModeManuel(course):
    retour = ""
    for g in Groupements :
        if g.nomStandard == course :
            retour = g.nomStandard
            break
    return retour
    

def addCourse(course, lettreCourse="") :
    """ addCourse prend en argument le nom personnalisé de la course (obligatoire)
    En mode CoursesManuelles, la lettre que l'on veut attribuer à cette course peut être fourni (facultatif)
    Comportement :
    en mode automatique, crée la course si elle n'existe pas. Crée le groupement du même nom (par défaut)
    en mode coursesManuelles, crée la course et le groupement mais peut également imposer la lettre, si fournie.
    Dans ce cas, les courses intermédiaires sont créées avec le nom standard,
    et le nom de la course avec la lettre est actualisé si le nom en cours est standard."""
##    # si CoursesManuelles, les courses portent le nom "A" comme entrée dans Courses.
##    # on doit trouver si une course existante c a pour propriété c.nom == categorie
##    # si ce n'est pas le cas, on crée la course et le groupement correspondant à l'identique en affectant le nom personnalisé avec la méthode adhoc
    if CoursesManuelles :
        if not lettreCourse :
            lettreCourse = lettreCourseEnModeCoursesManuelles(course)
            print("La lettre attribuée manuellement à la course est :",lettreCourse)
        #print("lettreCourse:",lettreCourse,"Courses => ", Courses.keys())
        if not lettreCourse in Courses :
            print("Création de la course manuelle", lettreCourse, "avec le nom :", course)
            Groupements.append(Groupement(lettreCourse,[lettreCourse]))
            Groupements[-1].setNom(course)
            c = Course(course)
            Courses.update({lettreCourse : c})
        return lettreCourse
    else :
        # compatibilité ascendante pour créer les groupements pour des courses qui existeraient déjà dans de vieilles bases de données.
        estPresent = False
        for grpment in Groupements :
            if course in grpment.listeDesCourses :
                estPresent = True
                break
        if not estPresent:
            print("Création du groupement ", course)
            Groupements.append(Groupement(course,[course]))
            #print("Groupements = ",[i.nom for i in Groupements])
        # création de la course si elle n'existe pas.
        #print(course, " est dans ", Courses,"?")
        if course not in Courses :
            print("Création de la course", course)
            c = Course(course)
            Courses.update({course : c})
        return course
        #print("cat",Courses[course].categorie)

def formateDossardNG(doss) :
    if doss :
        if not str(doss)[-1].isalpha(): # si le dernier caractère n'est pas une lettre, on ajoute le A. Sinon, on s'assure de la présence de majuscules.
            doss = str(doss) + "A"
        else :
            doss = str(doss).upper()
    return doss.replace(" ","")

def addArriveeDossard(dossard, dossardPrecedent=-1) :
    """ajoute un dossard sur la ligne d'arrivée dans l'ordre si non précisé. Si l'index est précisé, insère à l'endroit précisé par dossardPrecedent
        si dossardPrécedent vaut 0, insère en première position.
        Retourne 0 s'il n'y a pas d'erreur."""
    doss = formateDossardNG(dossard)
    dossPrecedent = formateDossardNG(dossardPrecedent)
    if Coureurs.existe(doss) :
        coureur = Coureurs.recuperer(doss)
        if CategorieDAge :
            infos = "dossard " + str(dossard) + " - " + coureur.nom + " " + coureur.prenom + " (" + coureur.categorie(Parametres["CategorieDAge"]) + ")."
        else :
            infos = "dossard " + str(dossard) + " - " + coureur.nom + " " + coureur.prenom + " (" + coureur.classe + ")."
        message = ""
        retour = Erreur(0)
        if doss in ArriveeDossards :
            ### ce cas commenté est géré proprement par un uid et un noTransmission uniques transmis par les smartphones pour chaque donnée transmise.
            ### Le doublon sera donc détecté proprement, y compris pour un temps transmis en doublon, y compris si ces deux transmissions ne sont pas dans deux lignes consécutives de DonneesSmartphone.txt, etc...
            message = "Dossard ayant déjà passé la ligne d'arrivée :\n" + infos
            print(message)
            retour=Erreur(401,message,elementConcerne=doss)
                # en conditions réelles, il arrive que le wifi ne fonctionne pas. Théoriquement l'appli smartphone empêche qu'un dossard soit scanné deux fois.
                #Mais si l'envoi des données du smartphone vers le serveur ne s'est pas vu accuser réception, le smartphone envoie une deuxième fois le dossard et on a un bloquant.
                # Désormais, le deuxième envoi est ignoré pour que l'interface ne se bloque plus sur cette erreur précise.
                # return message
        elif not Coureurs.existe(doss) :#doss > len(Coureurs) or doss < 1 :
            message = "Numéro de dossard incorrect :\n"  + infos
            print(message)
            retour=Erreur(411,message,elementConcerne=doss)
        elif Coureurs.recuperer(doss).absent :
            message = "Ce coureur ne devrait pas avoir passé la ligne d'arrivée car absent :\n" + infos
            print(message)
            retour=Erreur(421,message,elementConcerne=doss)
        elif Coureurs.recuperer(doss).dispense :
            message = "Ce coureur ne devrait pas avoir passé la ligne d'arrivée car dispensé :\n" + infos
            print(message)
            retour=Erreur(421,message,elementConcerne=doss)
        elif not Courses[Coureurs.recuperer(doss).course].depart :
            message = "La course " + groupementAPartirDeSonNom(Coureurs.recuperer(doss).course, nomStandard = True).nom + " n'a pas encore commencé. Ce coureur ne devrait pas avoir passé la ligne d'arrivée :\n" + infos
            print(message)
            retour=Erreur(431,message,elementConcerne=doss)
        ### changement comportemental du logiciel : Même s'il y a une erreur, on ajoute le dossard dans ArriveeDossards au bon endroit. Ainsi, il apparaîtra dans l'interface. L'erreur sera signalée et devra être corrigée.
        #print("dossard precéent:",dossPrecedent)
        #print("ArriveeDossards",ArriveeDossards)
        if dossPrecedent == "-1A" : # CAS COURANT : ajoute à la suite
            #position = len(ArriveeDossards)
            ArriveeDossards.append(doss)
            #print("Dossard arrivé :",doss)
        elif dossPrecedent == "0A" : # insère au début de liste
            print("insertion en début de liste d'arrivée")
            #position = 0
            ArriveeDossards.insert(0 , doss)
            Parametres["calculateAll"] = True
            #DonneesAAfficher.reinit() # on regénère le tableau GUI
        else :
            # insère juste après le dossard dossardPrecedent , si on le trouve.
            try :
                n = ArriveeDossards.index(dossPrecedent)
                print("Insertion du dossard", doss, "juste après", dossPrecedent, "à l'indice", n)
                #position = n+1
                ArriveeDossards.insert(n+1 , doss)
                Parametres["calculateAll"] = True
                #DonneesAAfficher.reinit() # on regénère le tableau GUI
            except ValueError :
                message = "DossardPrecedent non trouvé : cela ne devrait pas survenir via l'interface graphique :\n" + infos
                print(message)
                retour=Erreur(499,message, elementConcerne=doss)
    else :
        message = "Le dossard " + doss + " n'existe pas : il a été supprimé après la saisie des résultats de course. Anormal mais ignoré."
        print(message)
        retour=Erreur(411,message, elementConcerne=doss)
    return retour

def imprimePDF(pdf_file_name) :
    if os.path.exists(pdf_file_name) :
        win32api.ShellExecute (0, "print", pdf_file_name, None, ".", 0)
##        INCH = 1440
##        hDC = win32ui.CreateDC ()
##        hDC.CreatePrinterDC (win32print.GetDefaultPrinter ())
##        hDC.StartDoc (pdf_file_name)
##        hDC.StartPage ()
##        hDC.SetMapMode (win32con.MM_TWIPS)
##        hDC.DrawText ("TEST", (0, INCH * -1, INCH * 8, INCH * -2), win32con.DT_CENTER)
##        hDC.EndPage ()
##        hDC.EndDoc ()
        return True
    else :
        print("Le fichier", pdf_file_name, "n'existe pas.")
        return False

def calculeTousLesTemps(reinitialise = False):
    """ associe un temps à chaque dossard ayant passé la ligne d'arrivée permet les décalages positifs et négatifs.
    Si l'argument est True, recalcule tout depuis le début.
    Retourne une liste d'instances de la class Erreur pour traitement par l'interface GUI."""
    #print("Début de CalculeTousLesTemps(",reinitialise,")")
    global ligneTableauGUI,TableauGUI
    retour = []
    reinitTableauGUI()
    #print('Parametres["calculateAll"]', Parametres["calculateAll"])
    if Parametres["calculateAll"] or reinitialise :
        Parametres["positionDansArriveeTemps"] = 0
        Parametres["positionDansArriveeDossards"] = 0
        ligneTableauGUI = [1,0]
        print("on REINITIALISE")
        # print(globals())
        if 'tableau' in globals() :
            tableau.delTreeviewFrom(ligneTableauGUI[0])
            print("on efface treeview jusqu'à la ligne",ligneTableauGUI[0])
    i = Parametres["positionDansArriveeTemps"]
    j = Parametres["positionDansArriveeDossards"]
    chronosInutilesAvantLeDossard = 0
    ligneAjoutee = ligneTableauGUI[0]
    derniereLigneStabilisee = ligneTableauGUI[1]
    #print("DerniereLigneStabilisee au début",derniereLigneStabilisee, "i=", i, "j=",j,"retour",retour)
    #print("len(ArriveeDossards)",len(ArriveeDossards), "len(ArriveeTemps)",len(ArriveeTemps))
    while j < len(ArriveeDossards) and i < len(ArriveeTemps):
        # chaque dossard scanné doit se voir attribué un temps. i < len(ArriveeTemps) à tester plus loin.
        doss = formateDossardNG(ArriveeDossards[j])
        #print(ligneTableauGUI,"dossard", doss, reinitialise)
        ### debug
        tps = ArriveeTemps[i]
        dossardAffecteAuTps = formateDossardNG(ArriveeTempsAffectes[i])
        if (dossardAffecteAuTps != "0" and dossardAffecteAuTps != "0A") and Coureurs.existe(dossardAffecteAuTps) : # dossardAffecteAuTps <= len(Coureurs) :
            # 2ème test pour s'assurer que le dossard affecté existe. Prévient des bugs de saisie smartphones.
            # un dossard est affecté. On doit trouver le dossard dans ArriveeDossards
            if dossardAffecteAuTps == doss :
                # tout est désormais bien calé entre les deux listes aux indices i et j qui se correspondent à ce stade.
                #print("Le dossard", doss, "est affecté manuellement au temps indice n°", i, ". TempsCoureur=",tps.tempsCoureur, ", TempsReelCalculé=",tps.tempsReel)
                retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
                i += 1
                j += 1
            else :
                # on affecte au dossard rencontré le temps i-1 (tant que tout n'est pas recalé).
                tps = ArriveeTemps[i-1]
                dossardAffecteAuTps = formateDossardNG(ArriveeTempsAffectes[i-1])
                # retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps,ligneAjoutee, derniereLigneStabilisee, True)
                retour += affecteChronoAUnCoureur(doss, tps, '-',ligneAjoutee, derniereLigneStabilisee, True)
                #retour += "<p><red>Il manque un chrono juste avant le dossard " + str(dossardAffecteAuTps) + ". Le dossard " + str(doss) + " se voit affecté le temps de son prédécesseur.</red></p>\n"
                j += 1
        else :
            if doss in ArriveeTempsAffectes[i+1 : ] :
                # le temps actuel n'est pas affecté mais le dossard scanné est affecté à un autre temps (plus loin), on s'y rend en construisant le tableau affiché avec les temps intermédiaires.
                #iTrouve = ArriveeTempsAffectes.index(doss)
                #tps = ArriveeTemps[iTrouve]
                #decalage = iTrouve - i
                #i = iTrouve
                if chronosInutilesAvantLeDossard != doss :
                    message = "Il y a un (des) chrono(s) inutilisé(s) juste avant le dossard " + doss + "."
                    retour.append(Erreur(321,message, elementConcerne=doss))
                    chronosInutilesAvantLeDossard = doss
                #DonneesAAfficher.append(coureurVide,tps, dossardAffecteAuTps)
                coureur = Coureurs.recuperer(doss)
                alimenteTableauGUI (tableauGUI, coureurVide, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
                i += 1
                #retour += affecteChronoAUnCoureur(doss, tps)
            else :
                # pas d'affectation manuelle (cas le plus courant). On poursuit l'affectation dossard <=> temps aux rangs i et j des deux listes.
                retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
                i+=1
                j+=1
        ligneAjoutee += 1
    # à partir de cet indice, il faut actualiser systématiquement le tableau de l'interface GUI
    #print("derniere ligne stabilisée", ligneAjoutee - 1)
    derniereLigneStabilisee = ligneAjoutee - 1
    #print("derniereLigneStabilisee au milieu",derniereLigneStabilisee, "i=", i, "j=",j, "retour", retour)
    # Cas classique où des coureurs sont dans la file d'attente pour être scannés.
    if j == len(ArriveeDossards) :
        k = i
        while k < len(ArriveeTemps) :
            #print("Ajout des temps sans dossard affecté")
            tps = ArriveeTemps[k]
            dossardAffecteAuTps = formateDossardNG(ArriveeTempsAffectes[k])
            #DonneesAAfficher.append(coureurVide,tps, dossardAffecteAuTps, True)
            if k == i and (Parametres["CategorieDAge"]==0 or Parametres["CategorieDAge"]==2) and categorieDuDernierDepart() != "" :
                ### seul le premier temps prévisionnel sera affiché (pour disposer du premier d'une course).
                coureurPrevisionnel = Coureur("", "", "", "")
                arrivee = tps.tempsReel
                cat = categorieDuDernierDepart()
                depart = Courses[cat].temps
                coureurPrevisionnel.setTemps(arrivee- depart, Courses[cat].distance)
                alimenteTableauGUI (tableauGUI, coureurPrevisionnel, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
            else :
                ### affichage classique des heures de passage vides sans coureur affecté.
                alimenteTableauGUI (tableauGUI, coureurVide, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
            ligneAjoutee += 1
            k += 1
    # Cas tordu : s'il n'y a pas assez de temps saisis à la toute fin, affecte le dernier temps à tous les derniers dossards.
    #print("ArriveeTemps:", ArriveeTemps, len(ArriveeTemps))
    #print("ArriveeDossards :", ArriveeDossards, len(ArriveeDossards))
    if i == len(ArriveeTemps) and len(ArriveeTemps)>0 :
        k = j
        if k != len(ArriveeDossards) :
            message = "Pas assez de temps saisis pour le(s) dernier(s) dossards :\nil faut leur affecter un temps comme proposé."
            retour.append(Erreur(331,message))
            print(message)
        while k < len(ArriveeDossards) :
            #print("position dans ArriveeTemps" , i-1, "   poisiotn dans ArriveeDossard ",k)
            doss = formateDossardNG(ArriveeDossards[k])
            tps = ArriveeTemps[i-1]
            if k == j :
                dossardAffecteAuTps = formateDossardNG(ArriveeTempsAffectes[i-1])
            else :
                dossardAffecteAuTps = "0"
            retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
            ligneAjoutee += 1
            k += 1
    ligneTableauGUI = [derniereLigneStabilisee + 1 , derniereLigneStabilisee]
    #### ligneTableauGUI[0] = ligneTableauGUI[1] + 1
    if Parametres["calculateAll"] :
        #print("DONNEES UTILES GUI:",ligneTableauGUI, tableauGUI)
        Parametres["calculateAll"] = False
    Parametres["positionDansArriveeTemps"] = i
    Parametres["positionDansArriveeDossards"] = j
    #print("A la fin de calculeTousLesTemps",retour)
    #print("Erreur 0",retour[0].numero ,retour[0].description)
    return retour

def categorieDuDernierDepart() :
    ''' retourne l'heure du dernier départ lancé'''
    cat = ""
    tempsMax = 0
    for nom in Courses :
        c = Courses[nom]
        if c.depart and c.temps > tempsMax :
            cat = c.categorie
            tempsMax = c.temps
    return cat

##def calculeTousLesTemps(reinitialise = False):
##    """ associe un temps à chaque dossard ayant passé la ligne d'arrivée permet les décalages positifs et négatifs.
##    Si l'argument est True, recalcule tout depuis le début.
##    Retourne un code html (ou pas) à afficher dans l'interface GUI afin de montrer en temps réel les décalages calculés ("les erreurs")."""
##    global ligneTableauGUI,TableauGUI
##    retour = ""
##    reinitTableauGUI()
##    #print('Parametres["calculateAll"]', Parametres["calculateAll"])
##    if Parametres["calculateAll"] or reinitialise :
##        Parametres["positionDansArriveeTemps"] = 0
##        Parametres["positionDansArriveeDossards"] = 0
##        ligneTableauGUI = [1,0]
##    if 'tableau' in globals() :
##        tableau.delTreeviewFrom(ligneTableauGUI[0])
##    if ArriveeTemps :
##        i = Parametres["positionDansArriveeTemps"]
##        j = Parametres["positionDansArriveeDossards"]
##        chronosInutilesAvantLeDossard = 0
##        ligneAjoutee = ligneTableauGUI[0]
##        derniereLigneStabilisee = ligneTableauGUI[1]
##        while j < len(ArriveeDossards) and i < len(ArriveeTemps): # chaque dossard scanné doit se voir attribué un temps. i < len(ArriveeTemps) à tester plus loin.
##            doss = ArriveeDossards[j]
##            tps = ArriveeTemps[i]
##            dossardAffecteAuTps = ArriveeTempsAffectes[i]
##            if dossardAffecteAuTps != 0 and dossardAffecteAuTps <= len(Coureurs) : # 2ème test pour s'assurer que le dossard affecté existe. Prévient des bugs de saisie smartphones.
##                # un dossard est affecté. On doit trouver le dossard dans ArriveeDossards
##                if dossardAffecteAuTps == doss :
##                    # tout est désormais bien calé entre les deux listes aux indices i et j qui se correspondent à ce stade.
##                    #print("Le dossard", doss, "est affecté manuellement au temps indice n°", i, ". TempsCoureur=",tps.tempsCoureur, ", TempsReelCalculé=",tps.tempsReel)
##                    retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
##                    i += 1
##                    j += 1
##                else :
##                    # on affecte au dossard rencontré le temps i-1 (tant que tout n'est pas recalé).
##                    tps = ArriveeTemps[i-1]
##                    dossardAffecteAuTps = ArriveeTempsAffectes[i-1]
##                    # retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps,ligneAjoutee, derniereLigneStabilisee, True)
##                    retour += affecteChronoAUnCoureur(doss, tps, '-',ligneAjoutee, derniereLigneStabilisee, True)
##                    retour += "<p><red>Il manque un chrono juste avant le dossard " + str(dossardAffecteAuTps) + ". Le dossard " + str(doss) + " se voit affecté le temps de son prédécesseur.</red></p>\n"
##                    j += 1
##            else :
##                if doss in ArriveeTempsAffectes[i+1 : ] :
##                    # le temps actuel n'est pas affecté mais le dossard scanné est affecté à un autre temps (plus loin), on s'y rend en construisant le tableau affiché avec les temps intermédiaires.
##                    #iTrouve = ArriveeTempsAffectes.index(doss)
##                    #tps = ArriveeTemps[iTrouve]
##                    #decalage = iTrouve - i
##                    #i = iTrouve
##                    if chronosInutilesAvantLeDossard != doss :
##                        retour += "<p><red>Il y a un (des) chrono(s) inutilisé(s) juste avant le dossard " + str(doss) + "</red></p>\n"
##                        chronosInutilesAvantLeDossard = doss
##                    #DonneesAAfficher.append(coureurVide,tps, dossardAffecteAuTps)
##                    coureur = Coureurs[doss - 1]
##                    alimenteTableauGUI (tableauGUI, coureurVide, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
##                    i += 1
##                    #retour += affecteChronoAUnCoureur(doss, tps)
##                else :
##                    # pas d'affectation manuelle (cas le plus courant). On poursuit l'affectation dossard <=> temps aux rangs i et j des deux listes.
##                    retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
##                    i+=1
##                    j+=1
##            ligneAjoutee += 1
##        # à partir de cet indice, il faut actualiser systématiquement le tableau de l'interface GUI
##        #print("derniere ligne stabilisée", ligneAjoutee - 1)
##        ligneTableauGUI[1] = ligneAjoutee - 1
##        # Cas classique où des coureurs sont dans la file d'attente pour être scannés.
##        if j == len(ArriveeDossards) :
##            k = i
##            while k < len(ArriveeTemps) :
##                #print("Ajout des temps sans dossard affecté")
##                tps = ArriveeTemps[k]
##                dossardAffecteAuTps = ArriveeTempsAffectes[k]
##                #DonneesAAfficher.append(coureurVide,tps, dossardAffecteAuTps, True)
##                alimenteTableauGUI (tableauGUI, coureurVide, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
##                ligneAjoutee += 1
##                k += 1
##        # Cas tordu : s'il n'y a pas assez de temps saisis à la toute fin, affecte le dernier temps à tous les derniers dossards.
##        if i == len(ArriveeTemps) :
##            k = j
##            if k != len(ArriveeDossards) :
##                print("Pas assez de temps saisis pour le(s) dernier(s) dossards : on leur affecte le dernier temps")
##            while k < len(ArriveeDossards) :
##                doss = ArriveeDossards[k]
##                tps = ArriveeTemps[i-1]
##                if k == j :
##                    dossardAffecteAuTps = ArriveeTempsAffectes[i-1]
##                else :
##                    dossardAffecteAuTps = 0
##                retour += affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee)
##                ligneAjoutee += 1
##                k += 1
##        ligneTableauGUI[0] = ligneTableauGUI[1] + 1
##        #print(DonneesAAfficher.lignes)
##        #print("DONNEES UTILES GUI:",ligneTableauGUI, tableauGUI)
##        Parametres["calculateAll"] = False
##        Parametres["positionDansArriveeTemps"] = i
##        Parametres["positionDansArriveeDossards"] = j
##        ##transaction.commit()
##    else :
##        print("tableau GUI à vider")
##        tableauGUI=["reinit"]
##    if reinitialise :
##        retour = "<p>RECALCUL DE TOUS LES TEMPS :</p>\n" + retour
##    return retour


def affecteChronoAUnCoureur(doss, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee, tpsNonSaisi=False):
    arrivee = tps.tempsReel
    coureur = Coureurs.recuperer(doss)
    cat = coureur.course # categorie(Parametres["CategorieDAge"])
    retour = []
    try :
        categ = Courses[cat]
    except:
        categ = ""
    if categ and Courses[cat].depart :
        depart = categ.temps
        if arrivee- depart < 0 :
            coureur.setTemps(0)
            #print("Temps calculé pour le coureur ", coureur.nom, " négatif :", arrivee , "-", depart, "=" , arrivee- depart, " dossard:", doss)
            message = "Le coureur " + coureur.nom + " " + coureur.prenom + " (" + doss + ") a un temps négatif :\nDépart (" + str(categ.categorie)  +") : " + str(categ.departFormate())  + " / Arrivée : "+ str(formaterTempsALaSeconde(arrivee))
            retour.append(Erreur(211,message, elementConcerne=doss))
            print(message)
            # test pour afficher les erreurs dans l'interface GUI :
            alimenteTableauGUI (tableauGUI, coureur, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
        else :
            coureur.setTemps(arrivee- depart, categ.distance)
            if tpsNonSaisi :
                alimenteTableauGUI (tableauGUI, coureur, Temps(0,0,0), dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
                #DonneesAAfficher.append(coureur,Temps(0,0,0), dossardAffecteAuTps)
            else :
                alimenteTableauGUI (tableauGUI, coureur, tps , dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
                #DonneesAAfficher.append(coureur,tps, dossardAffecteAuTps)
            #if DEBUG :
                #print("On affecte le temps ",arrivee,"-",depart,"=",formateTemps(coureur.temps)," au coureur ",doss, "de rang", coureur.rang)
            message = "On affecte le temps " + str(arrivee) + " - " + str(depart) + " = " + formateTemps(coureur.temps) + " au coureur " + str(doss)+"."
            #print(message)
            retour.append(Erreur(0))
    else :
        coureur.setTemps(-1) # le chrono -1 signifie que la course n'est pas lancée.
        #print("La course ", cat, "n'est pas partie. Le coureur ", coureur.nom, " n'est pas censé avoir franchi la ligne")
        # test pour afficher les erreurs dans l'interface GUI :
        #alimenteTableauGUI (tableauGUI, coureur, tps, dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
        message = "La course " + cat+ " n'est pas partie. Le coureur "+ coureur.nom+ " n'est pas censé avoir franchi la ligne."
        print(message)
        ### cette erreur est déjà signalée par addArriveeDossard : on ne fait pas un 2ème signalement. retour.append(Erreur(431,message))
        #print("-----------------------------")
        #print("Erreur : le coureur", coureur.dossard, "est dans une catégorie non partie")
        #print("-----------------------------")
        ### il faut désormais alimenter le tableau avec les coureurs qui n'ont pas passé la ligne d'arrivée.
        alimenteTableauGUI (tableauGUI, coureur, tps , dossardAffecteAuTps, ligneAjoutee, derniereLigneStabilisee )
    #if retour[0].numero :
    #    print("retour affecteChronoAUnCoureur",retour, retour[0].numero)
    return retour
    
def tupleEtablissement() :
    '''retourne un tuple avec tous les établissements de tous les coureurs'''
    L = []
    for c in Coureurs.liste() :
        if not c.etablissement in L :
            L.append(c.etablissement)
    return tuple(L)

##def reindexDossards(numeroInitial) :
##    """ Réindexe tous les numéros de dossards des coureurs en cas d'insertion ou de suppression dans la liste Coureurs.\
##    Le dossard doit toujours être égal à l'index + 1 """
##    if not Parametres["CourseCommencee"] :
##        i = numeroInitial
##        while i < len(Coureurs) :
##            coureur = Coureurs[i]
##            #print("on rénumérote les dossards pour", coureur.nom, coureur.prenom)
##            coureur.setDossard(i+1)
##            i += 1


def delCoureur(dossard):
    if not Parametres["CourseCommencee"] :
        course = Coureurs.recuperer(dossard).course
        Coureurs.efface(dossard)
        delCourse(course)
        print("Coureur effacé :", dossard,".")
    # doss = int(dossard)
    # if not Parametres["CourseCommencee"] :
        # if doss-1 < len(Coureurs) :
            # categorie = Coureurs[doss - 1].categorie(Parametres["CategorieDAge"])
            # del Coureurs[doss-1]
            # # plus besoin car les dossards correspondront toujours avec le nouveau dictionnaire Coureurs : reindexDossards(doss-1)
            # delCourse(categorie)
            # print("Coureur effacé :", dossard,".")
        # else :
            # print("le dossard fourni n'existe pas :",dossard)

def delArriveeDossard(dossard, dossardPrecedent="-1"):
    """Nouvelle version de delArriveeDossard qui supprime dossard de ArriveeDossards. 
    
    Elle prend en compte l'argument dossardPrecedent qui peut valoir :
        -1 si non spécifié. Auquel cas, supprime la première occurence de dossard dans la liste ArriveeDossards
        0 s'il faut supprimer le premier dossard de ArriveeDossards
        un dossard valide (comme "2A", "1B", etc... qui précéde alors le numéro à supprimer.
        
        Cette fonction ne supprime rien si le dossard n'est pas trouvé ou si son prédécesseur, si spécifié, n'est pas trouvé."""
    #print("ArriveeDossards",ArriveeDossards)
    retour = Erreur(441, "Erreur inconnue dans delArriveeDossard()")
    if ArriveeDossards :
        doss = formateDossardNG(dossard)
        dossardPrec = formateDossardNG(dossardPrecedent)
        if dossardPrecedent == "-1" : 
            # l'ancienne version de chronoHB utilisait -1 pour indiquer qu'il n'y avait pas de prédécesseur.
            # dans ce cas, on supprime la première occurence de dossard, si possible (ancienne version de delArriveeDossard finalement)
            try :
                Coureurs.recuperer(doss).setTemps(0)
                Parametres["calculateAll"] = True
                ArriveeDossards.remove(doss)
                retour = Erreur(0)
                print("Dossard " + str(doss)  + " supprimé du passage sur la ligne d'arrivée en tant que première occurence. Pas de dossard prédécesseur spécifié.")
            except :
                message = "Le dossard " + str(doss) + " n'a pas encore passé la ligne d'arrivée et ne peut donc pas être supprimé."
                print(message)
                retour = Erreur(441, doss, message) # la suppression d'un dossard dans l'interface peut constituer une correction d'erreur. Elle ne doit pas provoquer elle-même une erreur .
        elif dossardPrecedent == "0A" or dossardPrecedent == "0" : # le dossard à supprimer est le premier de la liste, normalement.
            if ArriveeDossards[0] == doss :
                # on supprime le premier élément de la liste.
                print("Suppression du dossard", doss, "avec comme prédécesseur", dossardPrec)
                ArriveeDossards.pop(0)
                retour = Erreur(0)
            else :
                message = "Le premier dossard de la liste ArriveeDossards n'est pas " + str(doss) + " mais " + ArriveeDossards[0] +"."
                print(message)
                retour = Erreur(441, doss, message)
        elif len(dossardPrec) > 1 : # cas qui va devenir le plus classique via la nouvelle version 1.7 de l'interface et sur smartphone. 
        # Le dossard prédécesseur sera forcément spécifié pour ne pas supprimer n'importe lequel !
            i = 1
            pasTrouve = True
            while i < len(ArriveeDossards) and pasTrouve: 
                #print(ArriveeDossards[i]," == ",doss," and ",ArriveeDossards[i-1]," == ",dossardPrec)
                if ArriveeDossards[i] == doss and ArriveeDossards[i-1] == dossardPrec :
                    # suppression de l'élément i de la liste.
                    print("Suppression du dossard", doss, "avec comme prédécesseur", dossardPrec)
                    ArriveeDossards.pop(i)
                    pasTrouve = False
                i += 1
            if pasTrouve :
                message = "Il n'y a aucun dossard " + doss + " de la liste ArriveeDossards précédé par " + dossardPrec + "."
                print(message)
                retour = Erreur(441, doss, message)
            else :
                retour = Erreur(0)
        else :
            print("ArriveeDossards ne contient qu'un seul élément. Impossible de supprimer le dossard",dossard,"avec comme prédécesseur",dossardPrecedent,".")
    else :
        print("ArriveeDossards est vide. Impossible de supprimer le dossard",dossard,"avec comme prédécesseur",dossardPrecedent,".")
    return retour
    
# def delArriveeDossard(dossard):
# ##    PasDErreur = False
    # doss = formateDossardNG(dossard)
    # if not Parametres["CourseCommencee"] :
        # try :
            # Coureurs.recuperer(doss).setTemps(0)
            # #if doss != ArriveeDossards[-1] :
            # Parametres["calculateAll"] = True
            # #transaction.commit()
            # #DonneesAAfficher.reinit() # on regénère le tableau GUI
            # ArriveeDossards.remove(doss)
            # #calculeTousLesTemps()
            # retour = Erreur(0)
            # print("Dossard " + str(doss)  + " supprimé du passage sur la ligne d'arrivée.")
        # except :
            # message = "Le dossard " + str(doss) + " n'a pas encore passé la ligne d'arrivée et ne peut donc pas être supprimé."
            # print(message)
            # retour = Erreur(441, doss, message) # la suppression d'un dossard dans l'interface peut constituer une correction d'erreur. Elle ne doit pas provoquer elle-même une erreur .
# ##            PasDErreur = True
    # return retour

def delArriveeDossards():
##    if not Parametres["CourseCommencee"] :
    root['ArriveeDossards'].clear()
    ##transaction.commit()
    print("ArriveeDossards effacés")
##    else :
##        print("Course commencée : impossible d'effacer le listing des dossards arrivés.")


def delTousLesTempsDesCoureurs():
    for c in Coureurs.liste() :
        c.setRang(0)
        c.setTemps(0)
    print("Tous les temps des coureurs effacés")

def delTousLesDeparts():
    for key, value in Courses.items() :
        value.reset()
    print("Tous les top départs effacés")

def delArriveeTempss():
##    if not Parametres["CourseCommencee"] :
    root['ArriveeTemps'].clear()
    root['ArriveeTempsAffectes'].clear()
    ##transaction.commit()
    print("ArriveeTemps effacés")
##    else :
##        print("Course commencée : impossible d'effacer le listing des temps d'arrivée.")

def delDossardsEtTemps():
    global ligneTableauGUI
##    if not Parametres["CourseCommencee"] :
    Parametres["positionDansArriveeTemps"] = 0
    Parametres["positionDansArriveeDossards"] = 0
    Parametres["tempsDerniereRecuperationSmartphone"]=0
    Parametres["ligneDerniereRecuperationSmartphone"]=1
    delArriveeDossards()
    delArriveeTempss()
    delTousLesDeparts()
    delTousLesTempsDesCoureurs()
    effacerFichierDonnneesSmartphone()
    effacerFichierDonnneesLocales()
    genereAffichageTV([])# on vide le fichier Affichage.html pour qu'il ne contienne pas de vieilles données de course.
    genereAffichageWWW([])
    root["LignesIgnoreesSmartphone"] = []
    root["LignesIgnoreesLocal"] = []
    ligneTableauGUI = [1,0]
    if os.path.exists("./videos") :
        shutil.rmtree("./videos")
##    else :
##        print("Course commencée : impossible d'effacer le listing des coureurs")

def delCoureurs():
    global ligneTableauGUI
##    if not Parametres["CourseCommencee"] :
    Coureurs.effacerTout()
    CoureursParClasseUpdate()
    print("Coureurs effacés")
    delCourses()
    delDossardsEtTemps()
##    else :
##        print("Course commencée : impossible d'effacer le listing des coureurs")

def delCourses():
##    if not Parametres["CourseCommencee"] :
    dictUIDPrecedents.clear()
    root['Courses'].clear()
    root['Groupements'].clear()
    ##transaction.commit()
    print("Courses et groupements réinitialisés")
##    else :
##        print("Courses commencées : impossible de supprimer les courses en cours.")

def nettoieCoursesManuelles():
    global Courses, Groupements
    """ Supprime les courses qui n'ont aucun coureur inscrit et nettoie les groupements correspondants et réindexe"""
    # recherche des courses avec coureurs mises dans L
    L = []
    for c in Coureurs.liste() :
        if not c.course in L :
            L.append(c.course)
    # création de CoursesNew et GroupementsNew
    newCourses = {}
    newGroupements = []
    transcription = {} # dictionnaire pour traduire les anciens noms en nouveaux noms de course
    i = 1
    for nom in Courses :
        if nom in L : # si il y a des coureurs, on la copie
            nouveauNom = chr(64 + i)
            transcription[nom] = nouveauNom # destiné à garder une trace du changement pour la réindexatoin de tous les coureurs qui suit.
            newCourses[nouveauNom] = Courses[nom]
            newCourses[nouveauNom].setNomGroupement(nouveauNom)
            newGroupements.append(groupementAPartirDeSonNom(nom, nomStandard=True))
            newGroupements[-1].setNomStandard(nouveauNom)
            newGroupements[-1].setListeDesCourses([nouveauNom])
            i+=1
    # réindexation des noms de courses des coureurs
    for c in Coureurs.liste() :
        c.setCourse(transcription[c.course])
    Coureurs.reindexer(transcription)
    # nettoyage avec garbage collector
    Courses = newCourses
    Groupements = newGroupements
    # Coureurs.afficher()
    #print(newCourses, newGroupements, Courses, Groupements)
    return Courses, Groupements        
            

    
def delCourse(categorie) :
    if categorie in Courses :
        c = Courses[categorie]
        ### VERSION AVEC LA PROPRIETE effectif DONT L'UTILITE EST DISCUTABLE : il est RARE de supprimer un coureur importé.
        #effectif = c.effectif
##        if effectif == 0 :
##            print("Suppression de la course", categorie, "devenue inutile.")
##            Courses.pop(categorie)
##        #else :
##            #print("Categorie", categorie, "conservée.")
##            #c.decremente()
##            #transaction.commit()
##      ANCIENNE VERSION (restaurée) QUAND L'EFFECTIF N'ETAIT PAS UNE PROPRIETE DE l'objet Course
        i=0
        stop = False
        L = Coureurs.liste()
        while i < Coureurs.nombreDeCoureurs and not stop :
            coureur = L[i]
            if coureur.categorie(Parametres["CategorieDAge"]) == categorie :
                stop = True
            i += 1
        if stop :
            print("Categorie", categorie, "conservée.")
        else :
            print("Suppression de la course", categorie, "devenue inutile.")
            Courses.pop(categorie)
            ### ICI, actualisation des groupements en supprimant les groupements qui n'ont que cette course.
            for g in Groupements :
                if categorie in g.listeDesCourses :
                    g.listeDesCourses.remove(categorie)
                    if not g.listeDesCourses : # si le groupement est vide, on le vire
                        Groupements.remove(g)
                    break # une course ne peut être que dans un groupement. Poursuite du parcours inutile.
            ##transaction.commit()
    else :
        print("La course" , categorie, "n'existe pas. NE DEVRAIT PAS SURVENIR.")

##def actualiseCourses():
##    i = 0
##    while i < len(Coureurs) :
##        coureur = Coureurs[i]
##        creerCourse(coureur.categorie(CategorieDAge))
##        i += 1

##def start_server(path, port=8888):
##    '''Start a simple webserver serving path on port'''
##    server_address = ("", port)
##    server = http.server.HTTPServer
##    handler = http.server.CGIHTTPRequestHandler
##    handler.cgi_directories = path
##    print("Serveur actif sur le port :", port)
##    httpd = server(server_address, handler)
##    httpd.serve_forever()


def start_server(path, port=8888):
    '''Start a simple webserver serving path on port'''
    PORT = 8888
    server_address = (path, PORT)
##    httpd = HTTPServer(('', port), CGIHTTPRequestHandler)
##    httpd.serve_forever()
    server = HTTPServer
    handler = CGIHTTPRequestHandler
    handler.cgi_directories = ["/cgi"]
    print("Serveur actif sur le port :", port)
    httpd = server(server_address, handler)
    httpd.serve_forever()

# Start the server in a new thread
port = 8888
#start_server("/",8888)
daemon = threading.Thread(name='daemon_server', target=start_server, args=('', port))
daemon.setDaemon(True) # Set as a daemon so it will be killed once the main thread is dead.
daemon.start()
#time.sleep(1)

def simulateArriveesAleatoires():
    #listeDeCourses = [ '6-F', "6-G", '5-F', "5-G", '4-F', "4-G", '3-F', "3-G" ]
    #topDepart(listeDeCourses)
    addArriveeTemps(time.time(), time.time(),time.time())
    addArriveeTemps(time.time(), time.time(), time.time())
    nbreCoureursSimules = 11
    for n in range(nbreCoureursSimules) :
        time.sleep(random.randint(1,3)/10)
        addArriveeTemps(time.time(), time.time(),time.time())
        addArriveeDossard(random.randint(1,nbreCoureursSimules))


def listerDonneesTerminal():
    #print("Coureurs")
    #listCoureurs()
    print("Courses")
    for cat in listCourses():
        c = Courses[cat]
        print(c.label, "(",c.categorie,") :", c.temps, "(", c.distance,"km)")
    print("ArriveeDossards")
    listArriveeDossards()
    print("ArriveeTemps")
    listArriveeTemps()
##    print("TableauGUI")
##    print(DonneesAAfficher.lignes)
##    print(DonneesAAfficher.lignesActualisees)
##    print(DonneesAAfficher.tempsReferences)


############## AFFICHAGE TV HTML ############################

def estGroupement(obj):
    return isinstance(obj,Groupement)

def estNomDeGroupement(nom):
    return nom in listNomsGroupements()

def estChallenge(obj):
    return (isinstance(obj,str) and obj in listChallenges())

def genereAffichageWWW(listeDesGroupements) :
    """Génère toutes les pages html utiles pour l'affichage dynamique en temps réel depuis internet
    Retourne la liste des fichiers générés.
    """
    retour = []
    with open("modeles/index-en-ligne.html","r", encoding='utf8') as f:
        contenu = f.read()
    f.close()
    ## modèle d'onglet
    ongletModele = """
    <div id=tab@@indicePartantDe1@@ > <a href="#tab@@indicePartantDe1@@">@@groupement@@</a>
	  <div>
		  <h2> @@groupementTitre@@ <span id='chronotime@@indicePartantDe0@@'></span></h2>
		  <div id="conteneurGlobal@@indicePartantDe0@@" >
		  </div>
	  </div>
     </div>
    """
    ## à remettre dans onglet modèle, à côté du titre du groupement : 
    ## affichage tab modèle
    tabModele = """
    <html><head></head><body>
        @@tableauCourse@@
    <div id="testCharge@@indicePartantDe0@@"></div>
    <script>
    chronometres[@@indicePartantDe0@@] = @@heureDepartGroupement@@ ;
    </script>
    </body></html>
    """
    ## création des contenus à partir des données de courses.
    onglets = ""
    heuresDeparts = []
    timerID = []
    dureesActualisation = []
    i = 0
    for groupement in listeDesGroupements :
        chrono = not yATIlUCoureurArrive(groupement.nomStandard)
        onglet = ongletModele.replace("@@indicePartantDe1@@",str(i+1)).replace("@@indicePartantDe0@@",str(i))
        # print(groupement.nomStandard)
        groupementNomStandard = groupement.nomStandard
        if estChallenge(groupement) :
            #print("C'est un challenge par niveau")
            if Parametres["CategorieDAge"] == 2 :
                groupementTitre = "Challenge entre les établissements : catégorie " + groupement.nom + "."
            else :
                groupementTitre = "Challenge entre les classes : niveau " + groupement.nom + "ème."
        else :
            groupementTitre = "Course " + groupement.nom
            if not chrono :
                groupementTitre += " <span id='chronotime'></span>"
        onglet = onglet.replace("@@groupement@@",groupement.nom).replace("@@groupementTitre@@", groupementTitre)
        onglets += onglet
        hdep = genereHeureDepartHTML(groupementNomStandard)
        heuresDeparts.append(hdep)
        timerID.append(0)
        dureesActualisation.append(10000) # actualisation par défaut de 10 secondes. Varie ensuite selon le contexte.
        # création du fichier lié à l'onglet 
        tableauComplet = genereEnTetesHTML(groupementNomStandard, chrono, avecFermetureTABLE=False) + genereTableauHTML(groupementNomStandard, chrono, avecOuvertureTABLE=False)
        tabActuel = tabModele.replace("@@heureDepartGroupement@@",str(hdep)).replace("@@indicePartantDe0@@",str(i))\
            .replace("@@tableauCourse@@", tableauComplet)
        fichierTabActuel = "Affichage-tab" + str(i) + ".html"
        with open(fichierTabActuel,"w", encoding='utf8') as f :
            f.write(tabActuel)
        f.close()
        retour.append(fichierTabActuel)
        i += 1
    ### remplacement des données variables dans le modèle HTML (à partir de la BDD Parametres et des données de course).
    contenu = contenu.replace("@@onglets@@",onglets).replace("@@dureesActualisation@@", str(dureesActualisation))\
              .replace("@@heuresDeparts@@",str(heuresDeparts)).replace("@@timerID@@",str(timerID))
    # TableauxHTML = []
    # EnTetesHTML = []
    # TitresHTML = []
    # heuresDeparts = []
    # for groupement in listeDesGroupements :
    #     #print(groupement)
    #     if yATIlUCoureurArrive(groupement) :
    #         chrono = False
    #     else :
    #         chrono = True
    #     if estChallenge(groupement) :
    #         #print("C'est un challenge par niveau")
    #         if Parametres["CategorieDAge"] == 2 :
    #             TitresHTML.append( "<h2> Challenge entre les établissements : catégorie " + groupement + ". </h2><span id='chronotime'></span>" )
    #         else :
    #             TitresHTML.append( "<h2> Challenge entre les classes : niveau " + groupement + "ème. </h2><span id='chronotime'></span>" )
    #     else :
    #         nomGroupementAffiche = groupementAPartirDeSonNom(groupement, nomStandard=True).nom
    #         if chrono :
    #             TitresHTML.append( "<h2> Catégorie " + nomGroupementAffiche + "</h2>" )
    #         else :
    #             TitresHTML.append( "<h2> Catégorie " + nomGroupementAffiche + " ( <span id='chronotime'></span> )</h2>" )
    #     TableauxHTML.append(genereTableauHTML(groupement, chrono))
    #     EnTetesHTML.append(genereEnTetesHTML(groupement, chrono))
    #     heuresDeparts.append(genereHeureDepartHTML(groupement))
    # ### remplacement des données variables dans le modèle HTML (à partir de la BDD Parametres et des données de course).
    # contenu = contenu.replace("@@tempsPause@@",str(Parametres["tempsPause"]))\
    #           .replace("@@heuresDeparts@@",str(heuresDeparts))
    fichierIndex = "index-en-ligne.html"
    with open(fichierIndex,"w", encoding='utf8') as f :
        f.write(contenu)
    f.close()
    retour.append(fichierIndex)
    return retour

def genereAffichageTV(listeDesGroupements) :
    with open("modeles/Affichage-Contenu.html","r", encoding='utf8') as f:
        contenu = f.read()
    f.close()
    ## création des contenus à partir des données de courses.
    TableauxHTML = []
    EnTetesHTML = []
    TitresHTML = []
    heuresDeparts = []
    for groupement in listeDesGroupements :
        #print(groupement)
        if yATIlUCoureurArrive(groupement) :
            chrono = False
        else :
            chrono = True
        if estChallenge(groupement) :
            #print("C'est un challenge par niveau")
            if Parametres["CategorieDAge"] == 2 :
                TitresHTML.append( "<h2> Challenge entre les établissements : catégorie " + groupement + ". </h2><span id='chronotime'></span>" )
            else :
                TitresHTML.append( "<h2> Challenge entre les classes : niveau " + groupement + "ème. </h2><span id='chronotime'></span>" )
        else :
            nomGroupementAffiche = groupementAPartirDeSonNom(groupement, nomStandard=True).nom
            if chrono :
                TitresHTML.append( "<h2> Catégorie " + nomGroupementAffiche + "</h2>" )
            else :
                TitresHTML.append( "<h2> Catégorie " + nomGroupementAffiche + " ( <span id='chronotime'></span> )</h2>" )
        TableauxHTML.append(genereTableauHTML(groupement, chrono))
        EnTetesHTML.append(genereEnTetesHTML(groupement, chrono))
        heuresDeparts.append(genereHeureDepartHTML(groupement))
    ### remplacement des données variables dans le modèle HTML (à partir de la BDD Parametres et des données de course).
    contenu = contenu.replace("@@tempsPause@@",str(Parametres["tempsPause"]))\
              .replace("@@TitresHTML@@",str(TitresHTML))\
              .replace("@@EnTetesHTML@@",str(EnTetesHTML))\
              .replace("@@TableauxHTML@@",str(TableauxHTML))\
              .replace("@@heuresDeparts@@",str(heuresDeparts))\
              .replace("@@vitesseDefilement@@",str(Parametres["vitesseDefilement"]))

    with open("Affichage-Contenu.html","w", encoding='utf8') as f :
        f.write(contenu)
    f.close()


def genereHeureDepartHTML(groupement) :
    if estChallenge(groupement) :
        retour = [1,0,0,0,0,0] # les challenges n'ont pas d'heure de départ
    else :
        c = Courses[groupementAPartirDeSonNom(groupement, nomStandard = True).listeDesCourses[0]]
        #print("TEST HTML :",c.label, c.temps)
        if c.temps :
            retour = c.departFormate(affichageHTML=True) # le groupement a été lancé. On récupère son heure.
        else :
            retour = [0,0,0,0,0,0] # le groupement n'a pas commencé.
    return retour


def genereEnTetesHTML(groupement, chrono=False, avecFermetureTABLE = True) :
    if estChallenge(groupement) :
        tableau = "<table border='1' cellpadding='6' cellspacing='5' id='titres'><tbody>"
        tableau += '<thead> <tr><th class="rangC"> Classement</th>'
        if Parametres["CategorieDAge"] == 2 :
            tableau += '<th class="etabC">Etablissement</th>'
        else :
            tableau += '<th class="classeC">Classe</th>'
        tableau += '<th class="detailCTitre">Détail : <i>  … + Nom Prénom (rang à l\'arrivée) + ... </i></th>'
        tableau += '<th class="totalC">Total</th>'
        #tableau += '<th class="moyC"><div class=moyC> Moy. des temps des premiers de chaque catégorie. </div></th>'
        tableau += '</tr></thead>'
    else :
        if chrono :
            tableau = "<table border='1' cellpadding='6' cellspacing='5' id='titres'><tbody>"
            tableau += '<thead> <tr><th class="chronometre"> Chronomètre actuel</th> </tr></thead>'
        else :
            tableau = "<table border='1' cellpadding='6' cellspacing='5' id='titres'><tbody>"
            tableau += '<thead> <tr><th class="rang"> RANG</th> <th class="nomprenom">Prénom NOM</th>'
            if CategorieDAge == 0 :
                tableau += '<th class="classe">Classe</th>'
            else :
                if CategorieDAge == 2 :
                    tableau += '<th class="etab">Etablissement</th>'
                else :
                    tableau += '<th class="classe">Catégorie</th>'
            tableau += '<th class="chrono">TEMPS</th><th class="vitesse">VITESSE</th> </tr></thead>'
    if avecFermetureTABLE :
        tableau += '</table>'
    return tableau

def genereTableauHTML(courseName, chrono = False, avecOuvertureTABLE = True) :
    tableau = ""
    if avecOuvertureTABLE :
        tableau = "<table border='1' cellpadding='6' cellspacing='5' id='resultats' style='overflow:hidden;table-layout:fixed;'>"
    tableau += "<tbody>"
    #titre = "Catégorie " + Courses[courseName].label
    if estChallenge(courseName) :
        if courseName in ResultatsGroupements.keys() : # on sécurise si le challenge est vide.
            # challenge par classe
            i = 0
            print("Affichage du challenge sur la TV :",courseName)
            while i < len(ResultatsGroupements[courseName]) :
                #moy = Resultats[courseName][i].moyenneTemps
                score = ResultatsGroupements[courseName][i].score
                #print("Challenge",Resultats[courseName][i],Resultats[courseName][i].nom)
                classe = ResultatsGroupements[courseName][i].nom
                liste = ResultatsGroupements[courseName][i].listeCF + ResultatsGroupements[courseName][i].listeCG
                tableau += "<tr><td class='rangC'>"+ str(i+1) +"</td>"
                if Parametres["CategorieDAge"] == 2 :
                    tableau += "<td class='etabC'>" + classe[3:] # pour l'UNSS, afin de regrouper les coureurs d'un même établissement pour une catégorie, le nom de la "classe" est préfixé : on enlève ce préfixe à l'affichage.
                else :
                    tableau += "<td class='classeC'>" + classe
                tableau += "</td><td class='detailC'><p>" # + '<div class="detailCdiv">'
                tableau += listeNPremiersGF(ResultatsGroupements[courseName][i], htmlRetourLigne=True) #+ ", " + listeNPremiers(ResultatsGroupements[courseName][i].listeCG) 
                tableau += '</p></td>'
                tableau += "<td class='totalC'>"+str(ResultatsGroupements[courseName][i].scoreFormate()) +"</td>"
                #tableau += "<td class='moyC'>" + moy +"</td>"
                tableau += "</tr>"
                i += 1
        else :
            print("Impossible d'afficher le challenge vide sur la TV :",courseName)
    else :
        if chrono :
            tableau += "<tr><td class='chronometre'> <h1><span id='chronotime'></span></h1></td></tr>"
        else :
            Dossards = ResultatsGroupements[courseName]
            #print("ArriveeDossards",ArriveeDossards)
            for dossard in Dossards :
                if formateDossardNG(dossard) in ArriveeDossards : ### INUTILE ? puisque le dossard est dans Resultats, c'est qu'il est arrivé non ? 
                    tableau += genereLigneTableauHTML(formateDossardNG(dossard))
    return tableau + "</tbody> </table>"

def yATIlUCoureurArrive(groupement) :
    retour = False
    # print(ResultatsGroupements)
    try :
        # Fonctionne même si ResultatsGroupements contient les dossards des absents, dispensés et abandons... D'où la raison du parcours de ArriveeDossards.
        Dossards = ResultatsGroupements[groupement]
        for dossard in Dossards :
            if formateDossardNG(dossard) in ArriveeDossards :
                retour = True
                break
    except :
        print("ResultatsGroupements[",groupement,"] n'existe pas. Ignoré")
    return retour

############ LATEX #######################

def creerFichierChallenge(challenge, entete):
    if challenge == "LG" :
        challengeNomAffiche = "lycées généraux et technologiques"
    elif challenge == "LP" :
        challengeNomAffiche = "lycées professionnels"
    elif challenge == "MI" :
        challengeNomAffiche = "minimes"
    elif challenge == "BE" :
        challengeNomAffiche = "benjamin(e)s"
        
    titre = "{\\Large {} \\hfill Challenge " + challengeNomAffiche
    if Parametres["CategorieDAge"] == 0 :
        titre += "ème"
    titre += " \\hfill {}}"
    if Parametres["CategorieDAge"] == 2 :
        chaineSub = "Etabl."
    else :
        chaineSub = "Classe"
    tableau = """
\\begin{center}
\\begin{longtable}{| p{2cm} | p{2cm} | p{18cm} | p{2cm} |}
\\hline
{}\\hfill \\textbf{Rang} \\hfill {} & {} \\hfill \\textbf{@classe@} \\hfill {} & {}\\hfill \\textbf{Détail :} \ldots Prénom Nom (rang à l'arrivée) \ldots \\hfill {} & {}\\hfill \\textbf{Total} \\hfill{} \\\\
\\hline
\\endhead""".replace("@classe@",chaineSub)
    i = 0
    while i < len(ResultatsGroupements[challenge]) :
        #moy = ResultatsGroupements[challenge][i].moyenneTemps
        score = ResultatsGroupements[challenge][i].score
        classe = ResultatsGroupements[challenge][i].nom
        if CategorieDAge == 2 : ## pour l'UNSS , on ajoute le numéro d'établissement
            L = ResultatsGroupements[challenge][i].listeCG + ResultatsGroupements[challenge][i].listeCF
            #print("etablissementNoUNSS",L)
            #print(L[0].etablissementNoUNSS)
            classe += " ("+ L[0].etablissementNoUNSS +") "
        if Parametres["CategorieDAge"] == 2 :
            classe = classe[3:]
        #liste = ResultatsGroupements[challenge][i].listeCF + ResultatsGroupements[challenge][i].listeCG
        tableau += "{} \\hfill {} "+ str(i+1) +"{} \\hfill {}  &{}\\vspace{-2em}\\begin{center} "+ classe +"\\end{center}&  "
        tableau += '\\begin{minipage}{\\linewidth} \\medskip \n {} \\begin{center} '# + listeNPremiers(ResultatsGroupements[challenge][i].listeCF) + ", "
        #tableau += ' {} \\hfill {} \\\\ \n \n' + ' {} \\hfill {} ' + 
        tableau += listeNPremiersGF(ResultatsGroupements[challenge][i]) # listeNPremiers(ResultatsGroupements[challenge][i].listeCG)
        tableau += ' \\end{center} \\\\ \n \\end{minipage} \n & '
        tableau += "{} \\hfill {} "+str(ResultatsGroupements[challenge][i].scoreFormate()) +"{} \\hfill {} \\\\ \n"
        #tableau += "<td class='moyC'>" + moy +"</td>"
        tableau += "\\hline\n"
        i += 1
    return entete + "\n\n" + titre + "\n\n" + tableau

# def creerFichierCategorie(cat, entete):
    # titre = "{\\Large {} \\hfill \\textbf{CATEGORIE " + cat + "} \\hfill {}}"
    # tableau = """
# \\begin{center}
# \\begin{longtable}{| p{1.7cm} | p{6cm} | p{1.5cm} | p{4cm} | p{4.3cm} |}
# %\\begin{tabular}{|*{5}{c|}}
# \\hline
# {}\\hfill \\textbf{Rang} \\hfill {} & {} \\hfill \\textbf{Prénom Nom} \\hfill {} & {}\\hfill \\textbf{Classe} \\hfill {} & {}\\hfill \\textbf{Temps} \\hfill{} & {}\\hfill \\textbf{Vitesse} \\hfill {}\n\\\\
# \\hline
# \\endhead
# \n"""
    # Dossards = Resultats[cat]
    # for dossard in Dossards :
        # if dossard in ArriveeDossards :
            # tableau += genereLigneTableauTEX(dossard)
    # return entete + "\n\n" + titre + "\n\n" + tableau


def creerFichierCategories(groupement, entete):
    titre = "{\\Large {} \\hfill \\textbf{Course " + groupement.nom + "} \\hfill {}}"
    tableau = """
\\begin{center}
\\begin{longtable}{| p{1.7cm} | p{6cm} | p{1.5cm} | p{4cm} | p{4.3cm} |}
%\\begin{tabular}{|*{5}{c|}}
\\hline
{}\\hfill \\textbf{Rang} \\hfill {} & {} \\hfill \\textbf{Prénom Nom} \\hfill {} & {}\\hfill \\textbf{Classe} \\hfill {} & {}\\hfill \\textbf{Temps} \\hfill{} & {}\\hfill \\textbf{Vitesse} \\hfill {}\n\\\\
\\hline
\\endhead
\n"""
    Dossards = ResultatsGroupements[groupement.nomStandard]
    for dossard in Dossards :
        if dossard in ArriveeDossards :
            tableau += genereLigneTableauTEX(formateDossardNG(dossard))
    return entete + "\n\n" + titre + "\n\n" + tableau


def creerFichierClasse(nom, entete, estGroupement):
    titre = "{\\Large {} \\hfill \\textbf{@nom@} \\hfill {}}"
    colonneSuppl = ""
    titreSuppl = ""
    if CategorieDAge == 1 : # on affiche le sexe pour toutes les courses hors scolaire.
        colonneSuppl = "p{1.2cm} |"
        titreSuppl = "{} \\hfill \\textbf{Sexe} \\hfill {} &"
    tableau = "\\begin{center}\n\
\\begin{longtable}{| p{6cm} | " + colonneSuppl + " p{3cm} | p{3.2cm} | p{4.3cm} |}\
\\hline\
 {} \\hfill \\textbf{Nom Prénom } \\hfill {} & " + titreSuppl + " {}\\hfill \\textbf{Rang} \\hfill {} & {}\\hfill \\textbf{Temps} \\hfill{} & \
 {}\\hfill \\textbf{Vitesse} \\hfill {}\n\\\\  \
\\hline \
\\endhead \
"
    ### il faut tous les dossards d'une classe ou cétagorie ou groupement et non seulement ceux arrivés : Dossards = Resultats[classe]
    #print(nom, estGroupement)
    if estGroupement : #estNomDeGroupement(nom) :
        denomination = "Catégorie " + groupementAPartirDeSonNom(nom, nomStandard = True).nom
        Dossards = triParTemps(ResultatsGroupementsPourImpressions[nom])
        #print("Dossards du groupement :",Dossards)
        rangCourse = False
        if Parametres["CategorieDAge"] == 2 :
            garderAbandons = True
            garderAbsDispAbandons = True
        elif Parametres["CategorieDAge"] == 1 :
            garderAbandons = True
            garderAbsDispAbandons = False
        else :
            garderAbandons = False
            garderAbsDispAbandons = False
    else :
        if Parametres["CategorieDAge"] == 2 :
            garderAbandons = True
            garderAbsDispAbandons = True
            rangCourse = True
            denomination = "Catégorie " + nom
            Dossards = triParTemps(ResultatsPourImpressions[nom])### listDossardsDUneCategorie(nom))
        elif Parametres["CategorieDAge"] == 1 :
            garderAbandons = True
            garderAbsDispAbandons = False
            rangCourse = True
            denomination = "Catégorie " + nom
            Dossards = triParTemps(ResultatsPourImpressions[nom])### listDossardsDUneCategorie(nom))
        else :
            garderAbandons = True
            garderAbsDispAbandons = True
            rangCourse = False # pour une classe, on affiche le rang pour le cross du collège en général
            denomination = "Classe " + nom
            Dossards = triParNomPrenom(Resultats[nom]) # on trie par ordre alphabétique pour éviter le cas d'un import en plusieurs fois. listDossardsDUneClasse(nom)
            # les classes ne sont pas triées par temps car c'est plus pratique de garder l'ordre alpha et tous les abs, disp, abandons pour les collègues d'EPS
        #print("Dossards de l'établissement :",Dossards)
    #VMApresente = yATIlUneVMA(Dossards)
    ArrDispAbsAband = [0,0,0,0,0,0,0,0,[]] # le dernier élément contient tous les temps de la classe pour établir moyenne et médiane en bout de calcul
    for dossard in Dossards :
        if Coureurs.recuperer(dossard).temps >= 0 :
            newline, ArrDispAbsAband = genereLigneTableauTEXclasse(dossard, ArrDispAbsAband, rangCourse)
            if Coureurs.recuperer(dossard).temps > 0 or garderAbsDispAbandons or \
               (not Coureurs.recuperer(dossard).absent and not Coureurs.recuperer(dossard).dispense and garderAbandons) :
            # si tps >0 (a couru) OU on garde tout le monde OU si pas absent ni disp et que l'on garde les abandons, on le prend.
                tableau += newline
    return entete + "\n\n" + titre.replace("@nom@",denomination) + "\n\n" + tableau, ArrDispAbsAband


def yATIlUneVMA(listeDeDossards) :
    i = 0
    pasTrouveDeVMA = True
    while i < len(listeDeDossards) and pasTrouveDeVMA :
        coureur = Coureurs.recuperer(listeDeDossards[i])
        if coureur.VMA :
            pasTrouveDeVMA = False
        i += 1
    return not pasTrouveDeVMA


def genereLigneTableauTEXclasse(dossard, ArrDispAbsAbandon, rangCourse=False) :
    # le deuxième argument sera retourné imcrémenté : il représente le nombre d'Arrivées, Dispensés, Absents, Abandons rencontrés jusqu'alors.
    coureur = Coureurs.recuperer(dossard)
    if coureur.temps : # si pas de rang, équivalent à temps nul : sur les données initiales, le constructeur n'ajoutait pas la propriété self.temps.
        contenuTemps = coureur.tempsFormate()
        contenuVitesse = coureur.vitesseFormateeAvecVMAtex()# + supplVMA
        if rangCourse :
            contenuRang = str(coureur.rangCat)
        else :
            contenuRang = str(coureur.rang)
        if coureur.sexe == "F" :
            ArrDispAbsAbandon[0] = ArrDispAbsAbandon[0] + 1
        else :
            ArrDispAbsAbandon[1] = ArrDispAbsAbandon[1] + 1
        ArrDispAbsAbandon[8].append(coureur.temps)
    else :
        contenuVitesse = "-"
        contenuRang = "-"
        if coureur.dispense :
            contenuTemps = "Dispensé"
            if coureur.sexe == "F" :
                ArrDispAbsAbandon[2] = ArrDispAbsAbandon[2] + 1
            else :
                ArrDispAbsAbandon[3] = ArrDispAbsAbandon[3] + 1
        elif coureur.absent :
            contenuTemps = "Absent"
            if coureur.sexe == "F" :
                ArrDispAbsAbandon[4] = ArrDispAbsAbandon[4] + 1
            else :
                ArrDispAbsAbandon[5] = ArrDispAbsAbandon[5] + 1
        else :
            contenuTemps = "Abandon"
            if coureur.sexe == "F" :
                ArrDispAbsAbandon[6] = ArrDispAbsAbandon[6] + 1
            else :
                ArrDispAbsAbandon[7] = ArrDispAbsAbandon[7] + 1
    if Parametres["CategorieDAge"] and coureur.rangCat < 4 and coureur.rangCat != coureur.rang : # un coureur est dans les 3 premiers de sa catégorie
        if coureur.rangCat == 1 :
            if coureur.sexe == "G" :
                chEME = "er "
            else :
                chEME = "ère "
        else :
            chEME = "ème "
        contenuRangCat = " (" +str(coureur.rangCat) + chEME + coureur.categorieFFA() + ")"
    else :
        contenuRangCat = ""
    contenuSuppl = ""
    if CategorieDAge == 1 :
        contenuSuppl = "{} \\hfill " + coureur.sexe +" \\hfill {} & "
    ligne = " {} \\hfill " + coureur.prenom.replace("_","-") + " " + coureur.nom.replace("_","-") + "\\hfill {} & " \
    + contenuSuppl \
    + " {} \\hfill " + contenuRang + contenuRangCat +" \\hfill {} &  {} \\hfill "\
    + contenuTemps + " \\hfill {} &  {} \\hfill " + contenuVitesse \
    + " \\hfill {} \\\\\n"
    if CategorieDAge == 2 : #cas du cross UNSS
        ligne += "\n {} \\hfill ("+ coureur.etablissement + ")"   + " \\hfill {} &  & & \\\\\n"
    ligne += "\hline\n"
    return ligne, ArrDispAbsAbandon

def genereLigneTableauTEX(dossard) :
    coureur = Coureurs.recuperer(dossard)
    if coureur.temps : # si pas de rang, équivalent à temps nul : sur les données initiales, le constructeur n'ajoutait pas la propriété self.temps.
        contenuTemps = coureur.tempsFormate()
##        if coureur.VMA and coureur.VMA > coureur.vitesse :
##            supplVMA = " (" + str(int(coureur.vitesse/coureur.VMA*100)) + "\% VMA)"
##        else :
##            supplVMA = ""
        contenuVitesse = coureur.vitesseFormateeAvecVMAtex() #+ supplVMA
        contenuRang = str(coureur.rang)
        if Parametres["CategorieDAge"] and coureur.rangCat < 4 and coureur.rangCat != coureur.rang : # un coureur est dans les 3 premiers de sa catégorie
            contenuRangCat = " (" +str(coureur.rangCat) + " en " + coureur.categorieFFA() + ")"
        else :
            contenuRangCat = ""
        ligne = " {} \\hfill " + contenuRang  + contenuRangCat + " \\hfill {} &  {} \\hfill " + coureur.prenom + " " + coureur.nom +\
                " \\hfill {} &  {} \\hfill "\
            + coureur.classe + " \\hfill {} &  {} \\hfill " + contenuTemps + " \\hfill {} &  {} \\hfill " + contenuVitesse \
            + " \\hfill {} \\\\\n\hline\n"
    else :
        contenuVitesse = "-"
        contenuRang = "-"
        if coureur.dispense :
            contenuTemps = "Dispensé"
        elif coureur.absent :
            contenuTemps = "Absent"
        else :
            contenuTemps = "Abandon"
        ligne = ""
    return ligne
    
def listeNPremiersGF(equipe,htmlRetourLigne=False):
    ''' nouvelle version de listeNPremiers destinée à remplacer listeNPremiers dans tout le code'''
    listeCoureurs = equipe.listeCF + equipe.listeCG
    retour = ""
    i=0
    for coureur in listeCoureurs :
        if htmlRetourLigne and i == 2 :
            retour += "</p><p>"
            i = 0
        #print("coureur", coureur.nom, coureur.prenom)
        retour += coureur.nom + " " + coureur.prenom  + " ("
        if coureur.rang != coureur.scoreUNSS and Parametres["CategorieDAge"] == 2 :
            retour += str(coureur.rang)+ "/" + str(coureur.nbreArriveesGroupement) + "=>" + coureur.scoreUNSSFormate() + "pts"
        else :
            retour += str(coureur.rang)
        #print(coureur.nom, coureur.rang, coureur.scoreUNSS)
        retour += "), "
        i += 1
    return retour[:-2] + "."


# def listeNPremiers(listeCoureurs,htmlRetourLigne=False):
    # #print("liste des dossards premiers : ",listeCoureurs)
    # retour = ""
    # i=0
    # for coureur in listeCoureurs :
        # if i == 2 :
            # retour += "</p><p>"
            # i = 0
        # #print("coureur", coureur.nom, coureur.prenom)
        # retour += coureur.nom + " " + coureur.prenom  + " ("
        # if coureur.rang != coureur.scoreUNSS and Parametres["CategorieDAge"] == 2 :
            # retour += str(coureur.rang)+ "/" + str(coureur.nbreArriveesGroupement) + "=>" + coureur.scoreUNSSFormate() + "pts"
        # else :
            # retour += str(coureur.rang)
        # #print(coureur.nom, coureur.rang, coureur.scoreUNSS)
        # retour += "), "
        # i += 1
    # return retour[:-2]


def genereLigneTableauHTML(dossard) :
    coureur = Coureurs.recuperer(dossard)
    if coureur.sexe=="G" :
        masc = True
    else :
        masc = False
    ligne = "<tr><td class='rang'>"
    if coureur.rang < 4 :
        ligne += ajoutMedailleEnFonctionDuRang(coureur.rang, masculin=masc)
    else :
        ligne += str(coureur.rang) 
    ligne += "</td><td class='nomprenom'>"+ coureur.prenom + " " + coureur.nom
##    if not CategorieDAge :
##        ligne +=ajoutMedailleEnFonctionDuRang(coureur.rang, masculin=masc)
    if CategorieDAge == 2 and coureur.rang != coureur.rangCat :
        #print(coureur.nom, coureur.rangCat)
        medailleMeilleurDeSaCategorie = ajoutMedailleEnFonctionDuRang(coureur.rangCat, masculin=masc)
        if medailleMeilleurDeSaCategorie :
            ligne += " (" + medailleMeilleurDeSaCategorie + "" + coureur.categorieFFA()+")"
    ligne += "</td>"
    if not CategorieDAge :
        ligne += "<td class='classe'>"+coureur.classe + "</td>"
    else:
        if CategorieDAge == 1 :
            ligne += "<td class='classe'>"
            if coureur.rang != coureur.rangCat :
                ligne += ajoutMedailleEnFonctionDuRang(coureur.rangCat)
            ligne += coureur.categorieFFA()
        elif CategorieDAge == 2 :
            ligne += "<td class='etab'>"+coureur.etablissement
            #if coureur.rang < 4 :
            #    ligne += ajoutMedailleEnFonctionDuRang(coureur.rang, masculin=masc)
        ligne += "</td>"
    ligne += "<td class='chrono'>" + coureur.tempsFormate() +"</td><td class='vitesse'>" + coureur.vitesseFormateeAvecVMA() + "</td></tr>"
    return ligne

def ajoutMedailleEnFonctionDuRang(r,masculin=True) :
    ''' on fournit le rang r (int), cela génère le code HTML nécessaire pour l'insertion de l'image'''
    ligne = ""
    if r < 4 :
        if r == 1 :
            if masculin :
                ordinal = "1er"
            else :
                ordinal = "1ère"
            ligne += '<img style="vertical-align:middle" width="40" class="medailles" src="/media/or.webp" alt="(' + ordinal + ')">'
        elif r == 2 :
            ligne += '<img style="vertical-align:middle" width="40"  class="medailles" src="/media/argent.webp" alt="(2ème)">'
        elif r == 3 :
            ligne += '<img style="vertical-align:middle" width="40"  class="medailles" src="/media/bronze.webp" alt="(3ème)">'
    return ligne


#### catégories d'athlétisme

def categorieAthletisme(anneeNaissance, etablissementNature = "") :
    # pas de distinction dans les catégories Masters pour l'instant. Pas utile.
    # Facile à rajouter à l'aide du tableau categories-athletisme-2022.png
    # Toutes les années suivantes se calculeront par décalage par rapport à cette référence
    correspondanceAnneeCategories = [ [1937, "M10" ], [1942, "M9" ], [1947, "M8" ], [1952, "M7" ], [1957, "M6" ], [1962, "M5" ], [1967, "M4" ], [1972, "M3" ], [1977, "M2" ], [1982, "M1" ], [1987, "M0" ], [1999, "SE" ], [2002, "ES" ], [2004, "JU" ], [2006, "CA" ], [2008, "MI" ], [2010, "BE" ], [2012, "PO" ], [2015, "EA" ], [3000, "BB" ]]
    try :
        anneeNaissance = int(anneeNaissance)
        currentDateTime = datetime.datetime.now()
        date = currentDateTime.date()
        year = currentDateTime.year
        if currentDateTime.month > 8 :
            #changement d'année sportive au premier septembre.
            year += 1
        ecart2022 = year - 2022
        anneeCherchee = anneeNaissance - ecart2022
        i = 0
        continuer = True
        while i< len(correspondanceAnneeCategories) and continuer :
            if anneeCherchee <= correspondanceAnneeCategories[i][0] :
                continuer = False
                categorie = correspondanceAnneeCategories[i][1]
            i += 1
        # patch pour les catégories UNSS :  les redoublants courrent dans la catégorie en dessous. Les élèves en avance (en 2nde) courrent avec les lycéens.
        if Parametres["CategorieDAge"] == 2 :
            if etablissementNature == "CLG" and categorie == "CA" : # le cadet a redoublé
                categorie = "MI"
            elif etablissementNature and etablissementNature[0] == "L" and categorie == "MI" : # le minime a sauté une classe.
                categorie = "CA"
        return categorie
    except :
        print("argument fourni incorrect : pas au format nombre entier")
        return ""

#print(categorieAthletisme(2003))

#### Import des données nouvelle génération (post 2022) à tester...
def traitementDesDonneesAImporter(donneesBrutes) :
    ''' données brutes est un tableau (ou itérable) qui contient des lignes constituées des chaines de caractères (sans point virgule) issues d'un CSV ou tableur.
    Crée les coureurs à partir des informations de chacun, si les données indispensables sont présentes.
    Retourne False si certains éléments impératifs ne sont pas présents dans le fichier source'''
    i=0
    #retour= False
    BilanCreationModifErreur = [0,0,0,0] # nbres de [création, modif, erreurs, identiques]
    for row in donneesBrutes:
        if i == 0 :
            informations = [x.lower() for x in row]
            print("Informations disponibles dans le fichier importé",informations) # on accepte "cat" pour les fichiers OPUS UNSS qui contiennent le sexe dans cette information
## les informations valides ou non, suffisantes ou non sont testées lors de chaque import de coureur.
            if "nom" in informations and "prénom" in informations and "sexe" in informations  and \
            ((Parametres["CategorieDAge"] == 0 and "classe" in informations) \
             or (Parametres["CategorieDAge"] == 1 and "naissance" in informations) \
             or (Parametres["CategorieDAge"] == 1 and "course" in informations) \
             or (Parametres["CategorieDAge"] == 2 and "naissance" in informations and "établissement" in informations and \
                 ("établissementtype" in informations or "type" in informations))) :
             # si infos indispensables dans tous les cas
             # 0 - cas du cross du collège. On a besoin uniquement de la classe.
             # 1 - cas de courses organisées en fonction des catégories de la FFA
             # 2 - cas de courses UNSS (organisées en fonction des catégories de la FFA et des établissements)
                #retour = True
                print("Les éléments obligatoires de la documentation de chronoHB sont bien présents par rapport à la configuration choisie :", informations)
            else :
                print("Certains éléments obligatoires semblent manquer dans le fichier fourni (normal pour un cross UNSS où les champs sont non standards chronoHB) :", informations)
                #retour = False
                ### break
        else :
##             if i == 1 :
##                 print("Première ligne du fichier importé:")
##                 print(row)
             retourCreationModifErreur, d = creerCoureur(row, informations)
             #print("retour création :" ,retourCreationModifErreur)
             for i in range(4) : # actualisation de la liste dénombrant les ajouts, modifs, erreurs effectuées globalement.
                 #print(i,retourCreationModifErreur[i])
                 if retourCreationModifErreur[i] :
                    BilanCreationModifErreur[i] += 1
        i+=1
    return BilanCreationModifErreur, d


### Import XLSX
def recupImportNG(fichierSelectionne="") :
    ''' destiné à remplacer l'appel à recupCSVSIECLE(..) quand ce sera possible : ajout du paramètre categorieManuelle'''
    BilanCreationModifErreur = [0,0,0,0]
    if fichierSelectionne != "" and os.path.exists(fichierSelectionne) :
        if fichierSelectionne[-4:].lower() == "xlsx" :
            BilanCreationModifErreur, d = recupXLSX(fichierSelectionne)
        elif fichierSelectionne[-3:].lower() == "csv":
            BilanCreationModifErreur, d = recupCSV(fichierSelectionne)
    #if retour :
    print("IMPORT CSV ou XLSX TERMINE")
    generateListCoureursPourSmartphone()
    CoureursParClasseUpdate()
    print("Liste des coureurs pour smartphone actualisée.")
        # pas utile de créer une sauvegarde alors que rien n'a été modifié suite à l'import : ecrire_sauvegarde(sauvegarde, "-apres-IMPORT-DONNEES")
##    else :
##        print("Pas de fichier correct sélectionné. N'arrivera jamais avec l'interface graphique normalement.")
    return BilanCreationModifErreur, d


def recupXLSX(fichierSelectionne=""):
    ''' traite le fichier xlsx fourni en argument pour l'import des coureurs'''
    #try :
    #print(fichierSelectionne)
    wb_obj = load_workbook(fichierSelectionne)
    sheet = wb_obj.active # lis la feuille active
    donneesBrutes = [] # initialisation
    for row in sheet.iter_rows(max_row=sheet.max_row):
        ligne = []
        for cell in row:
            if cell.value == None :
                ligne.append("")
            else :
                valeur = str(cell.value)
                ### prétraitement pour les dates de naissances, selon les cas déjà rencontrés.
                if len(valeur)> 18 :
                    if valeur[:8] == "datetime" : # cas datetime.datetime(20,08,2008)
                        # cas des dates dans excel
                        #print("Date à importer au format datetime(...)",valeur)
                        valeur = time.strftime("%d/%m/%Y", valeur)
                    else :
                        try :
                            valeurInitiale = valeur
                            valeur = time.strftime("%d/%m/%Y",time.strptime(valeur, "%Y-%m-%d  %H:%M:%S")) #cas 2008-08-20 00:00:00
                            #print("Date à importer au format 2008-08-20 00:00:00",valeurInitiale, "vers",valeur)
                        except :
                            True # on ne fait rien si une date n'est pas reconnue.
                            #print("Probablement pas une date. Valeur conservée :", valeur)
                ligne.append(valeur)

        #print(chaine)
        donneesBrutes.append(ligne)
    ### traitement déporté dans la fonction ci-dessus traitementDesDonneesAImporter
    BilanCreationModifErreur, d = traitementDesDonneesAImporter(donneesBrutes)
    wb_obj.close()
    #except :
    #    print("Erreur : probablement pas un fichier xlsx valide...")
    return BilanCreationModifErreur, d


def recupCSV(fichierSelectionne=""):
    ''' traite le fichier csv (séparateur point virgule) fourni en argument pour l'import des coureurs'''
    #print("fichierSelectionne",fichierSelectionne)
    BilanCreationModifErreur = [0,0,0]
    try :
        with open(fichierSelectionne, encoding='utf-8') as csvfile:
            donneesBrutes = csv.reader(csvfile, delimiter=';')
            #print(donneesBrutes)
            ### traitement déporté dans la fonction ci-dessus traitementDesDonneesAImporter
            BilanCreationModifErreur, d = traitementDesDonneesAImporter(donneesBrutes)
    except :
        print("Erreur : probablement un mauvais encodage...")
    return BilanCreationModifErreur, d


#### Import CSV ancienne génération (avant 2022)

##def recupCSVSIECLE(fichierSelectionne=""):
####    if Parametres["CourseCommencee"] :
####        message = 'Une ou plusieurs courses ont commencé(es).\nNettoyer toutes les données de courses précédentes avant un import SIECLE.'
####        if __name__=="__main__":
####            print(message)
####        else :
####            reponse = showinfo("PAS D'IMPORT SIECLE",message)
####    else :
##    # Dans une première version non interfacée, on selectionnait le csv le plus récent
##    if fichierSelectionne=="" :
##        datePrecedente = 0
##        for file in glob.glob("./*.csv") :
##            if datePrecedente < os.path.getmtime(file) :
##                datePrecedente = os.path.getmtime(file)
##                fichierSelectionne = file
##        print("Le fichier listing le plus récent est :", fichierSelectionne)
##    # on procède à l'import si le fichier existe
##    if fichierSelectionne != "" and os.path.exists(fichierSelectionne) :
##        try :
##            with open(fichierSelectionne, encoding='utf-8') as csvfile:
##                spamreader = csv.reader(csvfile, delimiter=';')
##                ### traitement à déporter dans la fonction ci-dessus traitementDesDonneesAImporter
##                i=0
##                for row in spamreader:
##                    if i == 0 :
##                         informations = [x.lower() for x in row]
##                         print(informations)
##                         if "nom" in informations and "prénom" in informations and\
##                            "classe" in informations and "sexe" in informations :
##                             retour = True
##                         else :
##                             retour = False
##                             break
##                    else :
##                         #print(row, informations)
##                         creerCoureur(row, informations)
##                         #print("ligne :" ,row)
##                    i+=1
##        except :
##            retour = False
##            print("Erreur : probablement un mauvais encodage...")
##    else :
##        print("Pas de fichier CSV trouvé. N'arrivera jamais avec l'interface graphique normalement.")
##        retour = False
##    print("IMPORT CSV SIECLE TERMINE")
##    ecrire_sauvegarde(sauvegarde, "-apres-IMPORT-SIECLE")
##    if retour : # import effectué : on regénère la liste pour l'application smartphone.
##        generateListCoureursPourSmartphone()
##        CoureursParClasseUpdate()
##        print("Liste des coureurs pour smartphone créée.")
##    return retour

# def setDistances():
#     for nom in listCourses() :
#         print("ajout de la distance 1.2 km à", nom)
#         Courses[nom].setDistance(1.2)

def setDistanceToutesCourses(distance):
    for nom in listCourses() :
        setDistance(nom, distance)

def setDistance(nomCourse, distance):
    print("ajout de la distance " + str(distance) + " km à " + nomCourse)
    Courses[nomCourse].setDistance(distance)

def setParam(parametre, valeur) :
    Parametres[parametre] = valeur
    if parametre == "messageDefaut" : # cas particulier où le fichier texte nécessaire au serveur web doit être réactualisé immédiatement.
        with open("messageDefaut.txt", 'w') as f:
            f.write(valeur)
        f.close()

def setParametres() :
    if Parametres["CategorieDAge"] :
        if CoursesManuelles : 
            crossParClasse = "2"
        else :
            crossParClasse = "0"
    else :
        crossParClasse = "1"
    with open("params.txt", 'w') as f:
        f.write(crossParClasse)  # à compléter avec d'autres paramètres si besoin de les envoyer vers le smartphone.
    f.close()

def creerCoureur(listePerso, informations) :
    infos = {}
    i = 0
    while i < len(informations):
        if i < len(listePerso) :
            infos[informations[i].lower()] = listePerso[i]
        i += 1
    #print(infos)
    sexe=""
    clas = ""
    naiss = ""
    etab = ""
    nature = ""
    disp=False
    abse=False
    vma = 0
    comment = ""
    lic = ""
    doss = ""
    email=""
    emailDeux=""
    courseManuelle=""
    if "nom" in informations :
        nom = supprLF(infos["nom"])
    if "prénom" in informations :
        prenom = supprLF(infos["prénom"])
    if "sexe" in informations or "cat" in informations :
        try :
            sexe = supprLF(infos["sexe"])
        except:
            try :
                sexe = supprLF(infos["cat"])[-1].upper()
                ### dans les fichiers OPUS UNSS, le sexe n'est pas indiqué : on le déduit de la dernière lettre de la catégorie
            except :
                sexe = ""
    if sexe : # si spécifié, on transforme M en G
        if sexe == "M" :
            sexe = "G"
    if "dispensé" in informations :
        if infos["dispensé"] != "" :
            disp = True
    if "absent" in informations :
        if infos["absent"] != "" :
            abse = True
    if "classe" in informations :
        try :
            clas = supprLF(infos["classe"])
        except :
            clas = ""
    if "licence" in informations or "n° licence" in informations  :
        try :
            lic = supprLF(infos["licence"])
        except :
            try :
                lic = supprLF(infos["n° licence"])
            except :
                lic = ""
    if "naissance" in informations or "date naiss." in informations :
        try :
            naiss = supprLF(infos["naissance"])
        except :
            try :
                naiss = supprLF(infos["date naiss."])
            except :
                naiss = ""
    if "établissement" in informations or "nom étab." in informations :
        try :
            etab = supprLF(infos["établissement"])
        except :
            try :
                etab = supprLF(infos["nom étab."])
            except :
                etab = ""
    if "établissementtype" in informations or "type" in informations or "type étab." in informations :
        try :
            nature = supprLF(infos["établissementtype"])
        except :
            try :
                nature = supprLF(infos["type"])
            except :
                try :
                    nature = supprLF(infos["type étab."])
                except :
                    nature = ""
    if "email" in informations and emailEstValide(supprLF(infos["email"])):
        email = supprLF(str(infos["email"]))
    if "email2" in informations and emailEstValide(supprLF(infos["email2"])):
        emailDeux = supprLF(str(infos["email2"]))
    ### traitement des emails afin de s'assurer que email soit bien le principal et email2 le secondaire.
    ### que si email est égal à email2 alors, on vide email2.
    ### que si email est vide et email2 n'est pas vide, alors, on transfère email2 dans email.
    if email == "" and emailDeux != "" :
        email = emailDeux
        emailDeux = ""
    if email == emailDeux :
        emailDeux = ""
    #print("nature de " + supprLF(infos["nom"]) + ":" + nature + ".")
    if nature == "COL" :
        nature = "CLG"
    if nature == "LYC" :
        nature = "LG"
    if "vma" in informations :
        try :
            vma = float(infos["vma"].replace(",",".")) # on assure le coup si les VMA sont à virgule.
        except :
            vma = 0
    if "commentairearrivée" in informations :
        comment = supprLF(infos["commentairearrivée"])
    if "course" in informations :
        try :
            courseManuelle = supprLF(infos["course"])
        except :
            courseManuelle = ""
    if "dossard" in informations :
        doss = formateDossardNG(infos["dossard"])
        #print("Commentaire personnalisé :" + comment+ ".")
    # on crée le coureur avec toutes les informations utiles.
    print('addCoureur(',nom, prenom, sexe , 'classe=',supprLF(infos["classe"]), 'naissance=',naiss, 'absent=',abse, 'dispense=',disp, 'commentaireArrivee=',supprLF(comment), 'VMA=',vma, email, emailDeux)
    # try :
    if nom and prenom and (sexe.upper() == "G" or sexe.upper() =="F") : # trois informations essentielles OBLIGATOIRES VALIDES
        # print("test 06102023", type(email), email, type(emailDeux))
        retourCreationModifErreur, d = addCoureur(nom, prenom, sexe , classe=clas, \
                                            naissance=naiss, etablissement = etab, etablissementNature = nature, absent=abse, dispense=disp,\
                                            temps=0, commentaireArrivee=supprLF(comment), VMA=vma, licence=lic, course=courseManuelle, \
                                            dossard=doss, email=str(email), email2=str(emailDeux), CoureursParClasseUpdateActif=False)
        # print("retourCreationModifErreur",retourCreationModifErreur)
    else :
        if not supprLF(infos["nom"]) and not supprLF(infos["prénom"]) :
            # print("Probablement une ligne inutile dans le tableur. Pas de retour ! Le Nom et le Prénom sont vides.
            retourCreationModifErreur, d = [0,0,0,0], "0"
        else :
            print("Une ligne ne contient pas un des éléments indispensable (nom, prénom ou sexe) : nom=",supprLF(infos["nom"]),"; prénom=", supprLF(infos["prénom"]),"; sexe=", sexe)
            retourCreationModifErreur, d = [0,0,1,0] , "0"
    # except :
    #     retourCreationModifErreur, d = [0,0,1,0] , "0"
    #     print("Une ligne ne contient pas un des éléments indispensable (nom, prénom ou sexe).")
    return retourCreationModifErreur, d


def supprLF(ch) :
    # selon la dernière colonne du csv importée, choisie par l'utilisateur, on peut potentiellement avoir un LF dans n'importe quel champ.
    # tous les éléments importés passent donc dans ce filtre.
    return ch.replace("\n","")


##
##
##if __name__=="__main__":
####    # Start the server in a new thread
####    port = 8888
####    daemon = threading.Thread(name='daemon_server', target=start_server, args=('/', port))
####    daemon.setDaemon(True) # Set as a daemon so it will be killed once the main thread is dead.
####    daemon.start()
####    time.sleep(1)
##    while 1:
##        print("'g' to generate resultats; 'i' pour importer les données des smartphones ; 'g2' pour générer le fichier de coureurs pour les smartphones")
##        print("'g3' pour générer les pdfs de dossards , 'g4' pour l'impression des résultats, 'a4' affecte un dossard à un temps existant ; 'I' pour imprimer les dossards (test).")
##        #print("'s2' pour générer des départs et arrivées de coureurs, 'recup' pour importer le csv le plus récent.")
##        print("Press 'c' to calculate runners times ('c0' from beginning)")#'t' pour le top départ des courses existantes,")# 's' to simulate création de coureurs.")
##        print("Press 'a2' to add un dossard arrivé, 'a3' pour insérer un temps sur la ligne d'arrivée; 'd3' pour supprimer un temps associé à un dossard ou non")
##        print("d4 pour dissocier un dossard d'un temps associé ; dist pour affecter une distance à une course ; distall pour affecter une même distance à toutes les courses.")
##        print("'l' pour lister toutes les données ; 'l1' pour les courses ; 'l2' pour les dossards arrivés ; 'l3' pour les temps à l'arrivée ; 'l4' pour les coureurs.")
##        choice=input("'d1' to delete one runner, 'd2' to delete dossard ,'r' to reset all, 'A1' to add an Coureur, 'recup' pour siecle or 'Q' to quit:")
##        choice=choice.lower()
##        if choice=="l":
##            listerDonneesTerminal()
##        elif choice == "l1" :
##            print("Courses")
##            for cat in listCourses():
##                c = Courses[cat]
##                print(c.label, "(",c.categorie,") :", c.temps, "(", c.distance,"km)")
##        elif choice == "l2" :
##            print("ArriveeDossards")
##            listArriveeDossards()
##        elif choice == "l3" :
##            print("ArriveeTemps")
##            listArriveeTemps()
##        elif choice == 'l4' :
##            listCoureurs()
##        elif choice == "recup" :
##            recupCSVSIECLE()
##        elif choice == 'g2':
##            generateListCoureursPourSmartphone()
##        elif choice == 'g3':
##            generateDossards()
####            mon_thread2=Thread(target=generateDossards)
####            mon_thread2.start()
##        elif choice == 'g4' :
##            mon_thread=Thread(target=generateImpressions)
##            mon_thread.start()
##        elif choice =="test":
##            generateImpressions()
##        elif choice == "distall" :
##            saisie = input("Distance en km à affecter à toutes les courses :")
##            try :
##                d = float(saisie)
##                print(d)
##                setDistanceToutesCourses(d)
##            except:
##                print("Distance invalide : le séparateur décimal est un point")
##        elif choice == "dist" :
##            print("liste des courses", listCourses())
##            nom = input("Nom de la course :")
##            if nom in listCourses() :
##                saisie = input("Distance en km à affecter à " + nom + " : ")
##                try :
##                    d = float(saisie)
##                    setDistance( nom, d)
##                except:
##                    print("Distance invalide : le séparateur décimal est un point")
##        elif choice =="g":
##            genereResultatsCoursesEtClasses()
##            for key in Resultats :
##                if len(key) == 1 :
##                    print(key, " :")
##                    i=0
##                    while i < len(Resultats[key]) :
##                        score = Resultats[key][i].score
##                        classe = Resultats[key][i].nom
##                        liste = Resultats[key][i].listeCF + Resultats[key][i].listeCG
##                        print(" -" ,i+1,":", classe, "(score :", score, ") avec les coureurs ", end='')
##                        for element in liste :
##                            print(element.prenom, "-", element.dossard,"-rang:" , element.rang, ",", end='')
##                        i += 1
##                else :
##                    print(key , ":", Resultats[key])
##            genereAffichageTV(listCourses())
####        elif choice=="t":
####            listeDeCourses = listCourses()
####            topDepart(listeDeCourses)
##        elif choice=="c0":
##            print(calculeTousLesTemps(True))
##        elif choice=="c":
##            calculeTousLesTemps()
##        elif choice=="d1":
##            dossard = input("Dossard :")
##            delCoureur(dossard)
##        elif choice=="d2":
##            dossard = input("Dossard :")
##            delArriveeDossard(dossard)
##        elif choice=="d3":
##            #listArriveeTemps()
##            temps = input("Temps Réel à supprimer :")
##            saisie = input("Dossard associé :")
##            if saisie == "" :
##                dossard = 0
##            else :
##                dossard = saisie
##            delArriveeTemps(temps, dossard)
##        elif choice=="d4":
##            #listArriveeTemps()
##            temps = input("Temps Réel dont il faut dissocier le dossard :")
##            #dissocieArriveeTemps(temps)
##            delDossardAffecteArriveeTemps(temps)
##        elif choice=="r":
##            delCoureurs()
##        elif choice == 'a4':
##            temps = float(input("Temps à affecter :"))
##            dossard = input("Dossard associé (ne rien mettre pour effacer le dossard affecté) :")
##            if dossard == "" :
##                delDossardAffecteArriveeTemps(temps)
##            else :
##                affecteDossardArriveeTemps(temps, dossard)
##        elif choice == 'I' :
##            imprimePDF('dossards/0-tousLesDossards.pdf')
##        elif choice == 'a3':
##            temps = input("Temps à ajouter :")
##            saisie = input("Dossard associé :")
##            if saisie == "" :
##                dossard = 0
##            else :
##                dossard = saisie
##            addArriveeTemps(temps, time.time(), time.time(),dossard)
##        elif choice=="a2":
##            dossard = input("Dossard :")
##            listArriveeDossards()
##            dossardPrecedent = input("Dossard précédent :")
##            if dossardPrecedent != "" :
##                addArriveeDossard(dossard, dossardPrecedent)
##            else :
##                addArriveeDossard(dossard)
##        elif choice=="a1":
##            nom=input("nom :")
##            prenom=input("prenom :")
##            sexe=input("sexe :")
##            classe=input("classe :")
##            addCoureur(nom, prenom, sexe, classe)
##
##        elif choice=="q":
##            break
##        elif choice == "i" :
##            print("on traite les données venant des smartphones")
##            traiterDonneesSmartphone()
##        elif choice == "i0" :
##            print("on traite les données venant des smartphones et modifiées localement en réimportant tout.")
##            traiterDonneesSmartphone(True, True)
##            traiterDonneesLocales(True,True)
##        elif choice == "cross2021" :
##            for donnees in [["3-G",1634718699.42883],["3-F",1634717715.5224173],["4-G",1634716606.7038844],["4-F",1634715607.2591505],["M-G",1634716606.7038844],["5-F",1634713685.815892],["5-G",1634714642.7407954],["6-G",1634712769.324046],["6-F",1634711735.989033],["2-F",1634717715.5224173],["A-F",1634717715.5224173],["A-G",1634718699.42883],["B-F",1634715607.2591505]] :
##                fixerDepart(donnees[0],donnees[1])
##            root["LignesIgnoreesSmartphone"] = []
##            root["LignesIgnoreesLocal"] = []
##        elif choice == "teststats" :
##            testTMPStats()
##    ecrire_sauvegarde(sauvegarde)
##    ##transaction.commit()
##    # close database
##    #connection.close()
##    #db.close()
##
if __name__=="__main__":
    print("création du dictionnaire")
    C = DictionnaireDeCoureurs()
    C.effacerTout()
    print("ajout de coureurs")
    #ArriveeDossards = [Coureur("Lacroix","Olivier","G","21/09/1979"), Coureur("Lacroix","Marielle","F","25/09/1979"), Coureur("Lacroix","Marielle","F","25/09/1979"))
    print("ArriveeDossards",ArriveeDossards)
    print("ArriveeTempsAffectes",ArriveeTempsAffectes)
    C.ajouter(Coureur("Lacroix","Olivier","G","21/09/1979"),"A")
    C.ajouter(Coureur("Lacroix","Marielle","F","25/09/1979"),"A")
    C.ajouter(Coureur("Lacroix","Mathieu","G","26/09/1979"),"A")
    C.ajouter(Coureur("Lacroix","Olivier2","G","21/09/1979"),"B")
    Cmath = Coureur("Lacroix","Marielle2","F","25/09/1979")
    C.ajouter(Cmath,"B")
    C.ajouter(Coureur("Lacroix","Mathieu2","G","26/09/1979"),"B")
    C.afficher()
    print(C.recuperer("2B").prenom)
    C.effacer(Cmath)
    C.effacer("1B")
    C.ajouter(Coureur("Lax","Olive","G","21/09/1979"),"A")
    C.ajouter(Coureur("Lax","Olive2","G","21/09/1979"),"A")
    C.ajouter(Coureur("Lax","Olive3","G","21/09/1979"),"A")
    #print(C.liste())
    C.ajouter(Coureur("Lax","Olive4","G","21/09/1979"),"B")
    #C.effacer("1D")
    C.afficher()
    print(C.recuperer("3B").prenom)
    print(C.existe(Coureur("laX","oLIVE","F","23/09/1980")))
    print(C.existe(3))
    print(C.existe(7))

